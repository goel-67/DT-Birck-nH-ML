import os
import io
import json
import hashlib
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend to avoid Windows display issues
import matplotlib.pyplot as plt
import matplotlib.patheffects as pe
from matplotlib.lines import Line2D
from matplotlib.text import Text
import openpyxl
import datetime as dt
import argparse
from datetime import datetime, timezone

from openpyxl.styles import PatternFill
from sklearn.ensemble import RandomForestRegressor, ExtraTreesRegressor
from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.gaussian_process.kernels import ConstantKernel as C, RBF, WhiteKernel
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.model_selection import RandomizedSearchCV, KFold
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error

from dotenv import load_dotenv
load_dotenv()
# Optional low-discrepancy samplers
try:
    from scipy.stats import qmc
    _HAS_QMC = True
except Exception:
    _HAS_QMC = False

# Optional Graph (SharePoint) imports – only used if env creds present
try:
    import requests, msal  # pip install msal requests
    _HAS_MSAL = True
except Exception:
    _HAS_MSAL = False

CODE_VERSION = "v1.3.0"  # bumped

# ==================== CONFIG ====================
DATASET_CSV           = os.getenv("DT_DATASET_CSV", "full_dataset.csv")
ROOT_DIR              = os.getenv("DT_CACHE_ROOT", "dt-cache")
ROLLING_DIR           = os.path.join(ROOT_DIR, "rolling")
SNAPSHOTS_DIR         = os.path.join(ROOT_DIR, "snapshots")
MANIFESTS_DIR         = os.path.join(ROOT_DIR, "manifests")
ITERATIONS_DIR         = os.path.join(ROOT_DIR, "iterations")  # NEW: Iteration-based caching
TODAY                 = dt.datetime.now().strftime("%Y-%m-%d")
SNAP_DIR              = os.path.join(SNAPSHOTS_DIR, TODAY)
PLOTS_DIR             = os.path.join(SNAP_DIR, "plots")

K                     = int(os.getenv("DT_TOP_K", "3"))
SAMPLING_METHOD       = os.getenv("DT_SAMPLING", "sobol")
SAMPLING_SEED         = int(os.getenv("DT_SEED", "42"))

# If you want to silence the Sobol balance warning, set this true to round N_SAMPLES to nearest power-of-2
ROUND_SOBOL_TO_POW2   = os.getenv("DT_SOBOL_POW2", "true").lower() in ("1","true","yes")

N_SAMPLES             = int(os.getenv("DT_NSAMPLES", "200000"))

# ==================== PARETO OPTIMIZATION PARAMETERS ====================
# Weights for the combined score calculation in recipe selection
ALPHA                 = float(os.getenv("DT_ALPHA", "1.0"))    # Weight for exploitation (distance to Pareto front)
BETA                  = float(os.getenv("DT_BETA", "0.25"))     # Weight for exploration (prediction uncertainty)  
GAMMA                 = float(os.getenv("DT_GAMMA", "0.4"))    # Weight for diversity (spread in parameter space)

# Thresholds for Pareto front improvement (as fractions of current values)
RATE_IMPROVEMENT_THRESHOLD  = float(os.getenv("DT_RATE_THRESHOLD", "0.05"))   # 5% improvement in rate
RANGE_IMPROVEMENT_THRESHOLD = float(os.getenv("DT_RANGE_THRESHOLD", "0.05"))  # 5% improvement in range

# Model choice for first two iterations: "rf_both" (Random Forest for both) or "default" (ExtraTrees for rate, RF for range)
MODEL_CHOICE_FIRST_TWO_ITERATIONS = os.getenv("DT_MODEL_CHOICE", "rf_both").strip().lower()  # "rf_both" | "default"

TARGET_RATE_MIN       = float(os.getenv("DT_RATE_MIN", "35"))
TARGET_RATE_MAX       = float(os.getenv("DT_RATE_MAX", "110"))

# NEW: Iteration configuration
POINTS_PER_ITERATION  = int(os.getenv("DT_POINTS_PER_ITERATION", "3"))  # Number of points per iteration

# Ingestion toggle: "auto" or "manual"
INGEST_MODE           = os.getenv("DT_INGEST_MODE", "auto").strip().lower()  # "auto" | "manual"

# SharePoint / Graph envs (optional)
GRAPH_CLIENT_ID       = os.getenv("GRAPH_CLIENT_ID", "").strip()
GRAPH_CLIENT_SECRET   = os.getenv("GRAPH_CLIENT_SECRET", "").strip()
GRAPH_TENANT_ID       = os.getenv("GRAPH_TENANT_ID", "").strip()
GRAPH_TENANT_NAME     = os.getenv("GRAPH_TENANT_NAME", "purdue0").strip()
GRAPH_SITE_NAME       = os.getenv("GRAPH_SITE_NAME", "Birck-nanoHUB-DT").strip()
RECIPES_FILE_PATH     = os.getenv("RECIPES_FILE_PATH", "Experimental Data/Pareto Recipes.xlsx").strip()
LOCAL_RECIPES_XLSX    = os.getenv("DT_RECIPES_XLSX", "").strip()  # fallback local path

FEATURE_RANGES = {
    "Etch_AvgO2Flow":   (10.0, 90.0),
    "Etch_Avgcf4Flow":  (10.0, 90.0),
    "Etch_Avg_Rf1_Pow": ( 0.0,100.0),
    "Etch_Avg_Rf2_Pow": (50.0,700.0),
    "Etch_AvgPres":     ( 1.0,100.0)
}
FEATURES = list(FEATURE_RANGES.keys())

# IMPORTANT: Recipe constraint - Only recipes with rf2 > rf1 are proposed
# This ensures that the second RF power source is always higher than the first

# What columns to expect in the Excel
EXCEL_STATUS_COL      = os.getenv("DT_EXCEL_STATUS_COL", "Status")             # completed/pending
EXCEL_DATE_COL        = os.getenv("DT_EXCEL_DATE_COL", "Date_Completed")
EXCEL_LOT_COL         = os.getenv("DT_EXCEL_LOT_COL", "Lotname")  # Column 11 in Excel (contains lotnames)
EXCEL_INGEST_COL      = os.getenv("DT_EXCEL_INGEST_COL", "Ingestion_status")   # approved / waiting / not_approved / recipe_rejected / empty
EXCEL_PRED_RATE_COL   = os.getenv("DT_EXCEL_PRED_RATE_COL", "Pred_avg_etch_rate")  # Column 6 in Excel (predicted etch rate)
EXCEL_PRED_RANGE_COL  = os.getenv("DT_EXCEL_PRED_RANGE_COL", "Pred_Range")  # Column 7 in Excel (predicted range)
EXCEL_RATE_UNCERTAINTY_COL = os.getenv("DT_EXCEL_RATE_UNCERTAINTY_COL", "Etch_rate_uncertainty")  # Column for etch rate uncertainty
EXCEL_RANGE_UNCERTAINTY_COL = os.getenv("DT_EXCEL_RANGE_UNCERTAINTY_COL", "Range_uncertainty")  # Column for range uncertainty

# ==================== PLOTTING CONSTANTS ====================
HIGHLIGHT_COLORS = ["purple", "orangered", "green"]  # Match reference implementation

# Unicode symbols for iteration markers
CIRCLED_NUMBERS = ["①","②","③","④","⑤","⑥","⑦","⑧","⑨","⑩","⑪","⑫","⑬","⑭","⑮","⑯","⑰","⑱","⑲","⑳"]
SUBSCRIPT_NUMBERS = ["₁","₂","₃","₄","₅","₆","₇","₈","₉"]

# ==================== HIGHLIGHT PRESETS (optional) ====================
HIGHLIGHT_LOTS = [
    "07_30_2025[_367W]",
    "08_05_2025[_384W]",
    "08_05_2025[_594W]"
]

# ==================== IO HELPERS ====================
def _ensure_dirs():
    for d in [ROOT_DIR, ROLLING_DIR, SNAPSHOTS_DIR, MANIFESTS_DIR, SNAP_DIR, PLOTS_DIR]:
        os.makedirs(d, exist_ok=True)

def _atomic_write(path, data_bytes):
    tmp = path + ".__tmp__"
    with open(tmp, "wb") as f:
        f.write(data_bytes)
    os.replace(tmp, path)

def _read_csv(path):
    if not os.path.exists(path):
        return None
    return pd.read_csv(path)

def _append_csv_atomic(path, df):
    header = not os.path.exists(path)
    tmp = path + ".__tmp__"
    mode = "a" if os.path.exists(path) else "w"
    df.to_csv(tmp, index=False, header=header, mode=mode)
    if os.path.exists(path):
        os.remove(path)
    os.replace(tmp, path)

def _hash_bytes(b):
    import hashlib
    return hashlib.sha256(b).hexdigest()

def _hash_df(df, cols=None):
    if cols is not None:
        df = df[cols].copy()
    bio = io.BytesIO()
    df.to_csv(bio, index=False)
    return _hash_bytes(bio.getvalue())

def _manifest_write(obj):
    path = os.path.join(MANIFESTS_DIR, "latest.json")
    _atomic_write(path, json.dumps(obj, indent=2).encode("utf-8"))

def _manifest_read():
    path = os.path.join(MANIFESTS_DIR, "latest.json")
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

# ==================== DATA ====================
def _parse_date_from_lotname(x):
    if isinstance(x, str) and len(x) >= 10 and x[2] == "_" and x[5] == "_":
        m = x[:2]; d = x[3:5]; y = x[6:10]
        try:
            return pd.to_datetime(f"{y}-{m}-{d}")
        except Exception:
            return pd.NaT
    return pd.NaT

def _load_dataset():
    df = pd.read_csv(DATASET_CSV)
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"])
    else:
        if "LOTNAME" in df.columns:
            df["Date"] = df["LOTNAME"].apply(_parse_date_from_lotname)
        else:
            raise ValueError("Dataset needs a Date column or a LOTNAME with MM_DD_YYYY prefix")
    for c in ["AvgEtchRate", "RangeEtchRate"]:
        if c not in df.columns:
            raise ValueError(f"Missing column {c}")
    for f in FEATURES:
        if f not in df.columns:
            raise ValueError(f"Missing feature {f}")
    return df

# ==================== RECIPES EXCEL (Graph or local) ====================
def _read_recipes_excel():
    """
    Returns a dataframe with at least:
        EXCEL_STATUS_COL (completed/pending),
        EXCEL_DATE_COL   (parsed UTC),
        EXCEL_LOT_COL,
        optionally EXCEL_INGEST_COL,
        EXCEL_PRED_RATE_COL, EXCEL_PRED_RANGE_COL.
    """
    # Prefer Graph if creds provided
    if _HAS_MSAL and GRAPH_CLIENT_ID and GRAPH_CLIENT_SECRET and GRAPH_TENANT_ID:
        print("[recipes] Authenticating with Microsoft Graph…")
        authority = f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}"
        scopes = ["https://graph.microsoft.com/.default"]
        app = msal.ConfidentialClientApplication(
            GRAPH_CLIENT_ID, authority=authority, client_credential=GRAPH_CLIENT_SECRET
        )
        token_resp = app.acquire_token_for_client(scopes=scopes)
        if "access_token" not in token_resp:
            raise RuntimeError(f"MSAL token error: {token_resp}")
        token = token_resp["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        print("[recipes] Downloading Recipes Excel from SharePoint…")
        site_url = f"https://graph.microsoft.com/v1.0/sites/{GRAPH_TENANT_NAME}.sharepoint.com:/sites/{GRAPH_SITE_NAME}"
        site_id  = requests.get(site_url, headers=headers).json()["id"]
        drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
        drives = requests.get(drives_url, headers=headers).json()["value"]
        drive_id = next(d["id"] for d in drives if d["name"] == "Documents")
        download_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{RECIPES_FILE_PATH}:/content"
        resp = requests.get(download_url, headers=headers)
        resp.raise_for_status()
        df = pd.read_excel(io.BytesIO(resp.content))
    else:
        if not LOCAL_RECIPES_XLSX:
            print("[recipes] No Graph creds and no DT_RECIPES_XLSX set; skipping Excel features.")
            return None
        print(f"[recipes] Reading local Excel: {LOCAL_RECIPES_XLSX}")
        df = pd.read_excel(LOCAL_RECIPES_XLSX)

    # Normalize
    df.columns = df.columns.str.strip()
    if EXCEL_STATUS_COL in df.columns:
        df["Status_norm"] = df[EXCEL_STATUS_COL].astype(str).str.strip().str.lower()
    if EXCEL_INGEST_COL in df.columns:
        df["Ingest_norm"] = df[EXCEL_INGEST_COL].astype(str).str.strip().str.lower()
    if EXCEL_DATE_COL in df.columns:
        df[EXCEL_DATE_COL] = pd.to_datetime(df[EXCEL_DATE_COL], errors="coerce", utc=True)
    return df

# ==================== MODELS ====================
def _make_extratrees():
    """Create an ExtraTrees regression model for etch rate prediction (iterations 3+)"""
    return Pipeline([("scaler", StandardScaler()),
                     ("regressor", ExtraTreesRegressor(n_estimators=100, random_state=0, n_jobs=-1))])

def _make_rf():
    """Create a RandomForest regression model for range prediction (all iterations)"""
    return Pipeline([("scaler", StandardScaler()),
                     ("regressor", RandomForestRegressor(n_estimators=100, random_state=0, n_jobs=-1))])

def _make_rf_rate():
    """Create a RandomForest regression model for etch rate prediction (iterations 1-2)"""
    return Pipeline([("scaler", StandardScaler()),
                     ("regressor", RandomForestRegressor(n_estimators=100, random_state=0, n_jobs=-1))])

def _make_gpr():
    """Create a Gaussian Process Regression model with optimized kernel"""
    kernel = C(1.0, (1e-3, 1e3)) * RBF(1.0, (1e-2, 1e3)) + WhiteKernel(1.0, noise_level_bounds=(1e-6, 1e2))
    return Pipeline([
        ("scaler", StandardScaler()),
        ("regressor", GaussianProcessRegressor(
            kernel=kernel, alpha=1e-10, normalize_y=True,
            random_state=0, n_restarts_optimizer=5
        ))
    ])

def _tune_gpr(X, y):
    """Tune GPR hyperparameters using RandomizedSearchCV with comprehensive hyperparameter space"""
    print(f"[gpr_tuning] Starting GPR hyperparameter optimization for {len(X)} samples")
    
    base_model = _make_gpr()
    
    # Define comprehensive hyperparameter space for GPR
    param_space = {
        "regressor__kernel": [
            # ConstantKernel * RBF + WhiteKernel combinations
            C(c) * RBF(l) + WhiteKernel(w)
            for c in (0.01, 0.1, 1.0, 10.0, 100.0)  # Constant kernel variance
            for l in (0.01, 0.1, 1.0, 10.0, 100.0)  # RBF length scale
            for w in (1e-6, 1e-5, 1e-4, 1e-3, 1e-2, 1e-1)  # White noise
        ],
        "regressor__alpha": [1e-10, 1e-8, 1e-6, 1e-4, 1e-2],  # Regularization
        "regressor__n_restarts_optimizer": [3, 5, 10]  # Optimization restarts
    }
    
    # Use KFold cross-validation with appropriate number of splits
    n_splits = min(5, max(2, len(y) // 3))  # Ensure at least 2 samples per fold
    cv = KFold(n_splits=n_splits, shuffle=True, random_state=0)
    
    # Calculate number of iterations based on parameter space size
    n_iter = min(50, max(10, len(param_space["regressor__kernel"]) // 2))
    
    print(f"[gpr_tuning] Using {n_splits}-fold CV with {n_iter} iterations")
    
    # Randomized search
    search = RandomizedSearchCV(
        estimator=base_model,
        param_distributions=param_space,
        n_iter=n_iter,
        scoring="neg_root_mean_squared_error",
        cv=cv, n_jobs=-1,
        random_state=0,
        verbose=1, refit=True,
    )
    
    search.fit(X, y)
    
    print(f"[gpr_tuning] Best CV score: {-search.best_score_:.4f}")
    print(f"[gpr_tuning] Best parameters: {search.best_params_}")
    
    return search.best_estimator_, search.best_params_

def _pred_stats(pipeline_model, X_in):
    scaler = pipeline_model.named_steps["scaler"]
    regressor = pipeline_model.named_steps["regressor"]
    Xs = scaler.transform(X_in)
    
    if isinstance(regressor, GaussianProcessRegressor):
        # GPR provides mean and std directly
        mu, std = regressor.predict(Xs, return_std=True)
        return mu, std
    else:
        # Ensemble models (RF, ExtraTrees) - calculate mean and std across trees
        per_tree = np.stack([t.predict(Xs) for t in regressor.estimators_], axis=0)
        return per_tree.mean(axis=0), per_tree.std(axis=0)

def _calculate_uncertainties_for_iteration(iteration_num, recipes_df):
    """
    Calculate uncertainties for recipes in a specific iteration by training a model
    on the data available up to that iteration and predicting uncertainties.
    
    Args:
        iteration_num: The iteration number (1, 2, 3, ...)
        recipes_df: DataFrame containing the recipes Excel data
    
    Returns:
        DataFrame with added uncertainty columns
    """
    print(f"[uncertainty] Calculating uncertainties for iteration {iteration_num}")
    
    # Load the full dataset
    try:
        df = pd.read_csv(DATASET_CSV)
        print(f"[uncertainty] Loaded dataset with {len(df)} rows")
    except Exception as e:
        print(f"[uncertainty] Error loading dataset: {e}")
        return recipes_df
    
    # Get the proper training data for this iteration
    training_data = _get_training_data_for_iteration(df, recipes_df, iteration_num)
    
    if training_data.empty:
        print(f"[uncertainty] Iteration {iteration_num}: No training data available")
        return recipes_df
    
    # Prepare features and targets
    X = training_data[FEATURES].values
    y_rate = training_data['AvgEtchRate'].values
    y_range = training_data['RangeEtchRate'].values
    
    # Train Random Forest models with 100 estimators and random_state=0
    print(f"[uncertainty] Training Random Forest models with 100 estimators...")
    
    # Rate model - use appropriate model based on iteration
    if iteration_num <= 2:
        rate_model = _make_rf_rate()  # Random Forest for iterations 1-2
        print(f"[uncertainty] Iteration {iteration_num}: Using Random Forest for etch rate (100 estimators)")
    else:
        rate_model = _make_extratrees()  # Extra Trees for iterations 3+
        print(f"[uncertainty] Iteration {iteration_num}: Using Extra Trees for etch rate (100 estimators)")
    rate_model.fit(X, y_rate)
    
    # Range model - always use Random Forest
    range_model = _make_rf()  # Random Forest for all iterations
    print(f"[uncertainty] Iteration {iteration_num}: Using Random Forest for range (100 estimators)")
    range_model.fit(X, y_range)
    
    # Get the first 3 recipes from this iteration for uncertainty calculation
    iteration_recipes = recipes_df.head(3)
    
    if len(iteration_recipes) == 0:
        print(f"[uncertainty] No recipes found for iteration {iteration_num}")
        return recipes_df
    
    # Extract features for these recipes
    recipe_features = []
    for _, recipe in iteration_recipes.iterrows():
        # Extract feature values from the recipe
        # You may need to adjust this based on how features are stored in Excel
        feature_values = []
        for feature in FEATURES:
            if feature in recipe.index:
                feature_values.append(recipe[feature])
            else:
                # Use default values if features not found
                feature_values.append(np.mean(training_data[feature]))
        recipe_features.append(feature_values)
    
    recipe_features = np.array(recipe_features)
    
    # Predict uncertainties
    rate_mean, rate_std = _pred_stats(rate_model, recipe_features)
    range_mean, range_std = _pred_stats(range_model, recipe_features)
    
    # Convert range uncertainty to thickness range (multiply by 5)
    range_std_thickness = range_std * 5.0
    
    print(f"[uncertainty] Predicted uncertainties for {len(iteration_recipes)} recipes:")
    for i, (_, recipe) in enumerate(iteration_recipes.iterrows()):
        print(f"[uncertainty] Recipe {i+1}: Rate ±{rate_std[i]:.2f}, Range ±{range_std_thickness[i]:.2f}")
    
    # Add uncertainty columns to recipes DataFrame
    recipes_df_copy = recipes_df.copy()
    
    # Add uncertainty columns if they don't exist
    if 'Pred_etch_rate_uncertainty' not in recipes_df_copy.columns:
        recipes_df_copy['Pred_etch_rate_uncertainty'] = np.nan
    if 'Pred_range_uncertainty' not in recipes_df_copy.columns:
        recipes_df_copy['Pred_range_uncertainty'] = np.nan
    
    # Update uncertainties for the first 3 recipes
    for i in range(min(3, len(recipes_df_copy))):
        recipes_df_copy.loc[i, 'Pred_etch_rate_uncertainty'] = rate_std[i]
        recipes_df_copy.loc[i, 'Pred_range_uncertainty'] = range_std_thickness[i]
    
    print(f"[uncertainty] Added uncertainty columns to recipes DataFrame")
    return recipes_df_copy

# ==================== SAMPLING ====================
def _nearest_pow2(n):
    import math
    if n <= 1:
        return 1
    a = 1 << (n-1).bit_length()
    b = a >> 1
    return a if (a - n) <= (n - b) else b

def _sample_candidates(method, n, lower, upper, seed):
    d = len(lower)
    m = method.lower()
    if m == "sobol":
        # Always ensure Sobol uses powers of 2 for optimal performance
        n = _nearest_pow2(n)
        print(f"[sampling] Sobol sampling: adjusted N_SAMPLES from {n//_nearest_pow2(n)} to {n} (power of 2)")
    elif m == "sobol" and ROUND_SOBOL_TO_POW2:
        n = _nearest_pow2(n)
    if m == "random" or not _HAS_QMC:
        rng = np.random.default_rng(seed)
        u = rng.random((n, d))
        return lower + u * (upper - lower)
    if m == "sobol":
        eng = qmc.Sobol(d=d, scramble=True, seed=seed)
        u = eng.random(n)
        return qmc.scale(u, lower, upper)
    if m == "lhs":
        eng = qmc.LatinHypercube(d=d, seed=seed)
        u = eng.random(n)
        return qmc.scale(u, lower, upper)
    rng = np.random.default_rng(seed)
    u = rng.random((n, d))
    return lower + u * (upper - lower)

def _quantize(X, cols):
    Xq = X.copy()
    for j, col in enumerate(cols):
        kl = col.lower()
        if "flow" in kl or "pres" in kl:
            Xq[:, j] = np.round(Xq[:, j] * 10) / 10.0
        elif "pow" in kl:
            Xq[:, j] = np.round(Xq[:, j])
    return Xq

# ==================== PARETO ====================
def _is_pareto(points):
    P = np.array(points, dtype=float)
    flags = np.ones(len(P), dtype=bool)
    for i, (r, rng) in enumerate(P):
        # Check if this point is dominated by any other point
        for j, (r2, rng2) in enumerate(P):
            if i != j:  # Don't compare point with itself
                if (r2 >= r and rng2 <= rng and (r2 > r or rng2 < rng)):
                    flags[i] = False
                    break
    return flags

def _pareto_front(df_rates_ranges):
    pts = df_rates_ranges[["AvgEtchRate","Range_nm"]].values
    flags = _is_pareto(pts)
    return df_rates_ranges.loc[flags].sort_values("AvgEtchRate").reset_index(drop=True)

# ==================== SCORING ====================
def _norm01(v):
    v = np.asarray(v, float)
    vmin, vmax = float(np.min(v)), float(np.max(v))
    if vmax - vmin < 1e-12:
        return np.zeros_like(v)
    return (v - vmin) / (vmax - vmin)

def _objective_scales(rate_vals, range_nm_vals):
    r_scale  = np.percentile(rate_vals, 97.5) - np.percentile(rate_vals, 2.5)
    rn_scale = np.percentile(range_nm_vals, 97.5) - np.percentile(range_nm_vals, 2.5)
    if r_scale  < 1e-9: r_scale  = max(1.0, np.ptp(rate_vals))
    if rn_scale < 1e-9: rn_scale = max(1.0, np.ptp(range_nm_vals))
    return float(r_scale), float(rn_scale)

def _improves(p, front):
    """
    Check if point p improves the Pareto front.
    
    A point improves the front if it is NOT dominated by any existing front point.
    Point p is dominated by point q if:
    - q[0] >= p[0] (q has better or equal rate) AND q[1] <= p[1] (q has better or equal range)
    - AND at least one inequality is strict (q[0] > p[0] OR q[1] < p[1])
    
    A point can be added to the front even if it doesn't dominate any existing points.
    It just needs to be non-dominated (it can "fill a gap" between existing points).
    
    Args:
        p: tuple (rate, range) - the candidate point
        front: array of shape (n, 2) - existing Pareto front points
        
    Returns:
        bool: True if p improves the front (is non-dominated), False otherwise
    """
    if len(front) == 0:
        return True  # Empty front, any point improves it
    
    # Check if p is dominated by any existing point on the front
    for front_point in front:
        # front_point dominates p if front_point is better or equal in both objectives
        # AND strictly better in at least one objective
        if (front_point[0] >= p[0] and front_point[1] <= p[1] and 
            (front_point[0] > p[0] or front_point[1] < p[1])):
            return False  # p is dominated by this front point
    
    # If we get here, p is not dominated by any front point
    return True

def _dominates_existing(p, front):
    """
    Check if point p dominates ANY existing point on the Pareto front.
    
    Point p dominates point q if:
    - p[0] >= q[0] (p has better or equal rate) AND p[1] <= q[1] (p has better or equal range)
    - AND at least one inequality is strict (p[0] > q[0] OR p[1] < q[1])
    
    Args:
        p: tuple (rate, range) - the candidate point
        front: array of shape (n, 2) - existing Pareto front points
        
    Returns:
        bool: True if p dominates at least one existing point, False otherwise
    """
    if len(front) == 0:
        return False  # Empty front, nothing to dominate
    
    # Check if p dominates any existing point on the front
    for front_point in front:
        # p dominates front_point if p is better or equal in both objectives
        # AND strictly better in at least one objective
        if (p[0] >= front_point[0] and p[1] <= front_point[1] and 
            (p[0] > front_point[0] or p[1] < front_point[1])):
            return True  # p dominates this front point
    
    return False

def _significantly_improves(p, front, rate_threshold=0.05, range_threshold=0.05):
    """
    Check if point p significantly improves the Pareto front.
    
    A point significantly improves the front if it either:
    1. Dominates at least one existing point, OR
    2. Is non-dominated AND provides a significant improvement in at least one objective
       compared to the closest existing point
    
    Args:
        p: tuple (rate, range) - the candidate point
        front: array of shape (n, 2) - existing Pareto front points
        rate_threshold: minimum improvement in rate (as fraction of current best)
        range_threshold: minimum improvement in range (as fraction of current best)
        
    Returns:
        bool: True if p significantly improves the front, False otherwise
    """
    if len(front) == 0:
        return True  # Empty front, any point improves it
    
    # First check if it dominates any existing point
    if _dominates_existing(p, front):
        return True
    
    # If not dominating, check if it's non-dominated and provides significant improvement
    if not _improves(p, front):
        return False  # Point is dominated
    
    # Find the closest point on the front
    min_distance = float('inf')
    closest_point = None
    for front_point in front:
        # Calculate normalized distance
        rate_diff = abs(p[0] - front_point[0]) / max(1, abs(front_point[0]))  # Avoid division by zero
        range_diff = abs(p[1] - front_point[1]) / max(1, abs(front_point[1]))
        distance = np.sqrt(rate_diff**2 + range_diff**2)
        if distance < min_distance:
            min_distance = distance
            closest_point = front_point
    
    if closest_point is None:
        return True  # Shouldn't happen, but safety check
    
    # Check if there's significant improvement in at least one objective
    rate_improvement = (p[0] - closest_point[0]) / max(1, abs(closest_point[0]))
    range_improvement = (closest_point[1] - p[1]) / max(1, abs(closest_point[1]))  # Lower range is better
    
    # Return True if there's significant improvement in at least one objective
    return (rate_improvement > rate_threshold) or (range_improvement > range_threshold)

def _exploit_distance_norm(p, front, r_scale, rn_scale):
    dr = (p[0] - front[:,0]) / r_scale
    dn = (p[1] - front[:,1]) / rn_scale
    return np.min(np.hypot(dr, dn))

def _diversity_score(candidate_knob, selected_knobs, cols):
    if not selected_knobs:
        return 1.0
    def _norm_vec(vec):
        out = []
        for name in cols:
            lo, hi = FEATURE_RANGES[name]
            out.append((vec[cols.index(name)] - lo) / (hi - lo))
        return np.array(out)
    nc = _norm_vec(candidate_knob)
    dmin = float("inf")
    for sk in selected_knobs:
        ns = _norm_vec(sk)
        d = np.linalg.norm(nc - ns)
        if d < dmin:
            dmin = d
    return float(dmin)

# ==================== HASHES ====================
def _model_config_hash():
    payload = json.dumps({
        "K": K,
        "SAMPLING": SAMPLING_METHOD,
        "SEED": SAMPLING_SEED,
        "NSAMPLES": N_SAMPLES,
        "ALPHA": ALPHA,
        "BETA": BETA,
        "GAMMA": GAMMA,
        "models": {"rate": "ExtraTrees+Scaler/200", "range": "RF+Scaler/200"}
    }, sort_keys=True).encode("utf-8")
    return _hash_bytes(payload)

def _features_hash():
    payload = json.dumps({"features": FEATURES, "ranges": FEATURE_RANGES}, sort_keys=True).encode("utf-8")
    return _hash_bytes(payload)

def _code_hash():
    return _hash_bytes(CODE_VERSION.encode("utf-8"))

def _dataset_hash(df):
    cols = FEATURES + ["AvgEtchRate","RangeEtchRate","run_date"]  # Fixed: was "Date", should be "run_date"
    return _hash_df(df[cols])

def _cache_key(dataset_h, feat_h, model_h, code_h):
    return _hash_bytes("|".join([dataset_h, feat_h, model_h, code_h]).encode("utf-8"))

# ==================== BACKTEST APPEND (with optional cutoff) ====================
def _fit_predict_day(df, day, make_rate, make_range):
    """Fit models on data before the given day and predict on that day's data"""
    train = df[df["run_date"] < day]  # Fixed: was "Date", should be "run_date"
    test  = df[df["run_date"] == day]  # Fixed: was "Date", should be "run_date"
    if len(test) == 0 or len(train) == 0:
        return None
    Xtr = train[FEATURES].astype(float).values
    ytr_rate  = train["AvgEtchRate"].values
    ytr_range = train["RangeEtchRate"].values
    Xte = test[FEATURES].astype(float).values
    m_rate  = make_rate().fit(Xtr, ytr_rate)
    m_range = make_range().fit(Xtr, ytr_range)
    mu_rate,  sd_rate  = _pred_stats(m_rate,  Xte)
    mu_range, sd_range = _pred_stats(m_range, Xte)
    out = test[["LOTNAME","FIMAP_FILE","run_date","AvgEtchRate","RangeEtchRate"]].copy()  # Fixed: was "Date", should be "run_date"
    out = out.rename(columns={"run_date":"Date","AvgEtchRate":"y_true_rate","RangeEtchRate":"y_true_range"})  # Rename run_date to Date for output
    out["y_pred_rate"]  = mu_rate
    out["y_pred_range"] = mu_range
    out["y_std_rate"]   = sd_rate
    out["y_std_range"]  = sd_range
    return out

def _update_backtesting(df, cache_key_str, cutoff_day=None):
    """
    Enhanced backtesting that processes all available data to build comprehensive metrics.
    If cutoff_day is provided (manual mode), we only append days <= cutoff_day.
    """
    # Convert run_date column to datetime format for proper comparison
    df = df.copy()
    df["run_date"] = pd.to_datetime(df["run_date"])
    
    path = os.path.join(ROLLING_DIR, "predictions_by_date.csv")
    existing = _read_csv(path)
    if existing is None:
        existing = pd.DataFrame(columns=[
            "LOTNAME","FIMAP_FILE","Date","y_true_rate","y_true_range",
            "y_pred_rate","y_pred_range","y_std_rate","y_std_range",
            "train_end_date","cache_key"
        ])
    if len(existing):
        existing["Date"] = pd.to_datetime(existing["Date"])
        existing["train_end_date"] = pd.to_datetime(existing["train_end_date"])

    # Check if we need to rebuild the entire backtesting dataset
    need_rebuild = False
    if len(existing) == 0:
        need_rebuild = True
    else:
        # Check if cache key matches
        existing_cache_keys = existing["cache_key"].unique()
        if cache_key_str not in existing_cache_keys:
            need_rebuild = True
        else:
            # Check if we have data for all available dates
            existing_dates = set(existing.loc[existing["cache_key"] == cache_key_str, "train_end_date"].dt.date)
            # Convert run_date to datetime first, then extract date
            all_dates = set(pd.to_datetime(df["run_date"].dropna()).dt.date)
            if not existing_dates.issuperset(all_dates):
                need_rebuild = True

    if need_rebuild:
        print(f"[backtesting] Rebuilding comprehensive backtesting dataset for {len(df)} rows")
        # Clear existing data for this cache key
        if len(existing) > 0:
            existing = existing[existing["cache_key"] != cache_key_str]
        
        # Process all available dates
        unique_days = sorted(pd.to_datetime(df["run_date"].dropna()).unique())  # Convert to datetime first
        if cutoff_day is not None:
            unique_days = [d for d in unique_days if d <= cutoff_day]
        
        print(f"[backtesting] Processing {len(unique_days)} unique dates")

        rows = []
        for d in unique_days:
            out = _fit_predict_day(df, d, _make_extratrees, _make_rf)
            if out is None:
                continue
            out["train_end_date"] = d
            out["cache_key"] = cache_key_str
            rows.append(out)
        
        if rows:
            app = pd.concat(rows, ignore_index=True)
            # Combine with existing data from other cache keys
            if len(existing) > 0:
                app = pd.concat([existing, app], ignore_index=True)
            _atomic_write(path, app.to_csv(index=False).encode("utf-8"))
            print(f"[backtesting] Built comprehensive dataset with {len(rows)} date predictions")
        else:
            print("[backtesting] No valid predictions generated")
    else:
        print(f"[backtesting] Using existing backtesting data for cache key {cache_key_str}")
    
    # Return the rolling path first (for metrics building), then the backtest path
    return os.path.join(ROLLING_DIR, "metrics_over_time.csv"), path

def _build_metrics_over_time(backtest_csv_path, cache_key_str):
    print(f"[metrics] Building metrics from: {backtest_csv_path}")
    ex = _read_csv(backtest_csv_path)
    if ex is None or len(ex) == 0:
        print(f"[metrics] No backtest data found at {backtest_csv_path}")
        return None
    print(f"[metrics] Found {len(ex)} backtest records")
    ex = ex[ex["cache_key"] == cache_key_str].copy()
    if len(ex) == 0:
        print(f"[metrics] No records found for cache key {cache_key_str}")
        return None
    print(f"[metrics] {len(ex)} records match cache key {cache_key_str}")
    ex["train_end_date"] = pd.to_datetime(ex["train_end_date"])
    def _agg(g):
        ytr = g["y_true_rate"].values
        ypr = g["y_pred_rate"].values
        ys  = g["y_std_rate"].values
        yrmse = float(np.sqrt(mean_squared_error(ytr, ypr)))
        ymae  = float(mean_absolute_error(ytr, ypr))
        yr2   = float(r2_score(ytr, ypr)) if len(np.unique(ytr))>1 else np.nan
        cov1  = float(np.mean(np.abs(ytr - ypr) <= ys))
        cov2  = float(np.mean(np.abs(ytr - ypr) <= 2*ys))
        ytrg = g["y_true_range"].values
        yprg = g["y_pred_range"].values
        ysg  = g["y_std_range"].values
        grmse = float(np.sqrt(mean_squared_error(ytrg, yprg)))
        gmae  = float(mean_absolute_error(ytrg, yprg))
        gr2   = float(r2_score(ytrg, yprg)) if len(np.unique(ytrg))>1 else np.nan
        gcov1 = float(np.mean(np.abs(ytrg - yprg) <= ysg))
        gcov2 = float(np.mean(np.abs(ytrg - yprg) <= 2*ysg))
        return pd.Series({
            "rmse_rate": yrmse, "mae_rate": ymae, "r2_rate": yr2, "coverage_rate_1s": cov1, "coverage_rate_2s": cov2,
            "rmse_range": grmse, "mae_range": gmae, "r2_range": gr2, "coverage_range_1s": gcov1, "coverage_range_2s": gcov2,
            "n_points_up_to_date": len(g)
        })
    ex = ex.sort_values("train_end_date")
    groups = []
    dates = sorted(ex["train_end_date"].unique())
    for d in dates:
        g = ex[ex["train_end_date"] <= d]
        s = _agg(g)
        s.name = d
        groups.append(s)
    met = pd.DataFrame(groups)
    met.index.name = "train_end_date"
    met = met.reset_index()
    out_path_rolling = os.path.join(ROLLING_DIR, "metrics_over_time.csv")
    _atomic_write(out_path_rolling, met.to_csv(index=False).encode("utf-8"))
    
    # NEW: Create historical snapshots for all significant dates
    _create_historical_snapshots(met, ex, cache_key_str)
    
    # Also save to current snapshot directory
    out_path_snap = os.path.join(SNAP_DIR, "metrics_over_time.csv")
    _atomic_write(out_path_snap, met.to_csv(index=False).encode("utf-8"))
    print(f"[metrics] Built metrics with {len(met)} data points, saved to {out_path_rolling}")
    return out_path_snap

def _create_historical_snapshots(metrics_df, backtest_df, cache_key_str):
    """Create historical snapshots for all significant dates to enable historical analysis"""
    print(f"[snapshots] Creating historical snapshots for {len(metrics_df)} dates")
    
    # Create snapshots for key milestone dates (every 10th date and significant performance changes)
    snapshot_dates = []
    
    # Add every 10th date for regular intervals
    for i in range(0, len(metrics_df), 10):
        snapshot_dates.append(metrics_df.iloc[i]["train_end_date"])
    
    # Add the last date
    if len(metrics_df) > 0:
        snapshot_dates.append(metrics_df.iloc[-1]["train_end_date"])
    
    # Add dates with significant performance changes (large RMSE changes)
    if len(metrics_df) > 1:
        rmse_rate_changes = metrics_df["rmse_rate"].diff().abs()
        significant_changes = rmse_rate_changes > rmse_rate_changes.quantile(0.9)  # Top 10% changes
        significant_dates = metrics_df[significant_changes]["train_end_date"]
        snapshot_dates.extend(significant_dates)
    
    # Remove duplicates and sort
    snapshot_dates = sorted(list(set(snapshot_dates)))
    
    print(f"[snapshots] Creating {len(snapshot_dates)} historical snapshots")
    
    for date in snapshot_dates:
        date_str = date.strftime("%Y-%m-%d")
        snapshot_dir = os.path.join(SNAPSHOTS_DIR, date_str)
        os.makedirs(snapshot_dir, exist_ok=True)
        
        # Save metrics up to this date
        date_metrics = metrics_df[metrics_df["train_end_date"] <= date]
        metrics_path = os.path.join(snapshot_dir, "metrics_over_time.csv")
        _atomic_write(metrics_path, date_metrics.to_csv(index=False).encode("utf-8"))
        
        # Save backtesting data up to this date
        date_backtest = backtest_df[backtest_df["train_end_date"] <= date]
        backtest_path = os.path.join(snapshot_dir, "backtest_data.csv")
        _atomic_write(backtest_path, date_backtest.to_csv(index=False).encode("utf-8"))
        
        # Create plots for this date
        plots_dir = os.path.join(snapshot_dir, "plots")
        os.makedirs(plots_dir, exist_ok=True)
        
        # Generate plots for this historical date
        _generate_historical_plots(date_metrics, date_backtest, plots_dir, date_str)
    
    print(f"[snapshots] Historical snapshots created successfully")

def _generate_historical_plots(metrics_df, backtest_df, plots_dir, date_str):
    """Generate plots for a specific historical date"""
    if len(metrics_df) < 2:
        return  # Need at least 2 points for meaningful plots
    
    # RMSE plot with dual Y-axes
    fig, ax1 = plt.subplots(figsize=(12,7))
    ax2 = ax1.twinx()
    
    # Plot RMSE Rate on left Y-axis (blue, circles)
    line1 = ax1.errorbar(metrics_df["train_end_date"], metrics_df["rmse_rate"], 
                         fmt='o-', color='#1f77b4', label="RMSE Rate", linewidth=2, 
                         markersize=6, capsize=4, alpha=0.8)
    ax1.set_xlabel("Train end date")
    ax1.set_ylabel("RMSE Rate", color='#1f77b4')
    ax1.tick_params(axis='y', labelcolor='#1f77b4')
    
    # Plot RMSE Range on right Y-axis (orange, squares)
    line2 = ax2.errorbar(metrics_df["train_end_date"], metrics_df["rmse_range"], 
                         fmt='s-', color='#ff7f0e', label="RMSE Range", linewidth=2, 
                         markersize=6, capsize=4, alpha=0.8)
    ax2.set_ylabel("RMSE Range", color='#ff7f0e')
    ax2.tick_params(axis='y', labelcolor='#ff7f0e')
    
    # Set appropriate Y-axis limits
    ax1.set_ylim(0, max(metrics_df["rmse_rate"].max() * 1.1, 20))
    ax2.set_ylim(0, max(metrics_df["rmse_range"].max() * 1.2, 2))
    
    # Combine legends
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
    
    plt.title(f"RMSE over time - Historical Snapshot {date_str}")
    ax1.grid(True, alpha=0.3)
    plt.tight_layout()
    
    # Save the plot
    rmse_path = os.path.join(plots_dir, "metrics_rmse.png")
    plt.savefig(rmse_path, dpi=160)
    plt.close()
    
    # Coverage plot
    plt.figure(figsize=(10,6))
    plt.errorbar(metrics_df["train_end_date"], metrics_df["coverage_rate_1s"],
                fmt='o-', label="Rate within 1σ", linewidth=2, markersize=6, capsize=4)
    plt.errorbar(metrics_df["train_end_date"], metrics_df["coverage_range_1s"],
                fmt='s-', label="Range within 1σ", linewidth=2, markersize=6, capsize=4)
    
    plt.ylim(0, 1.05)
    plt.xlabel("Train end date")
    plt.ylabel("Coverage")
    plt.title(f"Uncertainty coverage over time - Historical Snapshot {date_str}")
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    coverage_path = os.path.join(plots_dir, "metrics_coverage.png")
    plt.savefig(coverage_path, dpi=160)
    plt.close()

# ==================== LOOCV ====================
def _loocv_path():
    return os.path.join(ROLLING_DIR, "loocv_predictions.csv")

def _needs_loocv(dataset_h, features_h, model_h, code_h):
    f = _read_csv(_loocv_path())
    if f is None or len(f) == 0:
        return True
    
    # Check if the required hash columns exist
    required_columns = ["dataset_hash", "features_hash", "model_hash", "code_hash"]
    if not all(col in f.columns for col in required_columns):
        return True
    
    row = f.iloc[-1]
    return not (row["dataset_hash"] == dataset_h and row["features_hash"] == features_h and
                row["model_hash"] == model_h and row["code_hash"] == code_h)

def _run_loocv(df):
    X = df[FEATURES].astype(float).values
    y_rate  = df["AvgEtchRate"].values
    y_range = df["RangeEtchRate"].values
    n = len(df)
    yp_r = np.zeros(n); ys_r = np.zeros(n)
    yp_g = np.zeros(n); ys_g = np.zeros(n)
    for i in range(n):
        m = np.ones(n, dtype=bool); m[i] = False
        mdl_r = _make_extratrees().fit(X[m], y_rate[m])
        mdl_g = _make_rf().fit(X[m], y_range[m])
        mu_r, sd_r = _pred_stats(mdl_r, X[i:i+1])
        mu_g, sd_g = _pred_stats(mdl_g, X[i:i+1])
        yp_r[i] = mu_r[0]; ys_r[i] = sd_r[0]
        yp_g[i] = mu_g[0]; ys_g[i] = sd_g[0]
    out = df[["LOTNAME","FIMAP_FILE","Date","AvgEtchRate","RangeEtchRate"]].copy()
    out = out.rename(columns={"AvgEtchRate":"loo_true_rate","RangeEtchRate":"loo_true_range"})
    out["loo_pred_rate"] = yp_r
    out["loo_std_rate"]  = ys_r
    out["loo_pred_range"] = yp_g
    out["loo_std_range"]  = ys_g
    return out

def _run_loocv_iteration_specific(df, iteration_num, recipes_df=None):
    """
    Run LOOCV on cumulative data up to a specific iteration with appropriate model selection.
    
    Args:
        df: Full dataset
        iteration_num: Iteration number (0, 1, 2, etc.)
        recipes_df: Excel recipes dataframe to determine which data belongs to which iteration
    
    Returns:
        LOOCV results dataframe with predictions for cumulative data up to the specified iteration
    """
    # For iterations 1 and 2, use Random Forest for both if configured
    use_rf_for_both = (MODEL_CHOICE_FIRST_TWO_ITERATIONS == "rf_both" and iteration_num <= 2)
    print(f"[loocv] Iteration {iteration_num}: MODEL_CHOICE_FIRST_TWO_ITERATIONS = '{MODEL_CHOICE_FIRST_TWO_ITERATIONS}'")
    print(f"[loocv] Iteration {iteration_num}: use_rf_for_both = {use_rf_for_both}")
    
    # Get the proper training data for this iteration
    if recipes_df is not None and iteration_num > 0:
        df_filtered = _get_training_data_for_iteration(df, recipes_df, iteration_num)
        print(f"[loocv] Iteration {iteration_num}: Using {len(df_filtered)} training points "
              f"(run_date < cutoff date for iteration {iteration_num})")
    else:
        # For iteration 0 or no recipes, use full dataset
        df_filtered = df.copy()
        print(f"[loocv] Using full dataset for baseline iteration {iteration_num}")
    
    X = df_filtered[FEATURES].astype(float).values
    y_rate = df_filtered["AvgEtchRate"].values
    y_range = df_filtered["RangeEtchRate"].values
    n = len(df_filtered)
    
    yp_r = np.zeros(n); ys_r = np.zeros(n)
    yp_g = np.zeros(n); ys_g = np.zeros(n)
    
    for i in range(n):
        m = np.ones(n, dtype=bool); m[i] = False
        
        # Choose model based on iteration and configuration
        if use_rf_for_both:
            mdl_r = _make_rf_rate().fit(X[m], y_rate[m])  # Random Forest for rate
            mdl_g = _make_rf().fit(X[m], y_range[m])      # Random Forest for range
        else:
            mdl_r = _make_extratrees().fit(X[m], y_rate[m])  # ExtraTrees for rate
            mdl_g = _make_rf().fit(X[m], y_range[m])         # Random Forest for range
        
        mu_r, sd_r = _pred_stats(mdl_r, X[i:i+1])
        mu_g, sd_g = _pred_stats(mdl_g, X[i:i+1])
        yp_r[i] = mu_r[0]; ys_r[i] = sd_r[0]
        yp_g[i] = mu_g[0]; ys_g[i] = sd_g[0]
    
    out = df_filtered[["LOTNAME","FIMAP_FILE","Date","AvgEtchRate","RangeEtchRate"]].copy()
    out = out.rename(columns={"AvgEtchRate":"loo_true_rate","RangeEtchRate":"loo_true_range"})
    out["loo_pred_rate"] = yp_r
    out["loo_std_rate"] = ys_r
    out["loo_pred_range"] = yp_g
    out["loo_std_range"] = ys_g
    out["iteration_num"] = iteration_num
    out["model_choice"] = "rf_both" if use_rf_for_both else "default"
    
    return out

def _update_loocv(df, dataset_h, features_h, model_h, code_h):
    if _needs_loocv(dataset_h, features_h, model_h, code_h):
        loocv = _run_loocv(df)
        loocv["dataset_hash"]  = dataset_h
        loocv["features_hash"] = features_h
        loocv["model_hash"]    = model_h
        loocv["code_hash"]     = code_h
        loocv["computed_at"]   = datetime.now(timezone.utc).isoformat()
        _atomic_write(_loocv_path(), loocv.to_csv(index=False).encode("utf-8"))
    return _loocv_path()

# ==================== PARETO FRONT HISTORY ====================
def _pareto_history_path():
    return os.path.join(ROLLING_DIR, "pareto_front_history.csv")

def _front_signature(df_front):
    arr = df_front[["AvgEtchRate","Range_nm"]].round(6).values
    return _hash_bytes(arr.tobytes())

def _update_pareto_history(summary_df):
    hist = _read_csv(_pareto_history_path())
    current_front = _pareto_front(summary_df)
    sig_now = _front_signature(current_front)
    now = datetime.now(timezone.utc).isoformat()
    if hist is None or len(hist) == 0:
        rows = []
        for i, r in current_front.iterrows():
            rows.append({
                "snapshot_time": now, "front_version": 1, "point_idx": i,
                "AvgEtchRate": r["AvgEtchRate"], "Range_nm": r["Range_nm"], "front_signature": sig_now
            })
        _atomic_write(_pareto_history_path(), pd.DataFrame(rows).to_csv(index=False).encode("utf-8"))
        return 1, current_front
    last_sig = hist.iloc[-1]["front_signature"]
    last_ver = int(hist["front_version"].max())
    if last_sig == sig_now:
        return last_ver, current_front
    rows = []
    for i, r in current_front.iterrows():
        rows.append({
            "snapshot_time": now, "front_version": last_ver+1, "point_idx": i,
            "AvgEtchRate": r["AvgEtchRate"], "Range_nm": r["Range_nm"], "front_signature": sig_now
        })
    _append_csv_atomic(_pareto_history_path(), pd.DataFrame(rows))
    return last_ver+1, current_front

# ==================== PROPOSALS ====================
def _build_and_write_proposals(X_band, mu_r, mu_rng_nm, sr, srg_nm, pareto_pts):
    r_scale, rn_scale = _objective_scales(
        np.concatenate([mu_r,  pareto_pts[:,0]]) if len(pareto_pts) else mu_r,
        np.concatenate([mu_rng_nm, pareto_pts[:,1]]) if len(pareto_pts) else mu_rng_nm
    )
    selected = []
    selected_knobs = []
    aug_front = pareto_pts.copy() if len(pareto_pts) else np.empty((0,2))
    rows = []
    for k in range(K):
        # First, filter to only Pareto-improving points that significantly push the front
        valid_candidates = []
        for i in range(len(mu_r)):
            if i not in selected:
                if len(aug_front) == 0 or _significantly_improves((mu_r[i], mu_rng_nm[i]), aug_front, 
                                                               RATE_IMPROVEMENT_THRESHOLD, RANGE_IMPROVEMENT_THRESHOLD):
                    valid_candidates.append(i)
        
        # NEW: Filter to ensure only recipes with rf2 > rf1 are proposed
        rf_filtered_candidates = []
        for i in valid_candidates:
            rf1_power = X_band[i, FEATURES.index("Etch_Avg_Rf1_Pow")]
            rf2_power = X_band[i, FEATURES.index("Etch_Avg_Rf2_Pow")]
            if rf2_power > rf1_power:
                rf_filtered_candidates.append(i)
            else:
                print(f"[proposals] Filtered out candidate {i}: Rf1={rf1_power:.1f}, Rf2={rf2_power:.1f} (rf2 must be > rf1)")
        
        if len(rf_filtered_candidates) == 0:
            print(f"⚠️ No candidates with rf2 > rf1 available after selecting {k} recipes!")
            print(f"   Trying with relaxed Pareto criteria...")
            
            # Fallback: try with more relaxed thresholds but still maintain rf2 > rf1
            relaxed_candidates = []
            for i in range(len(mu_r)):
                if i not in selected:
                    if len(aug_front) == 0 or _significantly_improves((mu_r[i], mu_rng_nm[i]), aug_front, 
                                                                   RATE_IMPROVEMENT_THRESHOLD * 0.5, 
                                                                   RANGE_IMPROVEMENT_THRESHOLD * 0.5):
                        relaxed_candidates.append(i)
            
            # Apply rf2 > rf1 filter to relaxed candidates
            rf_filtered_candidates = []
            for i in relaxed_candidates:
                rf1_power = X_band[i, FEATURES.index("Etch_Avg_Rf1_Pow")]
                rf2_power = X_band[i, FEATURES.index("Etch_Avg_Rf2_Pow")]
                if rf2_power > rf1_power:
                    rf_filtered_candidates.append(i)
            
            if len(rf_filtered_candidates) == 0:
                # Final fallback: use original _improves function but still maintain rf2 > rf1
                print(f"   No relaxed candidates with rf2 > rf1 found, using original non-dominated criteria")
                for i in range(len(mu_r)):
                    if i not in selected:
                        if len(aug_front) == 0 or _improves((mu_r[i], mu_rng_nm[i]), aug_front):
                            rf1_power = X_band[i, FEATURES.index("Etch_Avg_Rf1_Pow")]
                            rf2_power = X_band[i, FEATURES.index("Etch_Avg_Rf2_Pow")]
                            if rf2_power > rf1_power:
                                rf_filtered_candidates.append(i)
                
                if len(rf_filtered_candidates) == 0:
                    print(f"   ⚠️ No more Pareto-improving points with rf2 > rf1 available after selecting {k} recipes!")
                    print(f"   This suggests we've reached the limit of what the current model can predict while maintaining rf2 > rf1 constraint.")
                    break
        
        # Use the rf-filtered candidates
        valid_candidates = rf_filtered_candidates
        
        # Debug logging for recipe selection
        if len(aug_front) > 0:
            print(f"[proposals] Iteration {k+1}: Checking {len(mu_r)} candidates against Pareto front with {len(aug_front)} points")
            print(f"[proposals] Found {len(valid_candidates)} valid significantly improving candidates with rf2 > rf1")
            if len(valid_candidates) > 0:
                # Show a few examples of valid candidates with improvement details
                for j, idx in enumerate(valid_candidates[:3]):  # Show first 3
                    rate, range_nm = mu_r[idx], mu_rng_nm[idx]
                    rf1_power = X_band[idx, FEATURES.index("Etch_Avg_Rf1_Pow")]
                    rf2_power = X_band[idx, FEATURES.index("Etch_Avg_Rf2_Pow")]
                    dominates = _dominates_existing((rate, range_nm), aug_front)
                    print(f"[proposals]   Valid candidate {j+1}: Rate={rate:.1f}, Range={range_nm:.1f}, Rf1={rf1_power:.1f}, Rf2={rf2_power:.1f} (Dominates existing: {dominates})")
        else:
            print(f"[proposals] Iteration {k+1}: No existing Pareto front, all {len(mu_r)} candidates are valid")
        
        valid_candidates = np.array(valid_candidates)
        
        if len(valid_candidates) == 0:
            print(f"⚠️ No more significantly Pareto-improving points available after selecting {k} recipes!")
            print(f"   Trying with relaxed criteria...")
            
            # Fallback: try with more relaxed thresholds
            relaxed_candidates = []
            for i in range(len(mu_r)):
                if i not in selected:
                    if len(aug_front) == 0 or _significantly_improves((mu_r[i], mu_rng_nm[i]), aug_front, 
                                                                   RATE_IMPROVEMENT_THRESHOLD * 0.5, 
                                                                   RANGE_IMPROVEMENT_THRESHOLD * 0.5):
                        relaxed_candidates.append(i)
            
            if len(relaxed_candidates) > 0:
                print(f"   Found {len(relaxed_candidates)} candidates with relaxed criteria")
                valid_candidates = relaxed_candidates
            else:
                # Final fallback: use original _improves function
                print(f"   No relaxed candidates found, using original non-dominated criteria")
                for i in range(len(mu_r)):
                    if i not in selected:
                        if len(aug_front) == 0 or _improves((mu_r[i], mu_rng_nm[i]), aug_front):
                            valid_candidates.append(i)
                
                if len(valid_candidates) == 0:
                    print(f"   ⚠️ No more Pareto-improving points available after selecting {k} recipes!")
                    print(f"   This suggests we've reached the limit of what the current model can predict.")
                    break
        
        # Only consider valid candidates for scoring
        expl_raw = np.array([_exploit_distance_norm((mu_r[i], mu_rng_nm[i]), aug_front, r_scale, rn_scale) if len(aug_front) else 0 for i in valid_candidates])
        expl_norm = _norm01(expl_raw)
        
        explore_sigma_rn = sr[valid_candidates] / r_scale
        explore_sigma_gn = srg_nm[valid_candidates] / rn_scale
        explr_raw = np.hypot(explore_sigma_rn, explore_sigma_gn)
        explr_norm = _norm01(explr_raw)
        
        if not selected_knobs:
            center = np.array([0.5]*len(FEATURES))
            div_raw = []
            for i in valid_candidates:
                normalized_vec = []
                for n in FEATURES:
                    val = (X_band[i][FEATURES.index(n)] - FEATURE_RANGES[n][0]) / (FEATURE_RANGES[n][1] - FEATURE_RANGES[n][0])
                    normalized_vec.append(val)
                div_raw.append(np.linalg.norm(np.array(normalized_vec) - center))
            div_raw = np.array(div_raw)
        else:
            div_raw = np.array([_diversity_score(X_band[i], selected_knobs, FEATURES) for i in valid_candidates])
        div_norm = _norm01(div_raw)
        
        combined = ALPHA*expl_norm + BETA*explr_norm + GAMMA*div_norm
        
        # Select the best among valid candidates
        best_valid_idx = int(np.argmax(combined))
        best = valid_candidates[best_valid_idx]
        
        selected.append(best)
        selected_knobs.append(X_band[best])
        aug_front = np.vstack([aug_front, [mu_r[best], mu_rng_nm[best]]]) if len(aug_front) else np.array([[mu_r[best], mu_rng_nm[best]]])
        
        # Debug logging for selected recipe
        rate, range_nm = mu_r[best], mu_rng_nm[best]
        rf1_power = X_band[best, FEATURES.index("Etch_Avg_Rf1_Pow")]
        rf2_power = X_band[best, FEATURES.index("Etch_Avg_Rf2_Pow")]
        print(f"[proposals] Selected recipe {k+1}: Rate={rate:.1f}, Range={range_nm:.1f}, Rf1={rf1_power:.1f}, Rf2={rf2_power:.1f}")
        if len(pareto_pts) > 0:
            # Verify this point is actually non-dominated by the original Pareto front
            original_non_dominated = _improves((rate, range_nm), pareto_pts)
            dominates_original = _dominates_existing((rate, range_nm), pareto_pts)
            print(f"[proposals]   Verifies as non-dominated: {original_non_dominated}")
            print(f"[proposals]   Dominates existing points: {dominates_original}")
            if dominates_original:
                print(f"[proposals]   ✅ This point PUSHES the Pareto front forward!")
            else:
                print(f"[proposals]   ⚠️ This point fills a gap but doesn't dominate existing points")
        
        rows.append({
            "iter": k+1,
            "idx": best,
            "Exploit_raw": float(expl_raw[best_valid_idx]),
            "Exploit_norm": float(expl_norm[best_valid_idx]),
            "Explore_raw": float(explr_raw[best_valid_idx]),
            "Explore_norm": float(explr_norm[best_valid_idx]),
            "Diversity_raw": float(div_raw[best_valid_idx]),
            "Diversity_norm": float(div_norm[best_valid_idx]),
            "Alpha*Exploit_norm": float(ALPHA*expl_norm[best_valid_idx]),
            "Beta*Explore_norm": float(BETA*explr_norm[best_valid_idx]),
            "Gamma*Div_norm": float(GAMMA*div_norm[best_valid_idx]),
            "Combined_score": float(combined[best_valid_idx]),
            "AvgEtchRate_pred": float(mu_r[best]),
            "Range_nm_pred": float(mu_rng_nm[best]),
            "Rate_sigma": float(sr[best]),
            "Range_sigma_nm": float(srg_nm[best]),
        })
    sel_idx = [r["idx"] for r in rows]
    new_recs = X_band[sel_idx]
    rates = mu_r[sel_idx]
    ranges_nm = mu_rng_nm[sel_idx]

    # Summary of Pareto front improvements
    if len(pareto_pts) > 0:
        dominating_points = 0
        for i, (rate, range_nm) in enumerate(zip(rates, ranges_nm)):
            if _dominates_existing((rate, range_nm), pareto_pts):
                dominating_points += 1
                print(f"[proposals] 🎯 Recipe {i+1} dominates existing Pareto front points!")
        
        print(f"[proposals] 📊 Summary: {dominating_points}/{len(rates)} recipes actually push the Pareto front forward")
        if dominating_points == 0:
            print(f"[proposals] ⚠️ Warning: No recipes dominate existing points. Consider adjusting thresholds or sampling strategy.")
    
    # Get uncertainties for selected recipes
    rate_uncertainties = sr[sel_idx]
    range_uncertainties = srg_nm[sel_idx]
    
    df_props = pd.DataFrame({
        "O2_flow":            new_recs[:, FEATURES.index("Etch_AvgO2Flow")],
        "Rf1_Pow":            new_recs[:, FEATURES.index("Etch_Avg_Rf1_Pow")],
        "Rf2_Pow":            new_recs[:, FEATURES.index("Etch_Avg_Rf2_Pow")],
        "Pressure":           new_recs[:, FEATURES.index("Etch_AvgPres")],
        "cf4_flow":           new_recs[:, FEATURES.index("Etch_Avgcf4Flow")],
        "Pred_avg_etch_rate": rates,
        "Pred_Range":         ranges_nm,
        EXCEL_RATE_UNCERTAINTY_COL: rate_uncertainties,
        EXCEL_RANGE_UNCERTAINTY_COL: range_uncertainties
    })
    xlsx_path = os.path.join(SNAP_DIR, f"proposals_{TODAY}.xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
        df_props.to_excel(w, index=False, sheet_name="Proposals")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Proposals"]
    colors = ['FF1F77B4','FFFF7F0E','FF2CA02C']
    for i in range(min(K, len(colors))):
        fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
        for col in range(1, ws.max_column+1):
            ws.cell(row=2+i, column=col).fill = fill
    wb.save(xlsx_path)

    # Scores + Expression (unchanged)
    scores_rows, expr_rows = [], []
    for ridx, rec in enumerate(rows, start=1):
        idx = rec["idx"]
        scores_rows.append({
            "Recipe": f"Recipe {ridx}",
            "Alpha": ALPHA, "Beta": BETA, "Gamma": GAMMA,
            "Exploit_raw": rec["Exploit_raw"], "Exploit_norm": rec["Exploit_norm"],
            "Alpha*Exploit_norm": rec["Alpha*Exploit_norm"],
            "Explore_raw": rec["Explore_raw"], "Explore_norm": rec["Explore_norm"],
            "Beta*Explore_norm": rec["Beta*Explore_norm"],
            "Diversity_raw": rec["Diversity_raw"], "Diversity_norm": rec["Diversity_norm"],
            "Gamma*Div_norm": rec["Gamma*Div_norm"],
            "Combined_score": rec["Combined_score"],
            "AvgEtchRate_pred": rec["AvgEtchRate_pred"],
            "Range_nm_pred": rec["Range_nm_pred"],
            "Rate_sigma": rec["Rate_sigma"],
            "Range_sigma_nm": rec["Range_sigma_nm"],
            "Sampling": SAMPLING_METHOD,
            "N_SAMPLES": N_SAMPLES,
            "Seed": SAMPLING_SEED,
            # Knobs (store for audit)
            "O2_flow": float(X_band[idx, FEATURES.index("Etch_AvgO2Flow")]),
            "cf4_flow": float(X_band[idx, FEATURES.index("Etch_Avgcf4Flow")]),
            "Rf1_Pow": float(X_band[idx, FEATURES.index("Etch_Avg_Rf1_Pow")]),
            "Rf2_Pow": float(X_band[idx, FEATURES.index("Etch_Avg_Rf2_Pow")]),
            "Pressure": float(X_band[idx, FEATURES.index("Etch_AvgPres")]),
        })
        expr_rows.append({"Recipe": f"Recipe {ridx}",
                          "Expression": (f"Recipe {ridx}: score = αE + βX + γD = "
                                         f"{ALPHA:.6f}*{rec['Exploit_norm']:.6f} + "
                                         f"{BETA:.6f}*{rec['Explore_norm']:.6f} + "
                                         f"{GAMMA:.6f}*{rec['Diversity_norm']:.6f} "
                                         f"= {rec['Combined_score']:.6f}")})
    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
        pd.DataFrame(scores_rows).to_excel(w, index=False, sheet_name="Scores")
        pd.DataFrame(expr_rows).to_excel(w, index=False, sheet_name="Expression")
    return xlsx_path, rows, new_recs, rates, ranges_nm

# ==================== PLOTS ====================
def _plot_parity_from_loocv(loocv_csv):
    df = pd.read_csv(loocv_csv)
    df["Date"] = pd.to_datetime(df["Date"])
    y = df["loo_true_rate"].values
    yp = df["loo_pred_rate"].values
    ys = df["loo_std_rate"].values
    r2r = r2_score(y, yp) if len(np.unique(y))>1 else np.nan
    rmse_r = np.sqrt(mean_squared_error(y, yp)) if len(np.unique(y))>1 else np.nan
    plt.figure(figsize=(7,6))
    plt.errorbar(y, yp, yerr=ys, fmt='o', alpha=0.6, ecolor='gray', capsize=2, markersize=4)
    mn = min(y.min(), yp.min()); mx = max(y.max(), yp.max())
    plt.plot([mn,mx],[mn,mx],'k--',lw=1)
    plt.xlabel("Actual AvgEtchRate")
    plt.ylabel("Predicted AvgEtchRate")
    plt.title(f"LOOCV Parity — AvgEtchRate | R² = {r2r:.3f}, RMSE = {rmse_r:.3f}")
    plt.grid(True); plt.tight_layout()
    p1 = os.path.join(PLOTS_DIR, "parity_rate.png")
    plt.savefig(p1, dpi=160); plt.close()

    yg = df["loo_true_range"].values
    ypg = df["loo_pred_range"].values
    ysg = df["loo_std_range"].values
    r2g = r2_score(yg, ypg) if len(np.unique(yg))>1 else np.nan
    rmse_g = np.sqrt(mean_squared_error(yg, ypg)) if len(np.unique(yg))>1 else np.nan
    plt.figure(figsize=(7,6))
    plt.errorbar(yg, ypg, yerr=ysg, fmt='o', alpha=0.6, ecolor='gray', capsize=2, markersize=4)
    mn = min(yg.min(), ypg.min()); mx = max(yg.max(), ypg.max())
    plt.plot([mn,mx],[mn,mx],'k--',lw=1)
    plt.xlabel("Actual RangeEtchRate")
    plt.ylabel("Predicted RangeEtchRate")
    plt.title(f"LOOCV Parity — RangeEtchRate | R² = {r2g:.3f}, RMSE = {rmse_g:.3f}")
    plt.grid(True); plt.tight_layout()
    p2 = os.path.join(PLOTS_DIR, "parity_range.png")
    plt.savefig(p2, dpi=160); plt.close()
    return p1, p2

def _plot_parity_from_loocv_with_highlights(loocv_csv, highlight_lots, recipes_df=None, iteration_num=None, plots_dir=None):
    """Enhanced LOOCV parity plots with highlighted lots using Excel predictions and thickness range conversion"""
    df = pd.read_csv(loocv_csv)
    df["Date"] = pd.to_datetime(df["Date"])
    
    # Rate plot
    y = df["loo_true_rate"].values
    yp = df["loo_pred_rate"].values
    ys = df["loo_std_rate"].values
    r2r = r2_score(y, yp) if len(np.unique(y))>1 else np.nan
    
    # Increased figure size for better visibility
    plt.figure(figsize=(10, 8))
    
    # Plot all points (non-highlighted) - larger markers and error bars
    non_highlight_mask = ~df["LOTNAME"].isin(highlight_lots)
    if non_highlight_mask.any():
        y_nh = y[non_highlight_mask]
        yp_nh = yp[non_highlight_mask]
        ys_nh = ys[non_highlight_mask]
        plt.errorbar(y_nh, yp_nh, yerr=ys_nh, fmt='o', alpha=0.6, ecolor='gray', 
                    capsize=4, markersize=8, label="Historical (LOOCV)", linewidth=2)
    
    # Highlight points using Excel predictions (not LOOCV) - larger markers
    if highlight_lots and recipes_df is not None:
        # Use consistent colors from Pareto front plots
        highlight_colors = HIGHLIGHT_COLORS
        
        for i, lot in enumerate(highlight_lots):
            if lot in df["LOTNAME"].values and lot in recipes_df[EXCEL_LOT_COL].values:
                # Get actual values from LOOCV data
                lot_mask = df["LOTNAME"] == lot
                actual_rate = df.loc[lot_mask, "loo_true_rate"].iloc[0]
                
                # Get predicted values from Excel (not LOOCV)
                excel_mask = recipes_df[EXCEL_LOT_COL] == lot
                if excel_mask.any() and EXCEL_PRED_RATE_COL in recipes_df.columns:
                    pred_rate = recipes_df.loc[excel_mask, EXCEL_PRED_RATE_COL].iloc[0]
                    
                    # Plot highlighted point with consistent color and larger size
                    color = highlight_colors[i % len(highlight_colors)]
                    plt.scatter(actual_rate, pred_rate, s=120, c=color, edgecolors='k', 
                              linewidth=1.5, zorder=5, label=f"{lot} (Excel)")
    
    # Set plot limits and ideal line - thicker line
    all_rates = np.concatenate([y, [pred_rate] if 'pred_rate' in locals() else []])
    mn = min(all_rates.min(), yp.min()) if len(all_rates) > 0 else yp.min()
    mx = max(all_rates.max(), yp.max()) if len(all_rates) > 0 else yp.max()
    plt.plot([mn,mx],[mn,mx],'k--',lw=2, alpha=0.7)
    
    # Larger fonts for labels
    plt.xlabel("Actual AvgEtchRate (nm/min)", fontsize=16, weight='bold')
    plt.ylabel("Predicted AvgEtchRate (nm/min)", fontsize=16, weight='bold')
    
    # Calculate RMSE for the title
    rmse_r = np.sqrt(mean_squared_error(y, yp)) if len(np.unique(y))>1 else np.nan
    
    # Add iteration number to title if provided - larger font
    if iteration_num is not None:
        plt.title(f"Parity — AvgEtchRate | R² = {r2r:.3f}, RMSE = {rmse_r:.3f} - Iteration {iteration_num}", 
                 fontsize=18, weight='bold', pad=20)
    else:
        plt.title(f"Parity — AvgEtchRate | R² = {r2r:.3f}, RMSE = {rmse_r:.3f}", 
                 fontsize=18, weight='bold', pad=20)
    
    # Try to place legend inside below title, fallback to right if it covers points - larger font
    try:
        plt.legend(loc='upper left', bbox_to_anchor=(0.02, 0.98), fontsize=14)
    except:
        # Fallback to right side if there are issues
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize=14)
    
    # Larger grid and tick labels
    plt.grid(True, alpha=0.4, linewidth=1.5)
    plt.xticks(fontsize=14)
    plt.yticks(fontsize=14)
    plt.tight_layout()
    
    # Use provided plots directory or default
    target_plots_dir = plots_dir if plots_dir else PLOTS_DIR
    p1 = os.path.join(target_plots_dir, "parity_rate.png")
    plt.savefig(p1, dpi=160)
    plt.close()

    # Thickness Range plot (converted from RangeEtchRate * 5) - increased size
    yg = df["loo_true_range"].values * 5.0  # Convert to thickness range
    ypg = df["loo_pred_range"].values * 5.0  # Convert to thickness range
    ysg = df["loo_std_range"].values * 5.0   # Convert to thickness range
    r2g = r2_score(yg, ypg) if len(np.unique(yg))>1 else np.nan
    
    plt.figure(figsize=(10, 8))
    
    # Plot all points (non-highlighted) - larger markers and error bars
    if non_highlight_mask.any():
        yg_nh = yg[non_highlight_mask]
        ypg_nh = ypg[non_highlight_mask]
        ysg_nh = ysg[non_highlight_mask]
        plt.errorbar(yg_nh, ypg_nh, yerr=ysg_nh, fmt='o', alpha=0.6, ecolor='gray', 
                    capsize=4, markersize=8, label="Historical (LOOCV)", linewidth=2)
    
    # Highlight points using Excel predictions (not LOOCV) - larger markers
    if highlight_lots and recipes_df is not None:
        # Use consistent colors from Pareto front plots
        highlight_colors = HIGHLIGHT_COLORS
        
        for i, lot in enumerate(highlight_lots):
            if lot in df["LOTNAME"].values and lot in recipes_df[EXCEL_LOT_COL].values:
                # Get actual values from LOOCV data
                lot_mask = df["LOTNAME"] == lot
                actual_range = df.loc[lot_mask, "loo_true_range"].iloc[0] * 5.0  # Convert to thickness range
                
                # Get predicted values from Excel (not LOOCV)
                excel_mask = recipes_df[EXCEL_LOT_COL] == lot
                if excel_mask.any() and EXCEL_PRED_RANGE_COL in recipes_df.columns:
                    pred_range = recipes_df.loc[excel_mask, EXCEL_PRED_RANGE_COL].iloc[0]
                    
                    # Plot highlighted point with consistent color and larger size
                    color = highlight_colors[i % len(highlight_colors)]
                    plt.scatter(actual_range, pred_range, s=120, c=color, edgecolors='k', 
                              linewidth=1.5, zorder=5, label=f"{lot} (Excel)")
    
    # Thicker ideal line
    mn = min(yg.min(), ypg.min()); mx = max(yg.max(), ypg.max())
    plt.plot([mn,mx],[mn,mx],'k--',lw=2)
    
    # Larger fonts for labels
    plt.xlabel("Actual Thickness Range (nm)", fontsize=16, weight='bold')
    plt.ylabel("Predicted Thickness Range (nm)", fontsize=16, weight='bold')
    
    # Calculate RMSE for the title
    rmse_g = np.sqrt(mean_squared_error(yg, ypg)) if len(np.unique(yg))>1 else np.nan
    
    # Add iteration number to title if provided - larger font
    if iteration_num is not None:
        plt.title(f"Parity — Thickness Range | R² = {r2g:.3f}, RMSE = {rmse_g:.3f} - Iteration {iteration_num}", 
                 fontsize=18, weight='bold', pad=20)
    else:
        plt.title(f"Parity — Thickness Range | R² = {r2g:.3f}, RMSE = {rmse_g:.3f}", 
                 fontsize=18, weight='bold', pad=20)
    
    # Try to place legend inside below title, fallback to right if it covers points - larger font
    try:
        plt.legend(loc='upper left', bbox_to_anchor=(0.02, 0.98), fontsize=14)
    except:
        # Fallback to right side if there are issues
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize=14)
    
    # Larger grid and tick labels
    plt.grid(True, alpha=0.4, linewidth=1.5)
    plt.xticks(fontsize=14)
    plt.yticks(fontsize=14)
    plt.tight_layout()
    p2 = os.path.join(target_plots_dir, "parity_range.png")
    plt.savefig(p2, dpi=160)
    plt.close()
    
    return p1, p2

def _plot_front(summary_df, pareto_df, selected_points, highlight_lots, new_completed_lots=None, recipes_df=None, iteration_num=None, plots_dir=None, selected_uncertainties=None):
    """Enhanced Pareto front plot with new completed runs highlighted"""
    plt.figure(figsize=(12, 10))  # Increased figure size for better visibility
    
    # Plot historical data with larger markers and fonts
    plt.scatter(summary_df["AvgEtchRate"], summary_df["Range_nm"], s=80, edgecolor='k', alpha=0.6, label="Historical")
    
    # Plot Pareto front with thicker lines
    plt.plot(pareto_df["AvgEtchRate"], pareto_df["Range_nm"], 'r--', lw=2.5, label="Pareto front")
    plt.scatter(pareto_df["AvgEtchRate"], pareto_df["Range_nm"], s=100, facecolors='none', edgecolors='r', linewidth=1.5)
    
    # Plot new predicted recipes with new symbols (⓿₁, ①₂, ②₃, etc.)
    for i, (r, rng) in enumerate(selected_points):
        # Create the symbol: outside number = point number, inside = iteration number
        point_num = i + 1
        iter_num = iteration_num if iteration_num is not None else 1
        
        # Use Unicode symbols for better appearance
        if point_num == 1:
            symbol = f"⓿{iter_num}" if iter_num < 10 else f"⓿{iter_num}"
        elif point_num == 2:
            symbol = f"①{iter_num}" if iter_num < 10 else f"①{iter_num}"
        elif point_num == 3:
            symbol = f"②{iter_num}" if iter_num < 10 else f"②{iter_num}"
        else:
            symbol = f"③{iter_num}" if iter_num < 10 else f"③{iter_num}"
        
        # Plot the point with a large, visible symbol
        plt.scatter(r, rng, marker='o', s=300, c='gold', edgecolors='k', lw=2, 
                   label="Proposed Recipes" if i == 0 else "", zorder=10)
        
        # Add the symbol text on top
        plt.text(r, rng, symbol, color='black', fontsize=16, weight='bold', 
                ha='center', va='center', zorder=11)
        
        # Add uncertainty bars if provided
        if selected_uncertainties is not None and i < len(selected_uncertainties):
            rate_uncertainty, range_uncertainty = selected_uncertainties[i]
            
            # Vertical uncertainty bar for thickness range (y-axis)
            plt.errorbar(r, rng, yerr=range_uncertainty, fmt='none', color='gold', alpha=0.7, 
                        capsize=5, capthick=2, elinewidth=2)
            
            # Horizontal uncertainty bar for etch rate (x-axis)
            plt.errorbar(r, rng, xerr=rate_uncertainty, fmt='none', color='gold', alpha=0.7, 
                        capsize=5, capthick=2, elinewidth=2)
    
    # Add prediction legend entry right after Proposed Recipes for logical grouping
    if highlight_lots:
        plt.scatter([], [], marker='x', s=150, c='gray', linewidth=2, 
                  label="Predicted Outcomes")
    
    # Plot new completed runs with color coding to match predictions
    if new_completed_lots and recipes_df is not None:
        # Use darker colors for better visibility
        colors = ['#FF8C00', '#800080', '#228B22']  # Dark orange, darker purple, darker green
        
        for i, lot in enumerate(new_completed_lots):
            color = colors[i % len(colors)]
            
            # Check if we have actual experimental data for this lot
            has_actual_data = lot in summary_df["LOTNAME"].values
            
            # Get recipe information from Excel
            recipe_info = None
            if lot in recipes_df[EXCEL_LOT_COL].values:
                recipe_row = recipes_df[recipes_df[EXCEL_LOT_COL] == lot]
                pred_rate = recipe_row[EXCEL_PRED_RATE_COL].iloc[0] if EXCEL_PRED_RATE_COL in recipe_row.columns else None
                pred_range = recipe_row[EXCEL_PRED_RANGE_COL].iloc[0] if EXCEL_PRED_RANGE_COL in recipe_row.columns else None
                ingestion_status = recipe_row[EXCEL_INGEST_COL].iloc[0] if EXCEL_INGEST_COL in recipe_row.columns else None
                comment = recipe_row["Comment"].iloc[0] if "Comment" in recipe_row.columns else ""
                
                recipe_info = {
                    "pred_rate": pred_rate,
                    "pred_range": pred_range,
                    "ingestion_status": ingestion_status,
                    "comment": comment
                }
            
            if has_actual_data and recipe_info and recipe_info["pred_rate"] is not None and recipe_info["pred_range"] is not None:
                # We have both actual data and predictions - plot the actual run
                row = summary_df[summary_df["LOTNAME"] == lot]
                r = row["AvgEtchRate"].iloc[0]
                rng = row["Range_nm"].iloc[0]
                
                # Plot actual result with filled circle (real data)
                plt.scatter(r, rng, marker='o', s=80, c=color, edgecolors='k', lw=0.8, 
                          label=f"{lot} (Actual)")
                
                # Draw line from prediction to actual
                plt.plot([recipe_info["pred_rate"], r], [recipe_info["pred_range"], rng], '--', color=color, alpha=0.7, lw=1)
                
                # Plot prediction point with cross marker (predicted/theoretical)
                plt.scatter(recipe_info["pred_rate"], recipe_info["pred_range"], marker='x', s=150, c=color, linewidth=2, alpha=0.8)
                
                # Add comment if available and ingestion status isn't approved
                if (recipe_info["comment"] and str(recipe_info["comment"]).strip() and 
                    str(recipe_info["comment"]).lower() != "nan" and 
                    recipe_info["ingestion_status"] and recipe_info["ingestion_status"].lower() != "approved"):
                    plt.text(recipe_info["pred_rate"]+2, recipe_info["pred_range"]+0.3, f"Comment: {recipe_info['comment']}", 
                            color=color, fontsize=8, weight='bold', ha='left', 
                            bbox=dict(boxstyle="round,pad=0.3", facecolor='white', alpha=0.8))
            
            elif recipe_info and recipe_info["pred_rate"] is not None and recipe_info["pred_range"] is not None:
                # No actual data but we have predictions - this might be a completed run without approved ingestion
                # Plot only the prediction point and add comment if available
                plt.scatter(recipe_info["pred_rate"], recipe_info["pred_range"], marker='x', s=150, c=color, linewidth=2, alpha=0.8)
                
                # Add comment if available and not empty
                if recipe_info["comment"] and str(recipe_info["comment"]).strip() and str(recipe_info["comment"]).lower() != "nan":
                    plt.text(recipe_info["pred_rate"]+2, recipe_info["pred_range"]+0.3, f"Comment: {recipe_info['comment']}", 
                            color=color, fontsize=8, weight='bold', ha='left', 
                            bbox=dict(boxstyle="round,pad=0.3", facecolor='white', alpha=0.8))
                    plt.scatter([], [], marker='x', s=150, c=color, linewidth=2, 
                              label=f"{lot} (Comment only)")
                else:
                    plt.scatter([], [], marker='x', s=150, c=color, linewidth=2, 
                              label=f"{lot} (No data)")
            else:
                # No recipe info available - just add to legend
                plt.scatter([], [], marker='o', s=80, c=color, edgecolors='k', linewidth=0.8, 
                          label=f"{lot} (Unknown)")
    
    # Plot ALL highlighted lots with their predicted values from Excel
    # Use darker colors for better visibility
    highlight_colors = HIGHLIGHT_COLORS
    
    for i, lot in enumerate(highlight_lots):
        color = highlight_colors[i % len(highlight_colors)]
        
        # Check if we have actual experimental data for this lot
        has_actual_data = lot in summary_df["LOTNAME"].values
        
        # Get recipe information from Excel
        recipe_info = None
        if recipes_df is not None and lot in recipes_df[EXCEL_LOT_COL].values:
            recipe_row = recipes_df[recipes_df[EXCEL_LOT_COL] == lot]
            pred_rate = recipe_row[EXCEL_PRED_RATE_COL].iloc[0] if EXCEL_PRED_RATE_COL in recipe_row.columns else None
            pred_range = recipe_row[EXCEL_PRED_RANGE_COL].iloc[0] if EXCEL_PRED_RANGE_COL in recipe_row.columns else None
            ingestion_status = recipe_row[EXCEL_INGEST_COL].iloc[0] if EXCEL_INGEST_COL in recipe_row.columns else None
            comment = recipe_row["Comment"].iloc[0] if "Comment" in recipe_row.columns else ""
            
            recipe_info = {
                "pred_rate": pred_rate,
                "pred_range": pred_range,
                "ingestion_status": ingestion_status,
                "comment": comment
            }
        
        if has_actual_data and recipe_info and recipe_info["pred_rate"] is not None and recipe_info["pred_range"] is not None:
            # We have both actual data and predictions - plot the actual run
            row = summary_df[summary_df["LOTNAME"] == lot]
            r = row["AvgEtchRate"].iloc[0]
            rng = row["Range_nm"].iloc[0]
            
            # Plot actual result with filled circle (real data)
            plt.scatter(r, rng, marker='o', s=80, c=color, edgecolors='k', linewidth=0.8, 
                      label=f"{lot} (Actual)")
            
            # Draw line from prediction to actual
            plt.plot([recipe_info["pred_rate"], r], [recipe_info["pred_range"], rng], '--', color=color, alpha=0.7, lw=1)
            
            # Plot prediction point with cross marker (predicted/theoretical)
            plt.scatter(recipe_info["pred_rate"], recipe_info["pred_range"], marker='x', s=100, c=color, linewidth=1, alpha=0.8)
            
            # Add comment if available and ingestion status isn't approved
            if (recipe_info["comment"] and str(recipe_info["comment"]).strip() and 
                str(recipe_info["comment"]).lower() != "nan" and 
                recipe_info["ingestion_status"] and recipe_info["ingestion_status"].lower() != "approved"):
                plt.text(recipe_info["pred_rate"]+2, recipe_info["pred_range"]+0.3, f"Comment: {recipe_info['comment']}", 
                        color=color, fontsize=8, weight='bold', ha='left', 
                        bbox=dict(boxstyle="round,pad=0.3", facecolor='white', alpha=0.8))
        
        elif recipe_info and recipe_info["pred_rate"] is not None and recipe_info["pred_range"] is not None:
            # No actual data but we have predictions - this might be a completed run without approved ingestion
            # Plot only the prediction point and add comment if available
            plt.scatter(recipe_info["pred_rate"], recipe_info["pred_range"], marker='x', s=100, c=color, linewidth=1, alpha=0.8)
            
            # Add comment if available
            if recipe_info["comment"]:
                plt.text(recipe_info["pred_rate"]+2, recipe_info["pred_range"]+0.3, f"Comment: {recipe_info['comment']}", 
                        color=color, fontsize=8, weight='bold', ha='left', 
                        bbox=dict(boxstyle="round,pad=0.3", facecolor='white', alpha=0.8))
                plt.scatter([], [], marker='x', s=100, c=color, linewidth=1, 
                          label=f"{lot} (Comment only)")
            else:
                plt.scatter([], [], marker='x', s=100, c=color, linewidth=1, 
                          label=f"{lot} (No data)")
        else:
            # No recipe info available - just add to legend
            plt.scatter([], [], marker='o', s=80, c=color, edgecolors='k', linewidth=0.8, 
                      label=f"{lot} (Unknown)")
    
    # Larger fonts for labels and title
    plt.xlabel("Average Etch Rate (nm/min)", fontsize=22, weight='bold')
    plt.ylabel("Thickness Range (nm)", fontsize=22, weight='bold')
    
    # Add iteration number to title if provided - larger font
    if iteration_num is not None:
        plt.title(f"Pareto + Selected — Top {K} ({SAMPLING_METHOD.upper()} sampling) - Iteration {iteration_num}", 
                 fontsize=24, weight='bold', pad=20)
    else:
        plt.title(f"Pareto + Selected — Top {K} ({SAMPLING_METHOD.upper()} sampling)", 
                 fontsize=24, weight='bold', pad=20)
    
    # Larger legend and grid
    plt.legend(loc="upper left", fontsize=20)  # Move legend to top-left instead of right
    plt.grid(True, alpha=0.4, linewidth=1.5)
    plt.ylim(0, 25)  # Set y-axis limit to 25 for thickness range
    
    # Increase tick label sizes
    plt.xticks(fontsize=18)
    plt.yticks(fontsize=18)
    
    # Check for points that go above 30 nm (off the plot)
    off_plot_points = []
    
    # Check Pareto front values
    if len(pareto_df) > 0:
        for _, row in pareto_df.iterrows():
            if row["Range_nm"] > 30:
                off_plot_points.append(f"Pareto front: ({row['AvgEtchRate']:.1f}, {row['Range_nm']:.1f})")
    
    # Check selected points (proposed recipes)
    if selected_points:
        for i, (rate, rng) in enumerate(selected_points):
            if rng > 30:
                off_plot_points.append(f"Proposed recipe {i+1}: ({rate:.1f}, {rng:.1f})")
    
    # Check highlighted lots (predicted outcomes and actual runs)
    if highlight_lots and recipes_df is not None:
        for lot in highlight_lots:
            if lot in recipes_df[EXCEL_LOT_COL].values:
                recipe_row = recipes_df[recipes_df[EXCEL_LOT_COL] == lot]
                pred_rate = recipe_row[EXCEL_PRED_RATE_COL].iloc[0] if EXCEL_PRED_RATE_COL in recipe_row.columns else None
                pred_range = recipe_row[EXCEL_PRED_RANGE_COL].iloc[0] if EXCEL_PRED_RANGE_COL in recipe_row.columns else None
                
                if pred_range is not None and pred_range > 30:
                    off_plot_points.append(f"{lot} (predicted): ({pred_rate:.1f}, {pred_range:.1f})")
                
                # Also check if we have actual data for this lot
                if lot in summary_df["LOTNAME"].values:
                    row = summary_df[summary_df["LOTNAME"] == lot]
                    actual_rate = row["AvgEtchRate"].iloc[0]
                    actual_range = row["Range_nm"].iloc[0]
                    if actual_range > 30:
                        off_plot_points.append(f"{lot} (actual): ({actual_rate:.1f}, {actual_range:.1f})")
    
    # Check new completed lots
    if new_completed_lots:
        for lot in new_completed_lots:
            if lot in summary_df["LOTNAME"].values:
                row = summary_df[summary_df["LOTNAME"] == lot]
                actual_rate = row["AvgEtchRate"].iloc[0]
                actual_range = row["Range_nm"].iloc[0]
                if actual_range > 30:
                    off_plot_points.append(f"{lot} (new): ({actual_rate:.1f}, {actual_range:.1f})")
    
    # Print off-plot points if any exist
    if off_plot_points:
        warning_msg = f"[plots] Warning: {len(off_plot_points)} points are above 30 nm and may be off the plot:"
        print(warning_msg)
        
        # Log to file for later reference
        log_dir = os.path.join(PLOTS_DIR, "logs")
        os.makedirs(log_dir, exist_ok=True)
        log_file = os.path.join(log_dir, "off_plot_warnings.log")
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(log_file, "a") as f:
            f.write(f"\n{timestamp} - {warning_msg}\n")
            for point in off_plot_points:
                f.write(f"{timestamp} - [plots]   {point}\n")
            f.write(f"{timestamp} - [plots] Y-axis is fixed at 0-20 nm. Consider expanding if needed.\n")
        
        print(f"[plots] Warnings logged to: {log_file}")
        
        for point in off_plot_points:
            print(f"[plots]   {point}")
        print(f"[plots] Y-axis is fixed at 0-20 nm. Consider expanding if needed.")
    
    plt.tight_layout()
    
    # Determine target directory for saving plots
    target_plots_dir = plots_dir if plots_dir else PLOTS_DIR
    p = os.path.join(target_plots_dir, "front.png")
    plt.savefig(p, dpi=160)
    plt.close()
    return p

def _plot_metrics_over_time(csv_path, iteration_num=None, plots_dir=None):
    df = _read_csv(csv_path)
    if df is None or len(df) == 0:
        print(f"[plots] No metrics data found at {csv_path}")
        return None
    
    df["train_end_date"] = pd.to_datetime(df["train_end_date"])
    df = df.sort_values("train_end_date")
    
    # Check if we have enough data for meaningful plots
    if len(df) < 2:
        print(f"[plots] Only {len(df)} data point(s) available, skipping metrics plots")
        return None
    
    print(f"[plots] Creating metrics plots with {len(df)} data points")
    
    # Determine target directory for saving plots
    target_plots_dir = plots_dir if plots_dir else PLOTS_DIR
    
    # RMSE plot with dual Y-axes for better scale separation - increased size
    fig, ax1 = plt.subplots(figsize=(14, 9))
    
    # Calculate confidence intervals (standard error)
    n_points = df["n_points_up_to_date"].values
    rmse_rate_se = df["rmse_rate"].values / np.sqrt(n_points)
    rmse_range_se = df["rmse_range"].values / np.sqrt(n_points)
    
    # Create second Y-axis for range
    ax2 = ax1.twinx()
    
    # Plot RMSE Rate on left Y-axis (blue, circles) - larger markers and lines
    line1 = ax1.errorbar(df["train_end_date"], df["rmse_rate"], yerr=rmse_rate_se, 
                         fmt='o-', color='#1f77b4', label="RMSE Rate", linewidth=3, 
                         markersize=10, capsize=6, alpha=0.8)
    ax1.set_xlabel("Train end date", fontsize=16, weight='bold')
    ax1.set_ylabel("RMSE Rate", color='#1f77b4', fontsize=16, weight='bold')
    ax1.tick_params(axis='y', labelcolor='#1f77b4', labelsize=14)
    ax1.tick_params(axis='x', labelsize=14)
    
    # Plot RMSE Range on right Y-axis (orange, squares) - larger markers and lines
    line2 = ax2.errorbar(df["train_end_date"], df["rmse_range"], yerr=rmse_range_se, 
                         fmt='s-', color='#ff7f0e', label="RMSE Range", linewidth=3, 
                         markersize=10, capsize=6, alpha=0.8)
    ax2.set_ylabel("RMSE Range", color='#ff7f0e', fontsize=16, weight='bold')
    ax2.tick_params(axis='y', labelcolor='#ff7f0e', labelsize=14)
    
    # Set appropriate Y-axis limits for better visibility
    ax1.set_ylim(0, max(df["rmse_rate"].max() * 1.1, 20))  # Rate: 0 to max+10% or 20
    ax2.set_ylim(0, max(df["rmse_range"].max() * 1.2, 2))   # Range: 0 to max+20% or 2
    
    # Highlight latest iteration on both axes - larger markers and fonts
    if len(df) > 0:
        latest_date = df["train_end_date"].iloc[-1]
        latest_rate = df["rmse_rate"].iloc[-1]
        latest_range = df["rmse_range"].iloc[-1]
        
        # Highlight on left axis (Rate) - larger size
        ax1.scatter(latest_date, latest_rate, s=300, c='red', edgecolors='k', zorder=5, label="Latest")
        ax1.text(latest_date, latest_rate + max(rmse_rate_se[-1], 1.0), 
                f"Latest (Iteration {iteration_num})" if iteration_num is not None else "Latest", 
                color='red', fontsize=14, weight='bold', ha='center')
        
        # Highlight on right axis (Range) - larger size
        ax2.scatter(latest_date, latest_range, s=300, c='red', edgecolors='k', zorder=5)
    
    # Combine legends from both axes - larger font
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', fontsize=14)
    
    # Set title based on iteration - larger font
    if iteration_num == 1:
        title = f"RMSE over time - Baseline Model Performance ({len(df)} data points)"
    elif iteration_num is not None:
        title = f"RMSE over time - Iteration {iteration_num} ({len(df)} data points)"
    else:
        title = f"RMSE over time ({len(df)} data points)"
    
    plt.title(title, fontsize=18, weight='bold', pad=20)
    ax1.grid(True, alpha=0.4, linewidth=1.5)
    plt.tight_layout()
    
    # Save to appropriate directory
    p1 = os.path.join(target_plots_dir, "metrics_rmse.png")
    plt.savefig(p1, dpi=160)
    plt.close()

    # Coverage plot with confidence intervals - increased size
    plt.figure(figsize=(14, 8))
    
    # Calculate confidence intervals for coverage
    coverage_rate_se = np.sqrt(df["coverage_rate_1s"].values * (1 - df["coverage_rate_1s"].values) / n_points)
    coverage_range_se = np.sqrt(df["coverage_range_1s"].values * (1 - df["coverage_range_1s"].values) / n_points)
    
    # Larger markers and lines
    plt.errorbar(df["train_end_date"], df["coverage_rate_1s"], yerr=coverage_rate_se,
                fmt='o-', label="Rate within 1σ", linewidth=3, markersize=10, capsize=6)
    plt.errorbar(df["train_end_date"], df["coverage_range_1s"], yerr=coverage_range_se,
                fmt='s-', label="Range within 1σ", linewidth=3, markersize=10, capsize=6)
    
    # Highlight latest iteration - larger markers and fonts
    if len(df) > 0:
        latest_coverage_rate = df["coverage_rate_1s"].iloc[-1]
        latest_coverage_range = df["coverage_range_1s"].iloc[-1]
        plt.scatter(latest_date, latest_coverage_rate, s=300, c='red', edgecolors='k', zorder=5)
        plt.scatter(latest_date, latest_coverage_range, s=300, c='red', edgecolors='k', zorder=5)
        plt.text(latest_date, latest_coverage_rate + max(coverage_rate_se[-1], 0.02), 
                f"Latest (Iteration {iteration_num})" if iteration_num is not None else "Latest", 
                color='red', fontsize=14, weight='bold', ha='center')
    
    plt.ylim(0, 1.05)
    # Larger fonts for labels
    plt.xlabel("Train end date", fontsize=16, weight='bold')
    plt.ylabel("Coverage", fontsize=16, weight='bold')
    
    # Set title based on iteration - larger font
    if iteration_num == 1:
        title = f"Uncertainty Coverage over time - Baseline Model Performance ({len(df)} data points)"
    elif iteration_num is not None:
        title = f"Uncertainty Coverage over time - Iteration {iteration_num} ({len(df)} data points)"
    else:
        title = f"Uncertainty Coverage over time ({len(df)} data points)"
    
    plt.title(title, fontsize=18, weight='bold', pad=20)
    plt.legend(fontsize=14)
    plt.grid(True, alpha=0.4, linewidth=1.5)
    
    # Increase tick label sizes
    plt.xticks(fontsize=14)
    plt.yticks(fontsize=14)
    
    plt.tight_layout()
    
    # Save to appropriate directory
    p2 = os.path.join(target_plots_dir, "metrics_coverage.png")
    plt.savefig(p2, dpi=160)
    plt.close()
    
    return p1, p2

# ==================== INGESTION STATUS UPDATE ====================
def _update_ingestion_status(recipes_df, completed_lots):
    """Update ingestion status to 'approved' for successfully processed lots"""
    if not _HAS_MSAL or not all([GRAPH_CLIENT_ID, GRAPH_CLIENT_SECRET, GRAPH_TENANT_ID]):
        print("[ingestion] No Graph API credentials available, skipping status update")
        return False
    
    if not completed_lots:
        print("[ingestion] No completed lots to update")
        return False
    
    try:
        print(f"[ingestion] Updating ingestion status to 'approved' for {len(completed_lots)} lots")
        
        # Authenticate with Microsoft Graph
        authority = f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}"
        scopes = ["https://graph.microsoft.com/.default"]
        app = msal.ConfidentialClientApplication(
            GRAPH_CLIENT_ID, authority=authority, client_credential=GRAPH_CLIENT_SECRET
        )
        token_resp = app.acquire_token_for_client(scopes=scopes)
        if "access_token" not in token_resp:
            raise RuntimeError(f"MSAL token error: {token_resp}")
        token = token_resp["access_token"]
        headers = {"Authorization": f"Bearer {token}"}
        
        # Get site and drive information
        site_url = f"https://graph.microsoft.com/v1.0/sites/{GRAPH_TENANT_NAME}.sharepoint.com:/sites/{GRAPH_SITE_NAME}"
        site_id = requests.get(site_url, headers=headers).json()["id"]
        drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
        drives = requests.get(drives_url, headers=headers).json()["value"]
        drive_id = next(d["id"] for d in drives if d["name"] == "Documents")
        
        # Download current file
        download_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{RECIPES_FILE_PATH}:/content"
        resp = requests.get(download_url, headers=headers)
        resp.raise_for_status()
        
        # Load workbook and update status
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        
        # Find column indices
        lot_col_idx = None
        status_col_idx = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == EXCEL_LOT_COL:
                lot_col_idx = col
            elif ws.cell(row=1, column=col).value == EXCEL_INGEST_COL:
                status_col_idx = col
        
        if lot_col_idx is None or status_col_idx is None:
            print(f"[ingestion] Could not find required columns: {EXCEL_LOT_COL}, {EXCEL_INGEST_COL}")
            return False
        
        # Update status for completed lots
        updated_count = 0
        for row in range(2, ws.max_row + 1):
            lot_name = ws.cell(row=row, column=lot_col_idx).value
            if lot_name in completed_lots:
                ws.cell(row=row, column=status_col_idx).value = "approved"
                updated_count += 1
        
        if updated_count > 0:
            # Save updated workbook
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            
            # Upload back to SharePoint
            upload_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{RECIPES_FILE_PATH}:/content"
            upload_headers = {**headers, "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
            upload_resp = requests.put(upload_url, headers=upload_headers, data=bio.getvalue())
            upload_resp.raise_for_status()
            
            print(f"[ingestion] Successfully updated {updated_count} lots to 'approved' status")
            return True
        else:
            print("[ingestion] No lots found to update")
            return False
            
    except Exception as e:
        print(f"[ingestion] Error updating status: {e}")
        return False

# ==================== HISTORICAL PLOT VIEWER ====================
def _view_historical_plots(target_date=None):
    """View plots from a specific date or the latest available date"""
    if target_date is None:
        # Find the latest available snapshot
        if not os.path.exists(SNAPSHOTS_DIR):
            print("No snapshots directory found")
            return
        
        available_dates = [d for d in os.listdir(SNAPSHOTS_DIR) if os.path.isdir(os.path.join(SNAPSHOTS_DIR, d))]
        if not available_dates:
            print("No snapshots found")
            return
        
        # Sort dates and pick the latest
        available_dates.sort()
        target_date = available_dates[-1]
        print(f"Showing latest available plots from: {target_date}")
    
    # Check if target date exists
    target_dir = os.path.join(SNAPSHOTS_DIR, target_date)
    if not os.path.exists(target_dir):
        print(f"No snapshot found for date: {target_date}")
        return
    
    plots_dir = os.path.join(target_dir, "plots")
    if not os.path.exists(plots_dir):
        print(f"No plots directory found for date: {target_date}")
        return
    
    # Display available plots
    available_plots = [f for f in os.listdir(plots_dir) if f.endswith('.png')]
    if not available_plots:
        print(f"No plots found for date: {target_date}")
        return
    
    print(f"Available plots for {target_date}:")
    for plot in available_plots:
        print(f"  - {plot}")
    
    # Display plots
    for plot in available_plots:
        plot_path = os.path.join(plots_dir, plot)
        print(f"\nDisplaying: {plot}")
        
        img = plt.imread(plot_path)
        plt.figure(figsize=(10, 8))
        plt.imshow(img)
        plt.axis('off')
        plt.title(f"{plot} - {target_date}")
        plt.tight_layout()
        plt.close()
    
    print(f"\nAll plots from {target_date} have been displayed")

def _get_latest_snapshot_date():
    """Get the latest available snapshot date"""
    if not os.path.exists(SNAPSHOTS_DIR):
        return None
    
    available_dates = [d for d in os.listdir(SNAPSHOTS_DIR) if os.path.isdir(os.path.join(SNAPSHOTS_DIR, d))]
    if not available_dates:
        return None
    
    available_dates.sort()
    return available_dates[-1]

# ==================== ITERATION-BASED SYSTEM ====================
def _get_iteration_training_cutoff_date(recipes_df, iteration_num):
    """
    Get the training data cutoff date for a specific iteration.
    
    For iteration N, we train on all data with run_date < completion date of first row of iteration N.
    
    Args:
        recipes_df: Excel recipes dataframe
        iteration_num: Iteration number (1, 2, 3, ...)
    
    Returns:
        datetime object representing the cutoff date, or None if not available
    """
    if recipes_df is None or EXCEL_DATE_COL not in recipes_df.columns:
        return None
    
    # Calculate the row index for the first row of this iteration
    # Iteration 1: rows 0-2 (indices 0,1,2), first row = 0
    # Iteration 2: rows 3-5 (indices 3,4,5), first row = 3
    # Iteration 3: rows 6-8 (indices 6,7,8), first row = 6
    first_row_idx = (iteration_num - 1) * POINTS_PER_ITERATION
    
    if first_row_idx >= len(recipes_df):
        print(f"[training] Iteration {iteration_num}: No data available in Excel - this iteration needs new recipes proposed")
        return None
    
    # Get the completion date of the first row of this iteration
    first_row = recipes_df.iloc[first_row_idx]
    completion_date = first_row.get(EXCEL_DATE_COL)
    
    if pd.isna(completion_date) or completion_date is None:
        print(f"[training] Iteration {iteration_num}: No completion date for first row")
        return None
    
    # Convert Excel date format (MM/DD/YYYY) to datetime if needed
    if isinstance(completion_date, str):
        try:
            # Handle MM/DD/YYYY format
            completion_date = pd.to_datetime(completion_date, format='%m/%d/%Y')
        except:
            # Try other formats
            completion_date = pd.to_datetime(completion_date, errors='coerce')
    
    if pd.isna(completion_date):
        print(f"[training] Iteration {iteration_num}: Could not parse completion date: {first_row.get(EXCEL_DATE_COL)}")
        return None
    
    # Ensure timezone-naive datetime for consistent comparison
    if completion_date.tz is not None:
        completion_date = completion_date.tz_localize(None)
    
    print(f"[training] Iteration {iteration_num}: Training cutoff date = {completion_date.strftime('%Y-%m-%d')}")
    return completion_date

def _get_training_data_for_iteration(df, recipes_df, iteration_num):
    """
    Get the training data for a specific iteration based on the cutoff date.
    
    Args:
        df: Full dataset from full_dataset.csv
        recipes_df: Excel recipes dataframe
        iteration_num: Iteration number (1, 2, 3, ...)
    
    Returns:
        DataFrame containing the training data for this iteration
    """
    print(f"[training] Getting training data for iteration {iteration_num}")
    
    # Get the cutoff date for this iteration
    cutoff_date = _get_iteration_training_cutoff_date(recipes_df, iteration_num)
    
    if cutoff_date is None:
        print(f"[training] Iteration {iteration_num}: No cutoff date available, using full dataset")
        return df.copy()
    
    # Ensure run_date column is properly converted to datetime
    df_copy = _ensure_run_date_datetime(df)
    
    # Filter data to include all run_dates < cutoff_date
    # This ensures we don't include the experimental points from this iteration in training
    training_mask = df_copy["run_date"] < cutoff_date
    training_data = df_copy[training_mask].copy()
    
    print(f"[training] Iteration {iteration_num}: Training on {len(training_data)} points "
          f"(run_date < {cutoff_date.strftime('%Y-%m-%d')})")
    
    # Log the training data progression for verification
    if len(training_data) > 0:
        min_date = training_data["run_date"].min()
        max_date = training_data["run_date"].max()
        print(f"[training] Iteration {iteration_num}: Training data spans from {min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}")
    
    return training_data

def _ensure_run_date_datetime(df):
    """
    Ensure the run_date column in the dataset is properly converted to datetime.
    
    Args:
        df: DataFrame with run_date column
    
    Returns:
        DataFrame with run_date converted to datetime, or original if conversion fails
    """
    if "run_date" not in df.columns:
        print("[training] Warning: No run_date column found in dataset")
        return df
    
    if df["run_date"].dtype == 'object':
        try:
            df_copy = df.copy()
            df_copy["run_date"] = pd.to_datetime(df_copy["run_date"])
            print("[training] Successfully converted run_date column to datetime")
        except Exception as e:
            print(f"[training] Warning: Could not convert run_date to datetime: {e}")
            print("[training] Using original dataset (date comparisons may fail)")
            return df
    else:
        # Already datetime type
        df_copy = df.copy()
    
    # Ensure timezone-naive datetime for consistent comparison
    if df_copy["run_date"].dt.tz is not None:
        df_copy["run_date"] = df_copy["run_date"].dt.tz_localize(None)
        print("[training] Converted timezone-aware dates to timezone-naive")
    
    return df_copy

def _get_completed_experimental_data_for_iteration(df, recipes_df, iteration_num):
    """
    Get the completed experimental data for a specific iteration from full_dataset.csv.
    
    Args:
        df: Full dataset from full_dataset.csv
        recipes_df: Excel recipes dataframe
        iteration_num: Iteration number (1, 2, 3, ...)
    
    Returns:
        DataFrame containing the experimental data for this iteration
    """
    if recipes_df is None or EXCEL_LOT_COL not in recipes_df.columns:
        return pd.DataFrame()
    
    # Get the lotnames for this iteration
    start_idx = (iteration_num - 1) * POINTS_PER_ITERATION
    end_idx = start_idx + POINTS_PER_ITERATION
    
    if start_idx >= len(recipes_df):
        return pd.DataFrame()
    
    iter_recipes = recipes_df.iloc[start_idx:end_idx]
    
    # Get lotnames that have completion dates
    completed_lots = []
    for _, recipe in iter_recipes.iterrows():
        lotname = recipe.get(EXCEL_LOT_COL)
        completion_date = recipe.get(EXCEL_DATE_COL)
        
        if (isinstance(lotname, str) and lotname.strip() and 
            not pd.isna(completion_date) and completion_date is not None):
            completed_lots.append(lotname)
    
    if not completed_lots:
        print(f"[training] Iteration {iteration_num}: No completed lots found")
        return pd.DataFrame()
    
    # Find these lots in the full dataset
    experimental_data = df[df["LOTNAME"].isin(completed_lots)].copy()
    
    print(f"[training] Iteration {iteration_num}: Found {len(experimental_data)} experimental points "
          f"for lots: {', '.join(completed_lots)}")
    
    return experimental_data

def _get_training_data_for_main_proposals(df, recipes_df):
    """
    Get the training data for the main proposals based on the current iteration status.
    
    This function determines what data should be used for training the main models
    that generate new recipe proposals.
    
    Args:
        df: Full dataset from full_dataset.csv
        recipes_df: Excel recipes dataframe
    
    Returns:
        DataFrame containing the training data for main proposals
    """
    if recipes_df is None:
        print("[training] No recipes Excel available, using full dataset for main proposals")
        return df.copy()
    
    # Find the highest iteration that has proposed recipes
    iteration_status = _get_excel_iteration_status(recipes_df)
    if not iteration_status:
        print("[training] No iterations found in Excel, using full dataset for main proposals")
        return df.copy()
    
    max_iteration = max(iteration_status.keys())
    
    # For the main proposals, we want to train on data that would be available
    # for the next iteration (iteration max_iteration + 1)
    next_iteration = max_iteration + 1
    
    # Get the training data for the next iteration
    training_data = _get_training_data_for_iteration(df, recipes_df, next_iteration)
    
    print(f"[training] Main proposals: Using {len(training_data)} training points "
          f"for iteration {next_iteration} (current max: {max_iteration})")
    
    return training_data

def _ensure_iteration_dirs():
    """Ensure iteration directories exist"""
    for i in range(100):  # Support up to 100 iterations
        iter_dir = os.path.join(ITERATIONS_DIR, f"iteration_{i}")
        iter_plots_dir = os.path.join(iter_dir, "plots")
        os.makedirs(iter_dir, exist_ok=True)
        os.makedirs(iter_dir, exist_ok=True)

def _get_excel_iteration_status(recipes_df):
    """
    Analyze the Excel file to determine which iterations have proposed recipes and which are completed.
    Returns a dictionary with iteration status information.
    """
    if recipes_df is None or EXCEL_STATUS_COL not in recipes_df.columns:
        return {}
    
    # Group recipes by iteration (3 recipes per iteration)
    total_recipes = len(recipes_df)
    total_iterations = total_recipes // POINTS_PER_ITERATION
    
    iteration_status = {}
    
    for iter_num in range(total_iterations):
        start_idx = iter_num * POINTS_PER_ITERATION
        end_idx = start_idx + POINTS_PER_ITERATION
        
        # Get recipes for this iteration
        iter_recipes = recipes_df.iloc[start_idx:end_idx]
        
        # Count completed vs pending
        completed_count = len(iter_recipes[iter_recipes["Status_norm"] == "completed"])
        pending_count = len(iter_recipes[iter_recipes["Status_norm"] == "pending"])
        
        iteration_status[iter_num] = {
            "start_idx": start_idx,
            "end_idx": end_idx,
            "completed_count": completed_count,
            "pending_count": pending_count,
            "is_completed": completed_count == POINTS_PER_ITERATION,
            "recipes": iter_recipes
        }
    
    return iteration_status

def _get_proposed_recipes_for_iteration(recipes_df, iteration_num):
    """
    Get the proposed recipes for a specific iteration from the Excel file.
    Returns None if the iteration doesn't have proposed recipes yet.
    """
    if recipes_df is None or EXCEL_STATUS_COL not in recipes_df.columns:
        return None
    
    # For iteration 1, read rows 0,1,2; for iteration 2, read rows 3,4,5; etc.
    start_idx = (iteration_num - 1) * POINTS_PER_ITERATION
    end_idx = start_idx + POINTS_PER_ITERATION
    
    if start_idx >= len(recipes_df):
        return None
    
    iter_recipes = recipes_df.iloc[start_idx:end_idx]
    
    # Check if this iteration has proposed recipes
    if len(iter_recipes) == 0:
        return None
    
    # Extract the recipe data
    proposed_recipes = []
    for i, (_, recipe) in enumerate(iter_recipes.iterrows()):
        recipe_data = {
            "O2_flow": recipe.get("O2_flow", 0.0),
            "cf4_flow": recipe.get("cf4_flow", 0.0),
            "Rf1_Pow": recipe.get("Rf1_Pow", 0.0),
            "Rf2_Pow": recipe.get("Rf2_Pow", 0.0),
            "Pressure": recipe.get("Pressure", 0.0),
            "predicted_rate": recipe.get(EXCEL_PRED_RATE_COL, 0.0),
            "predicted_range": recipe.get(EXCEL_PRED_RANGE_COL, 0.0),
            "rate_uncertainty": recipe.get(EXCEL_RATE_UNCERTAINTY_COL, 0.0),
            "range_uncertainty": recipe.get(EXCEL_RANGE_UNCERTAINTY_COL, 0.0),
            "status": recipe.get(EXCEL_STATUS_COL, ""),
            "lotname": recipe.get(EXCEL_LOT_COL, ""),
            "date_completed": recipe.get(EXCEL_DATE_COL, "")
        }

        proposed_recipes.append(recipe_data)
    
    return proposed_recipes

def _should_propose_new_recipes(recipes_df, current_iteration):
    """
    Determine if we should propose new recipes for the next iteration.
    Returns True if:
    1. The current iteration doesn't exist in Excel (we need to propose recipes for it), OR
    2. The current iteration is completed and we don't already have recipes for the next iteration.
    """
    if recipes_df is None:
        return False
    
    iteration_status = _get_excel_iteration_status(recipes_df)
    
    # If current iteration doesn't exist in Excel, we definitely need to propose recipes for it
    if current_iteration not in iteration_status:
        return True
    
    # If current iteration exists but isn't completed, don't propose new recipes yet
    if not iteration_status[current_iteration]["is_completed"]:
        return False
    
    # Check if next iteration already has proposed recipes
    next_iteration = current_iteration + 1
    if next_iteration in iteration_status:
        return False
    
    return True

def _build_and_write_proposals_ei_batch(Xb, mur, mug, sdr, sdg, pareto_pts, training_df):
    """
    Build and write proposals using Expected Improvement (EI) + Euclidean distance + batch approach.
    
    Args:
        Xb: Candidate features
        mur: Predicted etch rates
        mug: Predicted ranges (in nm)
        sdr: Rate uncertainties
        sdg: Range uncertainties (in nm)
        pareto_pts: Current Pareto front points
        training_df: Training data for distance calculations
    
    Returns:
        Tuple of (xlsx_path, sel_rows, new_recs, rates, ranges_nm)
    """
    if len(Xb) == 0:
        print("[proposals] No candidates available for selection")
        return None
    
    # Import required functions locally to avoid import issues
    from scipy.stats import norm
    from scipy.spatial.distance import cdist
    
    # Calculate Expected Improvement (EI) for each candidate
    # EI = (μ - μ*) * Φ((μ - μ*)/σ) + σ * φ((μ - μ*)/σ)
    # where μ* is the best observed value, Φ is CDF, φ is PDF
    
    # For multi-objective optimization, we'll use a weighted sum approach
    # Normalize objectives to [0,1] range for fair comparison
    
    # Normalize etch rate (higher is better)
    rate_min, rate_max = mur.min(), mur.max()
    rate_norm = (mur - rate_min) / (rate_max - rate_min) if rate_max > rate_min else np.zeros_like(mur)
    
    # Normalize range (lower is better, so invert)
    range_min, range_max = mug.min(), mug.max()
    range_norm = 1.0 - (mug - range_min) / (range_max - range_min) if range_max > range_min else np.zeros_like(mug)
    
    # Combined objective: maximize rate + minimize range
    # Weight can be adjusted based on preference
    combined_objective = 0.6 * rate_norm + 0.4 * range_norm
    
    # Calculate EI for each candidate
    best_observed = combined_objective.max()
    ei_values = np.zeros(len(Xb))
    
    for i in range(len(Xb)):
        mu_i = combined_objective[i]
        sigma_i = np.sqrt(0.6**2 * sdr[i]**2 + 0.4**2 * sdg[i]**2)  # Combined uncertainty
        
        if sigma_i > 0:
            z = (mu_i - best_observed) / sigma_i
            ei_values[i] = (mu_i - best_observed) * norm.cdf(z) + sigma_i * norm.pdf(z)
        else:
            ei_values[i] = 0
    
    # Batch selection using EI + Euclidean distance
    batch_size = POINTS_PER_ITERATION
    selected_indices = []
    
    # First selection: highest EI
    first_idx = np.argmax(ei_values)
    selected_indices.append(first_idx)
    
    # Subsequent selections: balance EI with diversity (Euclidean distance)
    for batch_idx in range(1, batch_size):
        if len(selected_indices) >= len(Xb):
            break
            
        # Calculate distances to already selected points
        if len(selected_indices) > 0:
            selected_features = Xb[selected_indices]
            distances = cdist(Xb, selected_features, metric='euclidean')
            min_distances = np.min(distances, axis=1)
        else:
            min_distances = np.ones(len(Xb)) * np.inf
        
        # Score candidates: balance EI and diversity
        # Normalize both metrics to [0,1]
        ei_norm = (ei_values - ei_values.min()) / (ei_values.max() - ei_values.min()) if ei_values.max() > ei_values.min() else np.zeros_like(ei_values)
        dist_norm = (min_distances - min_distances.min()) / (min_distances.max() - min_distances.min()) if min_distances.max() > min_distances.min() else np.zeros_like(min_distances)
        
        # Combined score: 70% EI + 30% diversity
        combined_score = 0.7 * ei_norm + 0.3 * dist_norm
        
        # Remove already selected candidates
        combined_score[selected_indices] = -np.inf
        
        # Select next candidate
        next_idx = np.argmax(combined_score)
        selected_indices.append(next_idx)
    
    print(f"[proposals] Selected {len(selected_indices)} candidates using EI + Euclidean distance batch approach")
    
    # Use the existing _build_and_write_proposals function with selected indices
    Xb_selected = Xb[selected_indices]
    mur_selected = mur[selected_indices]
    mug_selected = mug[selected_indices]
    sdr_selected = sdr[selected_indices]
    sdg_selected = sdg[selected_indices]
    
    return _build_and_write_proposals(Xb_selected, mur_selected, mug_selected, sdr_selected, sdg_selected, pareto_pts)

def _propose_next_iteration_recipes(recipes_df, current_iteration, selected_points_with_features):
    """
    Propose the next 3 recipes for the next iteration and add them to Excel.
    Generates fresh recipes using the trained models instead of copying old ones.
    """
    if not _should_propose_new_recipes(recipes_df, current_iteration):
        return False
    
    next_iteration = current_iteration + 1
    print(f"[proposals] Generating new recipes for iteration {next_iteration}...")
    
    # Import required functions and constants
    from sklearn.ensemble import RandomForestRegressor, ExtraTreesRegressor
    from sklearn.preprocessing import StandardScaler
    from sklearn.pipeline import Pipeline
    from scipy.stats import norm
    from scipy.spatial.distance import cdist
    
    try:
        # Load the current dataset to retrain models with new data
        df = pd.read_csv(DATASET_CSV)
        if df.empty:
            print(f"[proposals] Error: Dataset is empty, cannot generate new recipes")
            return False
    
        # Get the proper training data for this iteration
        training_df = _get_training_data_for_iteration(df, recipes_df, next_iteration)
        
        if training_df.empty:
            print(f"[proposals] Error: No training data available for iteration {next_iteration}")
            return False
        
        # Prepare features and targets from training data
        X = training_df[FEATURES].astype(float).values
        y_rate = training_df["AvgEtchRate"].values
        y_range = training_df["RangeEtchRate"].values
        
        # Retrain models on iterative training data
        print(f"[proposals] Retraining models on {len(training_df)} training points for iteration {next_iteration}...")
        
        # Determine which models to use based on iteration and configuration
        if next_iteration <= 2:
            # Iterations 1-2: Random Forest for both rate and range
            print(f"[proposals] Using Random Forest for both rate and range (iteration {next_iteration})")
            model_rate = _make_rf_rate().fit(X, y_rate)
            model_range = _make_rf().fit(X, y_range)
        elif next_iteration <= 4:
            # Iterations 3-4: Extra Trees for rate, Random Forest for range
            print(f"[proposals] Using Extra Trees for rate, Random Forest for range (iteration {next_iteration})")
            model_rate = _make_extratrees().fit(X, y_rate)
            model_range = _make_rf().fit(X, y_range)
        else:
            # Iterations 5+: GPR with hyperparameter optimization for both rate and range
            print(f"[proposals] Using GPR with hyperparameter optimization for both rate and range (iteration {next_iteration})")
            model_rate, _ = _tune_gpr(X, y_rate)
            model_range, _ = _tune_gpr(X, y_range)
        
        # Sample new candidates using the same method as the main script
        print(f"[proposals] Sampling {N_SAMPLES} new candidate recipes...")
        lower = np.array([FEATURE_RANGES[c][0] for c in FEATURES], float)
        upper = np.array([FEATURE_RANGES[c][1] for c in FEATURES], float)
        
        # Use the same sampling function that's used in main()
        Xcand = _sample_candidates(SAMPLING_METHOD, N_SAMPLES, lower, upper, SAMPLING_SEED)
        Xcand = _quantize(Xcand, FEATURES)
        
        # Predict outcomes for new candidates
        print(f"[proposals] Predicting outcomes for new candidates...")
        mu_r, sd_r = _pred_stats(model_rate, Xcand)
        mu_g, sd_g = _pred_stats(model_range, Xcand)
        mu_g_nm = mu_g * 5.0  # Convert to nm
        
        # Apply target rate filter
        mask = (mu_r >= TARGET_RATE_MIN) & (mu_r <= TARGET_RATE_MAX)
        if not np.any(mask):
            mask = np.ones(len(mu_r), dtype=bool)
        
        # Apply rf2 > rf1 filter
        rf_mask = np.ones(len(mu_r), dtype=bool)
        for i in range(len(mu_r)):
            if mask[i]:  # Only check if already passed target rate filter
                rf1_power = Xcand[i, FEATURES.index("Etch_Avg_Rf1_Pow")]
                rf2_power = Xcand[i, FEATURES.index("Etch_Avg_Rf2_Pow")]
                if rf2_power <= rf1_power:
                    rf_mask[i] = False
        
        # Combine both filters
        combined_mask = mask & rf_mask
        
        if not np.any(combined_mask):
            print(f"[proposals] Warning: No candidates meet both target rate and rf2 > rf1 requirements")
            print(f"[proposals] Target rate filter: {np.sum(mask)} candidates")
            print(f"[proposals] Rf2 > rf1 filter: {np.sum(rf_mask)} candidates")
            print(f"[proposals] Combined filter: {np.sum(combined_mask)} candidates")
            # Fallback to just target rate filter if no candidates meet both
            combined_mask = mask
        
        Xb = Xcand[combined_mask]
        mur = mu_r[combined_mask]
        mug = mu_g_nm[combined_mask]
        sdr = sd_r[combined_mask]
        sdg = sd_g[combined_mask] * 5.0  # Convert to nm
        
        print(f"[proposals] Filtered candidates: {np.sum(mask)} (target rate) → {np.sum(combined_mask)} (with rf2 > rf1)")
        
        # Get current Pareto front to ensure new recipes improve it
        # Use the iterative training data for consistent comparison
        pareto_pts = []
        try:
            # Calculate Pareto front from the training data used for this iteration
            df_complete = training_df[["LOTNAME","FIMAP_FILE","AvgEtchRate","RangeEtchRate"]].copy()
            df_complete["Range_nm"] = df_complete["RangeEtchRate"] * 5.0
            
            # Calculate Pareto front from training data
            pts = df_complete[["AvgEtchRate","Range_nm"]].values
            flags = _is_pareto(pts)
            current_front = df_complete[flags].sort_values("AvgEtchRate")
            pareto_pts = current_front[["AvgEtchRate","Range_nm"]].values
            
            print(f"[proposals] Using iterative training data: {len(df_complete)} points → Pareto front with {len(pareto_pts)} points")
        except Exception as e:
            print(f"[proposals] Warning: Could not calculate Pareto front from training data: {e}")
            # Fallback to Excel recipes if calculation fails
            if len(recipes_df) > 0:
                completed_recipes = recipes_df[recipes_df[EXCEL_STATUS_COL].str.lower() == "completed"]
                if len(completed_recipes) > 0:
                    pareto_pts = completed_recipes[[EXCEL_PRED_RATE_COL, EXCEL_PRED_RANGE_COL]].values
                    print(f"[proposals] Fallback: Using Excel recipes: {len(pareto_pts)} points")
        
        # Generate new proposals using EI + Euclidean distance + batch approach
        print(f"[proposals] Selecting recipes using EI + Euclidean distance + batch approach from {len(Xb)} candidates...")
        new_recipes_data = _build_and_write_proposals_ei_batch(Xb, mur, mug, sdr, sdg, pareto_pts, training_df)
        
        if new_recipes_data is None:
            print(f"[proposals] Error: Failed to generate new proposals")
            return False
        
        xlsx_path, sel_rows, new_recs, rates, ranges_nm = new_recipes_data
        
        # Get uncertainties from the selected rows
        rate_uncertainties = []
        range_uncertainties = []
        for row in sel_rows:
            rate_uncertainties.append(row["Rate_sigma"])
            range_uncertainties.append(row["Range_sigma_nm"])
        
        # Convert to the format expected by the Excel addition function
        next_recipes = []
        for i in range(min(POINTS_PER_ITERATION, len(new_recs))):
            recipe_data = {
                "O2_flow": float(new_recs[i, FEATURES.index("Etch_AvgO2Flow")]),
                "cf4_flow": float(new_recs[i, FEATURES.index("Etch_Avgcf4Flow")]),
                "Rf1_Pow": float(new_recs[i, FEATURES.index("Etch_Avg_Rf1_Pow")]),
                "Rf2_Pow": float(new_recs[i, FEATURES.index("Etch_Avg_Rf2_Pow")]),
                "Pressure": float(new_recs[i, FEATURES.index("Etch_AvgPres")]),
                "rate": float(rates[i]),
                "range_nm": float(ranges_nm[i]),
                "rate_uncertainty": float(rate_uncertainties[i]),
                "range_uncertainty": float(range_uncertainties[i])
            }
            next_recipes.append(recipe_data)
        
        print(f"[proposals] Successfully generated {len(next_recipes)} new recipes for iteration {next_iteration}")
        
        # Log uncertainty information
        for i, recipe in enumerate(next_recipes):
            print(f"[proposals] Recipe {i+1}: Rate={recipe['rate']:.1f} ± {recipe['rate_uncertainty']:.1f}, Range={recipe['range_nm']:.1f} ± {recipe['range_uncertainty']:.1f}")
        
        # Add new rows to the Excel file
        return _add_proposed_recipes_to_excel(recipes_df, next_recipes, next_iteration)
        
    except Exception as e:
        print(f"[proposals] Error generating new recipes: {e}")
        print(f"[proposals] Falling back to copying existing recipes...")
        
        # Fallback: use existing recipes if new generation fails
        if len(selected_points_with_features) >= POINTS_PER_ITERATION:
            next_recipes = selected_points_with_features[:POINTS_PER_ITERATION]
            return _add_proposed_recipes_to_excel(recipes_df, next_recipes, next_iteration)
        else:
            print(f"[proposals] Cannot propose recipes: insufficient fallback recipes available")
            return False

def _create_iteration_plots_with_excel_data(df_summary, current_front, selected_points, 
                                          recipes_df, iteration_num, iteration_status):
    """
    Create plots for a specific iteration using data from Excel file.
    This ensures visual consistency across iterations.
    """
    print(f"[iteration] Creating plots for iteration {iteration_num}")
    
    # Create iteration directory
    iter_dir = os.path.join(ITERATIONS_DIR, f"iteration_{iteration_num}")
    iter_plots_dir = os.path.join(iter_dir, "plots")
    os.makedirs(iter_dir, exist_ok=True)
    os.makedirs(iter_plots_dir, exist_ok=True)
    
    # For iteration 1, we want to show the baseline Pareto front with proposed recipes
    if iteration_num == 1:
        # Get proposed recipes for iteration 1 from Excel
        proposed_recipes = _get_proposed_recipes_for_iteration(recipes_df, 1)
        
        if proposed_recipes:
            # Convert to the format expected by plotting functions
            selected_points_for_plot = [(r["predicted_rate"], r["predicted_range"]) for r in proposed_recipes]
            
            # Extract uncertainties for the proposed recipes
            selected_uncertainties = []
            for r in proposed_recipes:
                rate_uncertainty = r.get("Pred_etch_rate_uncertainty", 0.0)
                range_uncertainty = r.get("Pred_range_uncertainty", 0.0)
                selected_uncertainties.append((rate_uncertainty, range_uncertainty))
            
            # Create Pareto front plot with proposed recipes and uncertainties
            front_plot = _plot_front(df_summary, current_front, selected_points_for_plot, [], [], recipes_df, iteration_num, iter_plots_dir, selected_uncertainties)
            
            # Create parity plots with iteration-specific LOOCV (no highlighting for iteration 1)
            print(f"[iteration] Running iteration-specific LOOCV for iteration {iteration_num}...")
            try:
                # Get the main dataset for LOOCV
                df_main = _load_dataset()
                loocv_results = _run_loocv_iteration_specific(df_main, iteration_num, recipes_df)
                
                # Save iteration-specific LOOCV results
                loocv_path = os.path.join(iter_dir, "loocv_predictions.csv")
                loocv_results.to_csv(loocv_path, index=False)
                print(f"[iteration] Saved iteration-specific LOOCV to {loocv_path}")
                
                # Create parity plots using iteration-specific data with highlights
                _plot_parity_from_loocv_with_highlights(loocv_path, [], recipes_df, iteration_num, iter_plots_dir)
                
            except Exception as e:
                print(f"[iteration] Error running iteration-specific LOOCV: {e}")
                # Fallback to global LOOCV if available
                loocv_csv = os.path.join(ROLLING_DIR, "loocv_predictions.csv")
                if os.path.exists(loocv_csv):
                    _plot_parity_from_loocv_with_highlights(loocv_csv, [], recipes_df, iteration_num, iter_plots_dir)
                else:
                    print(f"[iteration] Warning: No LOOCV data available, skipping parity plots for iteration {iteration_num}")
            
            # Show metrics up to the point before any highlighting begins
            if os.path.exists(os.path.join(ROLLING_DIR, "metrics_over_time.csv")):
                _plot_metrics_over_time(os.path.join(ROLLING_DIR, "metrics_over_time.csv"), 
                                      iteration_num, iter_plots_dir)
            
            return front_plot
        else:
            # No proposed recipes yet, create basic plot
            front_plot = _plot_front(df_summary, current_front, selected_points, [], [], recipes_df, iteration_num, iter_plots_dir)
            return front_plot
    
    # For other iterations, we need to show the predicted outcomes from the previous iteration
    previous_iteration = iteration_num - 1
    
    if previous_iteration not in iteration_status:
        print(f"[iteration] No data for previous iteration {previous_iteration}, skipping iteration {iteration_num}")
        return None
    
    # Get the proposed recipes from the previous iteration (these become predicted outcomes)
    previous_proposed_recipes = _get_proposed_recipes_for_iteration(recipes_df, previous_iteration)
    
    if not previous_proposed_recipes:
        print(f"[iteration] No proposed recipes found for previous iteration {previous_iteration}")
        return None
    
    # Separate completed vs pending recipes
    completed_recipes = [r for r in previous_proposed_recipes if isinstance(r.get("status"), str) and r["status"].lower() == "completed"]
    pending_recipes = [r for r in previous_proposed_recipes if isinstance(r.get("status"), str) and r["status"].lower() == "pending"]
    
    # Get lotnames for completed recipes (these will be highlighted)
    completed_lots = [r["lotname"] for r in completed_recipes if isinstance(r.get("lotname"), str) and r["lotname"]]
    
    # Stars (proposed recipes) for this iteration should come from THIS iteration's
    # proposed rows in Excel, not from the previous iteration. The previous
    # iteration's proposals become the predicted outcomes (crosses), which are
    # rendered via highlight_lots + Excel lookup inside _plot_front.
    stars_points = []
    current_iter_props = _get_proposed_recipes_for_iteration(recipes_df, iteration_num)
    if current_iter_props and iteration_status.get(previous_iteration, {}).get("is_completed", False):
        stars_points = [(r["predicted_rate"], r["predicted_range"]) for r in current_iter_props 
                       if isinstance(r.get("predicted_rate"), (int, float)) and isinstance(r.get("predicted_range"), (int, float))]

    # Extract uncertainties for the proposed recipes (stars)
    selected_uncertainties = []
    if current_iter_props:
        for r in current_iter_props:
            rate_uncertainty = r.get("Pred_etch_rate_uncertainty", 0.0)
            range_uncertainty = r.get("Pred_range_uncertainty", 0.0)
            selected_uncertainties.append((rate_uncertainty, range_uncertainty))
    
    # Create Pareto front plot highlighting the completed runs (crosses) and
    # optionally stars for the next 3 proposed recipes (if available) with uncertainties
    front_plot = _plot_front(df_summary, current_front, stars_points, 
                            completed_lots, [], recipes_df, iteration_num, iter_plots_dir, selected_uncertainties)
    
    # Create parity plots highlighting the completed runs
    print(f"[iteration] Running iteration-specific LOOCV for iteration {iteration_num}...")
    try:
        # Get the main dataset for LOOCV
        df_main = _load_dataset()
        loocv_results = _run_loocv_iteration_specific(df_main, iteration_num, recipes_df)
        
        # Save iteration-specific LOOCV results
        loocv_path = os.path.join(iter_dir, "loocv_predictions.csv")
        loocv_results.to_csv(loocv_path, index=False)
        print(f"[iteration] Saved iteration-specific LOOCV to {loocv_path}")
        
        # Create parity plots using iteration-specific data with highlights
        _plot_parity_from_loocv_with_highlights(loocv_path, completed_lots, 
                                               recipes_df, iteration_num, iter_plots_dir)
        
    except Exception as e:
        print(f"[iteration] Error running iteration-specific LOOCV: {e}")
        # Fallback to global LOOCV if available
        loocv_csv = os.path.join(ROLLING_DIR, "loocv_predictions.csv")
        if os.path.exists(loocv_csv):
            _plot_parity_from_loocv_with_highlights(loocv_csv, completed_lots, 
                                                   recipes_df, iteration_num, iter_plots_dir)
        else:
            print(f"[iteration] Warning: No LOOCV data available, skipping parity plots for iteration {iteration_num}")
    
    # Create metrics plots showing performance with the completed runs
    if os.path.exists(os.path.join(ROLLING_DIR, "metrics_over_time.csv")):
        _plot_metrics_over_time(os.path.join(ROLLING_DIR, "metrics_over_time.csv"), 
                              iteration_num, iter_plots_dir)
    
    return front_plot

def _save_iteration_data_with_excel_info(iteration_num, df_summary, current_front, 
                                       selected_points, iteration_status, recipes_df):
    """Save iteration data to cache, including information about Excel status"""
    iter_dir = os.path.join(ITERATIONS_DIR, f"iteration_{iteration_num}")
    iter_plots_dir = os.path.join(iter_dir, "plots")
    
    # Create directories only when needed
    os.makedirs(iter_dir, exist_ok=True)
    os.makedirs(iter_plots_dir, exist_ok=True)
    
    # Save summary data
    summary_path = os.path.join(iter_dir, "summary_data.csv")
    df_summary.to_csv(summary_path, index=False)
    
    # Save Pareto front
    front_path = os.path.join(iter_dir, "pareto_front.csv")
    current_front.to_csv(front_path, index=False)
    
    # Save selected points
    selected_df = pd.DataFrame(selected_points, columns=['AvgEtchRate', 'Range_nm'])
    selected_path = os.path.join(iter_dir, "selected_points.csv")
    selected_df.to_csv(selected_path, index=False)
    
    # Save iteration status information
    status_info = {
        "iteration_num": iteration_num,
        "excel_status": {
            "start_idx": iteration_status.get(iteration_num, {}).get("start_idx"),
            "end_idx": iteration_status.get(iteration_num, {}).get("end_idx"),
            "completed_count": iteration_status.get(iteration_num, {}).get("completed_count"),
            "pending_count": iteration_status.get(iteration_num, {}).get("pending_count"),
            "is_completed": iteration_status.get(iteration_num, {}).get("is_completed", False)
        },
        "has_proposed_recipes": iteration_num in iteration_status,
        "is_completed": iteration_status.get(iteration_num, {}).get("is_completed", False)
    }
    
    status_path = os.path.join(iter_dir, "iteration_status.json")
    with open(status_path, 'w') as f:
        json.dump(status_info, f, indent=2)
    
    # Save highlight lots (completed recipes for this iteration)
    if iteration_num in iteration_status:
        completed_recipes = iteration_status[iteration_num]["recipes"]
        completed_lots = completed_recipes[completed_recipes["Status_norm"] == "completed"][EXCEL_LOT_COL].tolist()
    else:
        completed_lots = []
    
    highlight_path = os.path.join(iter_dir, "highlight_lots.txt")
    with open(highlight_path, 'w') as f:
        for lot in completed_lots:
            f.write(f"{lot}\n")
    
    print(f"[iteration] Saved data for iteration {iteration_num}")

def _get_completed_lots_by_iteration(recipes_df):
    """Group completed lots by iteration based on completion order"""
    if recipes_df is None or EXCEL_STATUS_COL not in recipes_df.columns:
        return {}
    
    # Get completed lots sorted by completion date
    completed_df = recipes_df[recipes_df["Status_norm"] == "completed"].copy()
    if completed_df.empty:
        return {}
    
    # Sort by completion date
    if EXCEL_DATE_COL in completed_df.columns:
        completed_df = completed_df.sort_values(EXCEL_DATE_COL)
    
    # Group into iterations
    iterations = {}
    current_iteration = 0
    current_points = []
    
    for _, row in completed_df.iterrows():
        lot_name = row[EXCEL_LOT_COL]
        current_points.append(lot_name)
        
        if len(current_points) == POINTS_PER_ITERATION:
            iterations[current_iteration] = current_points.copy()
            current_iteration += 1
            current_points = []
    
    # Handle partial iteration (remaining points)
    if current_points:
        iterations[current_iteration] = current_points
    
    return iterations

def _get_cumulative_highlight_lots(iterations, target_iteration):
    """Get cumulative highlight lots up to a specific iteration"""
    if target_iteration < 0:
        return []
    
    highlight_lots = []
    for i in range(target_iteration + 1):
        if i in iterations:
            highlight_lots.extend(iterations[i])
    
    return highlight_lots

def _create_iteration_plots(df_summary, current_front, selected_points, 
                           recipes_df, iteration_num, highlight_lots):
    """Create all plots for a specific iteration"""
    print(f"[iteration] Creating plots for iteration {iteration_num}")
    
    # Create iteration directory
    iter_dir = os.path.join(ITERATIONS_DIR, f"iteration_{iteration_num}")
    iter_plots_dir = os.path.join(iter_dir, "plots")
    os.makedirs(iter_dir, exist_ok=True)
    os.makedirs(iter_plots_dir, exist_ok=True)
    
    # For iteration 1, we want to show the baseline Pareto front
    if iteration_num == 1:
        # Extract uncertainties for the selected points if available
        selected_uncertainties = []
        if recipes_df is not None:
            for i, (rate, range_nm) in enumerate(selected_points):
                # Try to find matching recipe in Excel
                rate_uncertainty = 0.0
                range_uncertainty = 0.0
                # For now, use default uncertainties - these would be calculated elsewhere
                selected_uncertainties.append((rate_uncertainty, range_uncertainty))
        
        # Create Pareto front plot with no highlighting but with uncertainties
        front_plot = _plot_front(df_summary, current_front, selected_points, [], [], recipes_df, iteration_num, iter_plots_dir, selected_uncertainties)
        
        # Create parity plots with no highlighting
        loocv_csv = os.path.join(ROLLING_DIR, "loocv_predictions.csv")
        if os.path.exists(loocv_csv):
            _plot_parity_from_loocv_with_highlights(loocv_csv, [], recipes_df, iteration_num, iter_plots_dir)
        else:
            print(f"[iteration] Warning: LOOCV file not found, skipping parity plots for iteration {iteration_num}")
        
        # For iteration 1, show metrics up to the point before any highlighting begins
        # This will be the baseline model performance
        if os.path.exists(os.path.join(ROLLING_DIR, "metrics_over_time.csv")):
            _plot_metrics_over_time(os.path.join(ROLLING_DIR, "metrics_over_time.csv"), 
                                  iteration_num, iter_plots_dir)
        
        return front_plot
    
    # For other iterations, we want to show ONLY the new points for this iteration
    # Get the new points for this specific iteration (not cumulative)
    new_points_for_this_iteration = highlight_lots
    
    # Extract uncertainties for the selected points if available
    selected_uncertainties = []
    if recipes_df is not None:
        for i, (rate, range_nm) in enumerate(selected_points):
            # Try to find matching recipe in Excel
            rate_uncertainty = 0.0
            range_uncertainty = 0.0
            # For now, use default uncertainties - these would be calculated elsewhere
            selected_uncertainties.append((rate_uncertainty, range_uncertainty))
    
    # Create Pareto front plot highlighting ONLY the new points with uncertainties
    front_plot = _plot_front(df_summary, current_front, selected_points, 
                            new_points_for_this_iteration, [], recipes_df, iteration_num, iter_plots_dir, selected_uncertainties)
    
    # Create parity plots highlighting ONLY the new points
    loocv_csv = os.path.join(ROLLING_DIR, "loocv_predictions.csv")
    if os.path.exists(loocv_csv):
        _plot_parity_from_loocv_with_highlights(loocv_csv, new_points_for_this_iteration, 
                                               recipes_df, iteration_num, iter_plots_dir)
    else:
        print(f"[iteration] Warning: LOOCV file not found, skipping parity plots for iteration {iteration_num}")
    
    # Create metrics plots showing performance improvement with the new points
    if os.path.exists(os.path.join(ROLLING_DIR, "metrics_over_time.csv")):
        _plot_metrics_over_time(os.path.join(ROLLING_DIR, "metrics_over_time.csv"), 
                              iteration_num, iter_plots_dir)
    
    return front_plot

def _create_iteration_metrics_plots(df_summary, highlight_lots, iteration_num, iter_plots_dir):
    """Create RMSE and uncertainty plots for a specific iteration's new points"""
    # Filter data to just this iteration's points
    iter_data = df_summary[df_summary["LOTNAME"].isin(highlight_lots)].copy()
    
    if len(iter_data) < 2:
        print(f"[iteration] Skipping metrics for iteration {iteration_num}: insufficient data ({len(iter_data)} points)")
        return
    
    # Calculate metrics for this iteration's points
    rate_rmse = np.sqrt(mean_squared_error(iter_data["AvgEtchRate"], iter_data["AvgEtchRate"]))
    range_rmse = np.sqrt(mean_squared_error(iter_data["Range_nm"], iter_data["Range_nm"]))
    
    # Create RMSE plot
    plt.figure(figsize=(10, 6))
    plt.bar(['Etch Rate', 'Thickness Range'], [rate_rmse, range_rmse], 
            color=['blue', 'orange'], alpha=0.7)
    plt.ylabel('RMSE')
    plt.title(f'RMSE for Iteration {iteration_num} ({len(iter_data)} points)')
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    rmse_plot_path = os.path.join(iter_plots_dir, "metrics_rmse.png")
    plt.savefig(rmse_plot_path, dpi=160)
    plt.close()
    
    # Create uncertainty plot (scatter plot with numeric color mapping)
    plt.figure(figsize=(10, 6))
    
    # Use numeric indices for colors instead of LOTNAME strings
    color_indices = np.arange(len(iter_data))
    scatter = plt.scatter(iter_data["AvgEtchRate"], iter_data["Range_nm"], 
                         c=color_indices, cmap='viridis', s=100, alpha=0.7)
    
    # Add lotname labels
    for i, (_, row) in enumerate(iter_data.iterrows()):
        plt.annotate(row["LOTNAME"], (row["AvgEtchRate"], row["Range_nm"]), 
                    xytext=(5, 5), textcoords='offset points', fontsize=8)
    
    plt.xlabel('Average Etch Rate')
    plt.ylabel('Thickness Range (nm)')
    plt.title(f'Iteration {iteration_num} Points ({len(iter_data)} points)')
    plt.colorbar(scatter, label='Point Index')
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    uncertainty_plot_path = os.path.join(iter_plots_dir, "uncertainty_plot.png")
    plt.savefig(uncertainty_plot_path, dpi=160)
    plt.close()
    
    print(f"[iteration] Created metrics plots for iteration {iteration_num}")

def _save_iteration_data(iteration_num, df_summary, current_front, selected_points, highlight_lots):
    """Save iteration data to cache"""
    iter_dir = os.path.join(ITERATIONS_DIR, f"iteration_{iteration_num}")
    iter_plots_dir = os.path.join(iter_dir, "plots")
    
    # Create directories only when needed
    os.makedirs(iter_dir, exist_ok=True)
    os.makedirs(iter_plots_dir, exist_ok=True)
    
    # Save summary data
    summary_path = os.path.join(iter_dir, "summary_data.csv")
    df_summary.to_csv(summary_path, index=False)
    
    # Save Pareto front
    front_path = os.path.join(iter_dir, "pareto_front.csv")
    current_front.to_csv(front_path, index=False)
    
    # Save selected points
    selected_df = pd.DataFrame(selected_points, columns=['AvgEtchRate', 'Range_nm'])
    selected_path = os.path.join(iter_dir, "selected_points.csv")
    selected_df.to_csv(selected_path, index=False)
    
    # Save highlight lots
    highlight_path = os.path.join(iter_dir, "highlight_lots.txt")
    with open(highlight_path, 'w') as f:
        for lot in highlight_lots:
            f.write(f"{lot}\n")
    
    print(f"[iteration] Saved data for iteration {iteration_num}")

def _load_iteration_data(iteration_num):
    """Load cached data for a specific iteration"""
    iter_dir = os.path.join(ITERATIONS_DIR, f"iteration_{iteration_num}")
    
    if not os.path.exists(iter_dir):
        return None, None, None, None
    
    try:
        summary_path = os.path.join(iter_dir, "summary_data.csv")
        front_path = os.path.join(iter_dir, "pareto_front.csv")
        selected_path = os.path.join(iter_dir, "selected_points.csv")
        highlight_path = os.path.join(iter_dir, "highlight_lots.txt")
        
        df_summary = pd.read_csv(summary_path) if os.path.exists(summary_path) else None
        current_front = pd.read_csv(front_path) if os.path.exists(front_path) else None
        selected_points = pd.read_csv(selected_path).values.tolist() if os.path.exists(selected_path) else None
        
        highlight_lots = []
        if os.path.exists(highlight_path):
            with open(highlight_path, 'r') as f:
                highlight_lots = [line.strip() for line in f.readlines()]
        
        return df_summary, current_front, selected_points, highlight_lots
    except Exception as e:
        print(f"[iteration] Error loading iteration {iteration_num}: {e}")
        return None, None, None, None

def _list_available_iterations():
    """List all available iterations"""
    if not os.path.exists(ITERATIONS_DIR):
        return []
    
    iterations = []
    for item in os.listdir(ITERATIONS_DIR):
        if item.startswith("iteration_") and os.path.isdir(os.path.join(ITERATIONS_DIR, item)):
            try:
                iter_num = int(item.split("_")[1])
                iterations.append(iter_num)
            except ValueError:
                continue
    
    return sorted(iterations)

def view_iteration(iteration_num):
    """View plots for a specific iteration"""
    if not os.path.exists(ITERATIONS_DIR):
        print("No iterations directory found")
        return
    
    iter_dir = os.path.join(ITERATIONS_DIR, f"iteration_{iteration_num}")
    if not os.path.exists(iter_dir):
        print(f"Iteration {iteration_num} not found")
        return
    
    iter_plots_dir = os.path.join(iter_dir, "plots")
    if not os.path.exists(iter_plots_dir):
        print(f"No plots directory found for iteration {iteration_num}")
        return
    
    # Display available plots
    available_plots = [f for f in os.listdir(iter_plots_dir) if f.endswith('.png')]
    if not available_plots:
        print(f"No plots found for iteration {iteration_num}")
        return
    
    print(f"Available plots for iteration {iteration_num}:")
    for plot in available_plots:
        print(f"  - {plot}")
    
    # Display plots
    for plot in available_plots:
        plot_path = os.path.join(iter_plots_dir, plot)
        print(f"\nDisplaying: {plot}")
        
        img = plt.imread(plot_path)
        plt.figure(figsize=(10, 8))
        plt.imshow(img)
        plt.axis('off')
        plt.title(f"{plot} - Iteration {iteration_num}")
        plt.tight_layout()
        plt.close()
    
    print(f"\nAll plots from iteration {iteration_num} have been displayed")

def list_iterations():
    """List all available iterations"""
    iterations = _list_available_iterations()
    if not iterations:
        print("No iterations found")
        return
    
    print("Available iterations:")
    for iter_num in iterations:
        iter_dir = os.path.join(ITERATIONS_DIR, f"iteration_{iter_num}")
        iter_plots_dir = os.path.join(iter_dir, "plots")
        
        # Count plots
        plot_count = 0
        if os.path.exists(iter_plots_dir):
            plot_count = len([f for f in os.listdir(iter_plots_dir) if f.endswith('.png')])
        
        print(f"  Iteration {iter_num}: {plot_count} plots")
        
        # Show highlight lots info
        highlight_path = os.path.join(iter_dir, "highlight_lots.txt")
        if os.path.exists(highlight_path):
            with open(highlight_path, 'r') as f:
                highlight_lots = [line.strip() for line in f.readlines()]
            print(f"    Highlight lots: {', '.join(highlight_lots) if highlight_lots else 'None'}")

def clear_iteration_cache():
    """Clear all iteration cache data"""
    if not os.path.exists(ITERATIONS_DIR):
        print("No iterations directory found")
        return
    
    import shutil
    shutil.rmtree(ITERATIONS_DIR)
    print("All iteration cache data cleared")

# ==================== MAIN ====================
def main():
    _ensure_dirs()
    df = _load_dataset()

    # -------- Read recipes Excel (optional, for highlights & gating) --------
    recipes = _read_recipes_excel()

    # Prepare summary + Pareto
    df_summary = df[["LOTNAME","FIMAP_FILE","AvgEtchRate","RangeEtchRate"]].copy()
    df_summary = df_summary[df_summary["AvgEtchRate"] <= 250].copy()
    df_summary["Range_nm"] = df_summary["RangeEtchRate"] * 5.0
    _atomic_write(os.path.join(SNAP_DIR, "fimap_rate_vs_range.csv"), df_summary[["LOTNAME","FIMAP_FILE","AvgEtchRate","Range_nm"]].to_csv(index=False).encode("utf-8"))

    # Hashes
    dataset_h  = _dataset_hash(df)
    features_h = _features_hash()
    model_h    = _model_config_hash()
    code_h     = _code_hash()
    cache_key_str = _cache_key(dataset_h, features_h, model_h, code_h)

    # ----- Ingestion gating (manual) or highlight logic (auto) -----
    cutoff_day = None
    pending_df = pd.DataFrame()
    highlight_lots_auto = set()
    new_completed_lots = set()  # New: track lots that need ingestion status update

    if recipes is not None:
        # Normalize expected columns if present
        if EXCEL_STATUS_COL in recipes.columns:
            recipes["Status_norm"] = recipes["Status_norm"] if "Status_norm" in recipes.columns else recipes[EXCEL_STATUS_COL].astype(str).str.strip().str.lower()
        if EXCEL_INGEST_COL in recipes.columns:
            recipes["Ingest_norm"] = recipes["Ingest_norm"] if "Ingest_norm" in recipes.columns else recipes[EXCEL_INGEST_COL].astype(str).str.strip().str.lower()
        if EXCEL_DATE_COL in recipes.columns:
            recipes[EXCEL_DATE_COL] = pd.to_datetime(recipes[EXCEL_DATE_COL], errors="coerce", utc=True)

        if INGEST_MODE == "auto":
            # latest completed lots (for highlight)
            if EXCEL_STATUS_COL in recipes.columns and EXCEL_DATE_COL in recipes.columns:
                completed_all = recipes[recipes["Status_norm"] == "completed"].copy()
                latest_date = completed_all[EXCEL_DATE_COL].max()
                if pd.notna(latest_date):
                    latest_completed = completed_all.loc[completed_all[EXCEL_DATE_COL] == latest_date].copy()
                    # form a set of lots that have both pred & actual in df_summary
                    pred_latest = latest_completed.rename(columns={EXCEL_LOT_COL:"LOTNAME"})
                    highlight_lots_auto = set(pred_latest["LOTNAME"].astype(str))
                    
                    # NEW: Identify lots that need ingestion status update
                    if EXCEL_INGEST_COL in recipes.columns:
                        for _, row in latest_completed.iterrows():
                            lot_name = row[EXCEL_LOT_COL]
                            ingest_status = row[EXCEL_INGEST_COL] if pd.notna(row[EXCEL_INGEST_COL]) else ""
                            if (lot_name in df_summary["LOTNAME"].values and 
                                (pd.isna(ingest_status) or ingest_status.strip().lower() in ["", "waiting"])):
                                new_completed_lots.add(lot_name)
            
            # pending list for plot (stars)
            if EXCEL_STATUS_COL in recipes.columns:
                pending_df = recipes[recipes["Status_norm"] == "pending"].copy()
        else:
            # MANUAL: Gate by latest approved ingestion status
            if (EXCEL_INGEST_COL in recipes.columns) and (EXCEL_DATE_COL in recipes.columns):
                approved = recipes.loc[recipes["Ingest_norm"] == "approved"].copy()
                approved = approved.dropna(subset=[EXCEL_DATE_COL]).sort_values(EXCEL_DATE_COL)
                if not approved.empty:
                    latest_approved = approved.iloc[-1]
                    cutoff_day = latest_approved[EXCEL_DATE_COL].normalize().tz_convert(None)
                    # prev_approved (may be NaN if only one)
                    prev_approved = approved.iloc[-2] if len(approved) >= 2 else None
                    print(f"[ingest/manual] cutoff at latest approved date = {cutoff_day.date()}")
                    if prev_approved is not None:
                        print(f"[ingest/manual] previous approved date = {prev_approved[EXCEL_DATE_COL].date()}")
                else:
                    print("[ingest/manual] No approved rows found; no gating applied.")

    # ----- Backtesting (with optional cutoff in manual mode) -----
    met_snap_path, backtest_path = _update_backtesting(df, cache_key_str, cutoff_day=cutoff_day)
    met_snap = _build_metrics_over_time(backtest_path, cache_key_str)
    
    # Debug info about metrics
    if met_snap:
        print(f"[main] Metrics built successfully: {met_snap}")
    else:
        print(f"[main] No metrics built, met_snap_path: {met_snap_path}")

    # ----- LOOCV (skips unless config changed) -----
    loocv_csv = _update_loocv(df, dataset_h, features_h, model_h, code_h)
    
    # Print model configuration for first two iterations
    if MODEL_CHOICE_FIRST_TWO_ITERATIONS == "rf_both":
        print(f"[config] First two iterations will use Random Forest for both etch rate and range prediction")
    else:
        print(f"[config] Using default models: ExtraTrees for etch rate, RandomForest for range prediction")
    
    # NEW: Use enhanced LOOCV plotting with highlights for new completed runs
    if new_completed_lots:
        _plot_parity_from_loocv_with_highlights(loocv_csv, new_completed_lots, recipes)
    elif recipes is not None:
        # Even if no new completed lots, use enhanced plotting for HIGHLIGHT_LOTS
        hl_union = list(set(HIGHLIGHT_LOTS) | set(highlight_lots_auto))
        if hl_union:
            _plot_parity_from_loocv_with_highlights(loocv_csv, hl_union, recipes)
        else:
            _plot_parity_from_loocv(loocv_csv)
    else:
        _plot_parity_from_loocv(loocv_csv)

    # ----- Pareto front history -----
    # Calculate Pareto front from complete dataset (not filtered summary)
    df_complete = df[["LOTNAME","FIMAP_FILE","AvgEtchRate","RangeEtchRate"]].copy()
    df_complete["Range_nm"] = df_complete["RangeEtchRate"] * 5.0
    
    front_ver, current_front = _update_pareto_history(df_complete)
    pareto_pts = current_front[["AvgEtchRate","Range_nm"]].values

    # ----- Fit models on iterative training data for proposals -----
    # Use iterative training data based on current iteration status
    training_df = _get_training_data_for_main_proposals(df, recipes)
    
    X = training_df[FEATURES].astype(float).values
    y_rate  = training_df["AvgEtchRate"].values
    y_range = training_df["RangeEtchRate"].values
    
    # Determine which model to use for etch rate based on current iteration
    iteration_status = _get_excel_iteration_status(recipes)
    if iteration_status:
        current_max_iteration = max(iteration_status.keys())
        next_iteration = current_max_iteration + 1
        
        if next_iteration <= 2:
            # Use Random Forest for iterations 1-2
            model_rate = _make_rf_rate().fit(X, y_rate)
            print(f"[main] Using Random Forest for etch rate (iteration {next_iteration})")
        else:
            # Use Extra Trees for iterations 3+
            model_rate = _make_extratrees().fit(X, y_rate)
            print(f"[main] Using Extra Trees for etch rate (iteration {next_iteration})")
    else:
        # Fallback to Extra Trees if no iteration info
        model_rate = _make_extratrees().fit(X, y_rate)
        print("[main] No iteration info, using Extra Trees for etch rate")
    
    # Always use Random Forest for range
    model_range = _make_rf().fit(X, y_range)
    print(f"[main] Using Random Forest for range (100 estimators)")

    # ----- Candidate sampling & predictions -----
    lower = np.array([FEATURE_RANGES[c][0] for c in FEATURES], float)
    upper = np.array([FEATURE_RANGES[c][1] for c in FEATURES], float)
    Xcand = _sample_candidates(SAMPLING_METHOD, N_SAMPLES, lower, upper, SAMPLING_SEED)
    Xcand = _quantize(Xcand, FEATURES)

    mu_r, sd_r = _pred_stats(model_rate, Xcand)
    mu_g, sd_g = _pred_stats(model_range, Xcand)
    mu_g_nm = mu_g * 5.0
    sd_g_nm = sd_g * 5.0

    mask = (mu_r >= TARGET_RATE_MIN) & (mu_r <= TARGET_RATE_MAX)
    if not np.any(mask):
        mask = np.ones(len(mu_r), dtype=bool)

    Xb = Xcand[mask]
    mur = mu_r[mask]
    mug = mu_g_nm[mask]
    sdr = sd_r[mask]
    sdg = sd_g_nm[mask]

    xlsx_path, sel_rows, new_recs, rates, ranges_nm = _build_and_write_proposals(Xb, mur, mug, sdr, sdg, pareto_pts)

    # ----- Front plot (note: we keep highlight list from HIGHLIGHT_LOTS plus auto-latest set) -----
    hl_union = list(set(HIGHLIGHT_LOTS) | set(highlight_lots_auto))
    
    # NEW: Use enhanced front plotting with new completed runs
    front_plot = _plot_front(df_summary, current_front, list(zip(rates, ranges_nm)), hl_union, 
                            new_completed_lots, recipes)
    
    # Use the rolling metrics file for plotting, not the snapshot copy
    met_plots = _plot_metrics_over_time(os.path.join(ROLLING_DIR, "metrics_over_time.csv")) if met_snap else None

        # ----- NEW: Iteration-based system -----
    if recipes is not None:
        print("[iteration] Processing iterations...")
        
        # First, analyze the Excel file to see what iterations already have proposed recipes
        iteration_status = _get_excel_iteration_status(recipes)
        print(f"[iteration] Excel analysis: {len(iteration_status)} iterations with proposed recipes")
        for iter_num, status in iteration_status.items():
            print(f"[iteration] Iteration {iter_num}: {status['completed_count']}/{POINTS_PER_ITERATION} completed")
        
        # Prepare selected points with feature values for potential recipe proposals
        print("[iteration] Preparing selected points with feature values for recipe proposals...")
        selected_points_with_features = []
        for i, (rate, range_nm) in enumerate(zip(rates, ranges_nm)):
            # Get the corresponding feature values from the selected recipes
            recipe_idx = sel_rows[i]["idx"] if i < len(sel_rows) else 0
            features = {
                "O2_flow": float(new_recs[i, FEATURES.index("Etch_AvgO2Flow")]),
                "cf4_flow": float(new_recs[i, FEATURES.index("Etch_Avgcf4Flow")]),
                "Rf1_Pow": float(new_recs[i, FEATURES.index("Etch_Avg_Rf1_Pow")]),
                "Rf2_Pow": float(new_recs[i, FEATURES.index("Etch_Avg_Rf2_Pow")]),
                "Pressure": float(new_recs[i, FEATURES.index("Etch_AvgPres")]),
                "rate": rate,
                "range_nm": range_nm
            }
            selected_points_with_features.append(features)
            print(f"[iteration] Recipe {i+1}: O2={features['O2_flow']:.1f}, cf4={features['cf4_flow']:.1f}, "
                  f"Rf1={features['Rf1_Pow']:.1f}, Rf2={features['Rf2_Pow']:.1f}, "
                  f"P={features['Pressure']:.1f} → Rate={features['rate']:.1f}, Range={features['range_nm']:.1f}")
        
        # Process each iteration based on Excel data
        max_iteration_in_excel = max(iteration_status.keys()) if iteration_status else -1
        
        # Calculate uncertainties for all available iterations
        if recipes is not None:
            print(f"[uncertainty] Calculating uncertainties for {max_iteration_in_excel + 1} iterations...")
            recipes_with_uncertainties = recipes.copy()
            
            # Calculate uncertainties for each iteration
            for iter_num in range(1, max_iteration_in_excel + 2):
                recipes_with_uncertainties = _calculate_uncertainties_for_iteration(iter_num, recipes_with_uncertainties)
            
            # Save the updated recipes with uncertainties back to the Excel file
            # Note: This would require write access to the Excel file
            print("[uncertainty] Uncertainties calculated and added to recipes DataFrame")
        
        for iteration_num in range(1, max_iteration_in_excel + 2):  # Start from iteration 1, +2 to include next iteration if needed
            print(f"[iteration] Processing iteration {iteration_num}...")
            
            # Create comprehensive plots for this iteration using the new 8-plot system
            # Get the proposed recipes for this specific iteration from Excel
            proposed_recipes_for_iteration = _get_proposed_recipes_for_iteration(recipes_with_uncertainties, iteration_num)
            
            if proposed_recipes_for_iteration:
                # Convert to the format expected by plotting functions
                selected_points_for_plot = [(r["predicted_rate"], r["predicted_range"]) for r in proposed_recipes_for_iteration]
            else:
                # Fallback to generated Pareto points if no Excel recipes exist
                selected_points_for_plot = list(zip(rates, ranges_nm))
            
            front_plot = _create_comprehensive_iteration_plots(iteration_num, df_summary, current_front, 
                                                             selected_points_for_plot, 
                                                             recipes_with_uncertainties, iteration_status)
            
            if front_plot is None:
                print(f"[iteration] Skipping iteration {iteration_num} - no data available")
                continue
            
            # Note: Previous iteration points are now handled directly in the comprehensive plotting system
            # No need for separate updates that were overwriting the main plots
            
            # Check if we should propose new recipes for the next iteration
            if _should_propose_new_recipes(recipes, iteration_num):
                if _propose_next_iteration_recipes(recipes, iteration_num, selected_points_with_features):
                    print(f"[iteration] Successfully proposed recipes for iteration {iteration_num + 1}")
                else:
                    print(f"[iteration] Failed to propose recipes for iteration {iteration_num + 1}")
            
            print(f"[iteration] Completed iteration {iteration_num}")
        
        # Handle iteration 5 (or next iteration) if no Excel data available
        next_iteration = max_iteration_in_excel + 1
        print(f"[iteration] Processing iteration {next_iteration}...")
        
        # Check if this iteration needs new recipes proposed
        if _should_propose_new_recipes(recipes, next_iteration - 1):
            print(f"[iteration] Generating new recipes for iteration {next_iteration}...")
            
            # Get the proper training data for this iteration
            training_df = _get_training_data_for_iteration(df, recipes, next_iteration)
            
            if training_df.empty:
                print(f"[iteration] Error: No training data available for iteration {next_iteration}")
                print(f"[iteration] Skipping iteration {next_iteration}")
            else:
                # Generate new recipes for this iteration
                if _propose_next_iteration_recipes(recipes, next_iteration - 1, selected_points_with_features):
                    print(f"[iteration] Successfully proposed recipes for iteration {next_iteration}")
                    
                    # Create the first 3 plots for iteration 5
                    print(f"[iteration] Creating first 3 plots for iteration {next_iteration}...")
                    
                    # Get the proposed recipes for this iteration
                    proposed_recipes_for_iteration = _get_proposed_recipes_for_iteration(recipes, next_iteration)
                    
                    if proposed_recipes_for_iteration:
                        # Convert to the format expected by plotting functions
                        selected_points_for_plot = [(r["predicted_rate"], r["predicted_range"]) for r in proposed_recipes_for_iteration]
                    else:
                        # Fallback to generated Pareto points if no Excel recipes exist
                        selected_points_for_plot = list(zip(rates, ranges_nm))
                    
                    # Create comprehensive plots for iteration 5
                    front_plot_iter5 = _create_comprehensive_iteration_plots(next_iteration, df_summary, current_front, 
                                                                         selected_points_for_plot, 
                                                                         recipes, iteration_status)
                    
                    if front_plot_iter5:
                        print(f"[iteration] Successfully created plots for iteration {next_iteration}")
                    else:
                        print(f"[iteration] Failed to create plots for iteration {next_iteration}")
                else:
                    print(f"[iteration] Failed to propose recipes for iteration {next_iteration}")
        
        print(f"[iteration] Completed iteration {next_iteration}")
    
    # ----- NEW: Update ingestion status for processed lots -----
    if new_completed_lots and recipes is not None:
        print(f"[main] Updating ingestion status for {len(new_completed_lots)} newly processed lots")
        _update_ingestion_status(recipes, new_completed_lots)

    manifest = _manifest_read()
    manifest.update({
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "code_version": CODE_VERSION,
        "ingest_mode": INGEST_MODE,
        "cutoff_day": cutoff_day.isoformat() if cutoff_day is not None else None,
        "new_completed_lots": list(new_completed_lots),  # NEW: track processed lots
        "hashes": {"dataset": dataset_h, "features": features_h, "model": model_h, "code": code_h, "cache_key": cache_key_str},
        "rolling": {
            "predictions_by_date": os.path.relpath(os.path.join(ROLLING_DIR, "predictions_by_date.csv"), ROOT_DIR),
            "loocv_predictions": os.path.relpath(_loocv_path(), ROOT_DIR),
            "pareto_front_history": os.path.relpath(_pareto_history_path(), ROOT_DIR),
            "metrics_over_time": os.path.relpath(os.path.join(ROLLING_DIR, "metrics_over_time.csv"), ROOT_DIR),
        },
        "snapshot": {
            "date": TODAY,
            "dir": os.path.relpath(SNAP_DIR, ROOT_DIR),
            "proposals_xlsx": os.path.relpath(xlsx_path, ROOT_DIR),
            "plots": {
                "front": os.path.relpath(front_plot, ROOT_DIR),
                "parity_rate": os.path.relpath(os.path.join(PLOTS_DIR, "parity_rate.png"), ROOT_DIR),
                "parity_range": os.path.relpath(os.path.join(PLOTS_DIR, "parity_range.png"), ROOT_DIR),
                "metrics_rmse": os.path.relpath(os.path.join(PLOTS_DIR, "metrics_rmse.png"), ROOT_DIR) if met_plots else None,
                "metrics_coverage": os.path.relpath(os.path.join(PLOTS_DIR, "metrics_coverage.png"), ROOT_DIR) if met_plots else None,
            }
        },
        "front_version": front_ver
    })
    _manifest_write(manifest)

    print("OK")
    print(json.dumps(manifest, indent=2))

def view_plots(target_date=None):
    """View plots from a specific date or the latest available date"""
    if target_date is None:
        # Interactive prompt for date
        print("\nAvailable snapshot dates:")
        if not os.path.exists(SNAPSHOTS_DIR):
            print("No snapshots directory found")
            return
        
        available_dates = [d for d in os.listdir(SNAPSHOTS_DIR) if os.path.isdir(os.path.join(SNAPSHOTS_DIR, d))]
        if not available_dates:
            print("No snapshots found")
            return
        
        available_dates.sort()
        for i, date in enumerate(available_dates, 1):
            print(f"  {i}. {date}")
        
        try:
            choice = input(f"\nEnter date number (1-{len(available_dates)}) or date (YYYY-MM-DD): ").strip()
            
            # Check if it's a number
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(available_dates):
                    target_date = available_dates[idx]
                else:
                    print("Invalid number selection")
                    return
            else:
                # Assume it's a date string
                target_date = choice
                
        except (KeyboardInterrupt, EOFError):
            print("\nCancelled")
            return
    
    _view_historical_plots(target_date)

def view_latest():
    """View the latest available plots"""
    latest_date = _get_latest_snapshot_date()
    if latest_date:
        print(f"Viewing latest plots from: {latest_date}")
        _view_historical_plots(latest_date)
    else:
        print("No snapshots found")

def _clear_cache():
    """Clear all cached data"""
    for f in os.listdir(ROLLING_DIR):
        if f.endswith('.csv'):
            os.remove(os.path.join(ROLLING_DIR, f))
    for f in os.listdir(PLOTS_DIR):
        if f.endswith('.png'):
            os.remove(os.path.join(PLOTS_DIR, f))
    print("Cache cleared")

def _clear_backtesting_cache():
    """Clear only the backtesting cache to force rebuild"""
    backtest_path = os.path.join(ROLLING_DIR, "predictions_by_date.csv")
    if os.path.exists(backtest_path):
        os.remove(backtest_path)
        print("Backtesting cache cleared - will rebuild on next run")
    else:
        print("No backtesting cache found")

def _find_last_table_row(ws):
    """Helper function to find the last row with actual table data"""
    last_table_row = 1  # Start with header row
    
    # First, find the column indices for key recipe columns
    key_columns = ["O2_flow", "cf4_flow", "Rf1_Pow", "Rf2_Pow", "Pressure"]
    column_indices = {}
    
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value in key_columns:
            column_indices[header_value] = col
    
    # If we can't find the key columns, fall back to checking first 8 columns
    if not column_indices:
        print("[proposals] Warning: Key recipe columns not found, falling back to first 8 columns")
        for col in range(1, min(8, ws.max_column + 1)):
            column_indices[f"col_{col}"] = col
    
    # Scan from bottom to top to find the last row with actual recipe data
    for row in range(ws.max_row, 1, -1):
        has_data = False
        
        # Check if this row has data in the key recipe columns
        for col_idx in column_indices.values():
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value is not None and str(cell_value).strip():  # Non-empty cell
                has_data = True
                break
        
        if has_data:
            last_table_row = row
            break  # Found the last row with data, stop searching
    
    print(f"[proposals] Last table row with data: {last_table_row}")
    return last_table_row

def _expand_excel_table(ws, new_row_count):
    """Expand Excel Table to include new rows"""
    try:
        # Check if worksheet has any tables
        if hasattr(ws, 'tables') and ws.tables:
            # Get the first table (assuming there's one main table)
            table = list(ws.tables.values())[0]
            print(f"[proposals] Found Excel Table: {table.name}")
            
            # Parse current table range (e.g., "A1:M10")
            current_range = table.ref
            print(f"[proposals] Current table range: {current_range}")
            
            # Extract start and end cells
            start_cell, end_cell = current_range.split(':')
            
            # Parse the end cell to get row number (e.g., "M10" -> row 10)
            import re
            match = re.match(r'([A-Z]+)(\d+)', end_cell)
            if match:
                end_col = match.group(1)
                current_end_row = int(match.group(2))
                
                # Calculate new end row
                new_end_row = current_end_row + new_row_count
                new_end_cell = f"{end_col}{new_end_row}"
                
                # Update table range
                new_range = f"{start_cell}:{new_end_cell}"
                table.ref = new_range
                
                print(f"[proposals] Expanded table range to: {new_range}")
                return True
        else:
            print("[proposals] No Excel Table found - adding as regular rows")
            return False
    except Exception as e:
        print(f"[proposals] Error expanding table: {e}")
        return False

def _cleanup_empty_rows(ws):
    """Remove any empty rows from the worksheet to keep table clean"""
    try:
        empty_rows = []
        
        # Find the last row with actual data first
        last_data_row = _find_last_table_row(ws)
        
        # Scan from bottom to top, but only up to the last data row
        for row in range(ws.max_row, last_data_row, -1):
            row_empty = True
            
            # Check if this row has any data
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    row_empty = False
                    break
            
            if row_empty:
                empty_rows.append(row)
        
        # Delete empty rows
        if empty_rows:
            print(f"[proposals] Found {len(empty_rows)} empty rows to clean up")
            for row in empty_rows:
                ws.delete_rows(row)
            print(f"[proposals] Cleaned up {len(empty_rows)} empty rows")
        
        return len(empty_rows)
    except Exception as e:
        print(f"[proposals] Error cleaning up empty rows: {e}")
        return 0

def _remove_gaps_between_data(ws):
    """Remove any gaps between data rows to ensure continuous data"""
    try:
        # Find the last row with actual data
        last_data_row = _find_last_table_row(ws)
        
        # Check for gaps between rows 2 and last_data_row
        rows_to_delete = []
        
        for row in range(2, last_data_row + 1):
            row_empty = True
            
            # Check if this row has any data in key recipe columns
            key_columns = ["O2_flow", "cf4_flow", "Rf1_Pow", "Rf2_Pow", "Pressure"]
            for col in range(1, ws.max_column + 1):
                header_value = ws.cell(row=1, column=col).value
                if header_value in key_columns:
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        row_empty = False
                        break
            
            if row_empty:
                rows_to_delete.append(row)
        
        # Delete empty rows from bottom to top to avoid index shifting
        if rows_to_delete:
            print(f"[proposals] Found {len(rows_to_delete)} gap rows to remove")
            for row in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row)
            print(f"[proposals] Removed {len(rows_to_delete)} gap rows")
        
        return len(rows_to_delete)
    except Exception as e:
        print(f"[proposals] Error removing gaps: {e}")
        return 0

def _add_proposed_recipes_to_excel(recipes_df, next_recipes, iteration_num):
    """
    Add new proposed recipes to the Excel file.
    Returns True if successful, False otherwise.
    """
    try:
        # Check if we have Graph API access
        if _HAS_MSAL and GRAPH_CLIENT_ID and GRAPH_CLIENT_SECRET and GRAPH_TENANT_ID:
            return _add_proposed_recipes_to_sharepoint(next_recipes, iteration_num)
        elif LOCAL_RECIPES_XLSX:
            return _add_proposed_recipes_to_local_excel(next_recipes, iteration_num)
        else:
            print("[proposals] No Excel access available (neither Graph API nor local file)")
            return False
    except Exception as e:
        print(f"[proposals] Error adding proposed recipes: {e}")
        return False

def _add_proposed_recipes_to_sharepoint(next_recipes, iteration_num):
    """Add proposed recipes to SharePoint Excel file"""
    try:
        print(f"[proposals] Adding {len(next_recipes)} recipes to SharePoint Excel for iteration {iteration_num}")
        
        # SharePoint-specific authentication and processing
        # Authenticate with Microsoft Graph
        authority = f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}"
        scopes = ["https://graph.microsoft.com/.default"]
        app = msal.ConfidentialClientApplication(
            GRAPH_CLIENT_ID, authority=authority, client_credential=GRAPH_CLIENT_SECRET
        )
        token_resp = app.acquire_token_for_client(scopes=scopes)
        if "access_token" not in token_resp:
            raise RuntimeError(f"MSAL token error: {token_resp}")
        token = token_resp["access_token"]
        headers = {"Authorization": f"Bearer {token}"}
        
        # Get site and drive information
        site_url = f"https://graph.microsoft.com/v1.0/sites/{GRAPH_TENANT_NAME}.sharepoint.com:/sites/{GRAPH_SITE_NAME}"
        site_id = requests.get(site_url, headers=headers).json()["id"]
        drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
        drives = requests.get(drives_url, headers=headers).json()["value"]
        drive_id = next(d["id"] for d in drives if d["name"] == "Documents")
        
        # Download current file
        download_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{RECIPES_FILE_PATH}:/content"
        resp = requests.get(download_url, headers=headers)
        resp.raise_for_status()
        
        # Load workbook and add new rows
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        
        # Step 1: Find ALL recipe_rejected rows to delete (SharePoint)
        recipe_rejected_rows = []
        ingest_status_col = None
        
        # Find the ingestion status column
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == EXCEL_INGEST_COL:
                ingest_status_col = col
                break
        
        if ingest_status_col:
            for row in range(2, ws.max_row + 1):  # Skip header row
                try:
                    ingest_status = ws.cell(row=row, column=ingest_status_col).value
                    if ingest_status and str(ingest_status).lower() == "recipe_rejected":
                        recipe_rejected_rows.append(row)
                except:
                    continue
        
        # Step 2: Delete ALL recipe_rejected rows completely
        deleted_count = 0
        if recipe_rejected_rows:
            print(f"[proposals] Found {len(recipe_rejected_rows)} recipe_rejected rows to delete")
            # Delete from bottom to top to avoid index shifting issues
            for row in sorted(recipe_rejected_rows, reverse=True):
                ws.delete_rows(row)
                deleted_count += 1
        
        # Step 3: Check if recipes for this iteration already exist (after deletion)
        existing_iteration_rows = []
        
        for row in range(2, ws.max_row + 1):  # Skip header row
            try:
                pred_rate_col = None
                pred_range_col = None
                ingest_status_col = None
                
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == EXCEL_PRED_RATE_COL:
                        pred_rate_col = col
                    if ws.cell(row=1, column=col).value == EXCEL_PRED_RANGE_COL:
                        pred_range_col = col
                    if ws.cell(row=1, column=col).value == EXCEL_INGEST_COL:
                        ingest_status_col = col
                
                if pred_rate_col and pred_range_col and ingest_status_col:
                    pred_rate = ws.cell(row=row, column=pred_rate_col).value
                    pred_range = ws.cell(row=row, column=pred_range_col).value
                    ingest_status = ws.cell(row=row, column=ingest_status_col).value
                    
                    # Check if this row has the same predicted values as the new recipes
                    if (pred_rate == float(next_recipes[0]["rate"]) and 
                        pred_range == float(next_recipes[0]["range_nm"])):
                        
                        if ingest_status and str(ingest_status).lower() in ["approved", "waiting"]:
                            # This is a valid existing recipe, don't add duplicates
                            existing_iteration_rows.append(row)
            except:
                continue
        
        if existing_iteration_rows:
            print(f"[proposals] Recipes for iteration {iteration_num} already exist (approved or waiting), skipping duplicate addition")
            return True
        
        # Step 4: Add new recipes right after the last table row (SharePoint)
        next_row = _find_last_table_row(ws) + 1
        
        # Verify that next_row is actually empty (safety check)
        while next_row <= ws.max_row:
            row_empty = True
            for col in range(1, min(8, ws.max_column + 1)):
                cell_value = ws.cell(row=next_row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    row_empty = False
                    break
            if row_empty:
                break
            next_row += 1
        
        print(f"[proposals] Adding recipes starting at row: {next_row}")
        
        # Step 5: Expand Excel Table to include new rows
        _expand_excel_table(ws, len(next_recipes))
        
        # Add new recipes
        for i, recipe_data in enumerate(next_recipes):
            # Use the actual feature values from the selected recipe
            # Note: We'll only add columns that exist in the Excel file
            row_data = {}
            
            # Try to add recipe parameters if columns exist
            if "O2_flow" in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data["O2_flow"] = float(recipe_data["O2_flow"])
            if "cf4_flow" in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data["cf4_flow"] = float(recipe_data["cf4_flow"])
            if "Rf1_Pow" in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data["Rf1_Pow"] = float(recipe_data["Rf1_Pow"])
            if "Rf2_Pow" in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data["Rf2_Pow"] = float(recipe_data["Rf2_Pow"])
            if "Pressure" in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data["Pressure"] = float(recipe_data["Pressure"])
            
            # Add the constant values that should always be present
            if "Chamber_temp" in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data["Chamber_temp"] = 50
            if "Electrode_temp" in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data["Electrode_temp"] = 15
            if "Etch_time" in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data["Etch_time"] = 5
            
            # Add predicted outcomes (these should exist)
            row_data[EXCEL_PRED_RATE_COL] = float(recipe_data["rate"])
            row_data[EXCEL_PRED_RANGE_COL] = float(recipe_data["range_nm"])
            
            # Add uncertainties if available
            if "rate_uncertainty" in recipe_data:
                row_data[EXCEL_RATE_UNCERTAINTY_COL] = float(recipe_data["rate_uncertainty"])
            if "range_uncertainty" in recipe_data:
                row_data[EXCEL_RANGE_UNCERTAINTY_COL] = float(recipe_data["range_uncertainty"])
            
            # Add status and metadata
            row_data[EXCEL_STATUS_COL] = "pending"
            if EXCEL_DATE_COL in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data[EXCEL_DATE_COL] = ""
            if EXCEL_LOT_COL in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data[EXCEL_LOT_COL] = ""
            if "IDRun" in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data["IDRun"] = ""
            if EXCEL_INGEST_COL in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data[EXCEL_INGEST_COL] = "waiting"
            if "Comment" in [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]:
                row_data["Comment"] = ""
            
            # Add the row to the worksheet
            for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                # Find the column index for this column name
                col_found = False
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == col_name:
                        ws.cell(row=next_row, column=col, value=value)
                        col_found = True
                        break
                
                if not col_found:
                    print(f"[proposals] Warning: Column '{col_name}' not found in Excel")
            
            next_row += 1
        
        # Step 6: Clean up any empty rows to keep table clean
        _cleanup_empty_rows(ws)
        
        # Step 7: Additional cleanup - ensure no gaps between data rows
        _remove_gaps_between_data(ws)
        
        # Step 8: Final verification - ensure recipes were added consecutively
        final_last_row = _find_last_table_row(ws)
        expected_last_row = next_row - 1 + len(next_recipes)
        if final_last_row != expected_last_row:
            print(f"[proposals] Warning: Expected last row {expected_last_row}, but found {final_last_row}")
            print(f"[proposals] This might indicate gaps in the data")
        
        # Save updated workbook
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        
        # Upload back to SharePoint
        upload_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{RECIPES_FILE_PATH}:/content"
        upload_headers = {**headers, "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        upload_resp = requests.put(upload_url, headers=upload_headers, data=bio.getvalue())
        upload_resp.raise_for_status()
        
        print(f"[proposals] Successfully added {len(next_recipes)} recipes to SharePoint Excel")
        return True
        
    except Exception as e:
        print(f"[proposals] Error updating SharePoint Excel: {e}")
        return False

def _add_proposed_recipes_to_local_excel(next_recipes, iteration_num):
    """Add proposed recipes to local Excel file"""
    try:
        print(f"[proposals] Adding {len(next_recipes)} recipes to local Excel for iteration {iteration_num}")
        
        # Local Excel-specific processing
        # Load existing workbook or create new one
        if os.path.exists(LOCAL_RECIPES_XLSX):
            wb = openpyxl.load_workbook(LOCAL_RECIPES_XLSX)
            ws = wb.active
        else:
            # Create new workbook with headers
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = [
                "Chamber", "Electrode", "Etch_time", "O2_flow", "cf4_flow", "Rf1_Pow", "Rf2_Pow", "Pressure",
                EXCEL_PRED_RATE_COL, EXCEL_PRED_RANGE_COL, EXCEL_RATE_UNCERTAINTY_COL, EXCEL_RANGE_UNCERTAINTY_COL,
                EXCEL_STATUS_COL, EXCEL_DATE_COL, EXCEL_LOT_COL, "IDRun", EXCEL_INGEST_COL, "Comment"
            ]
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
        
        # Step 1: Find ALL recipe_rejected rows to delete (Local Excel)
        recipe_rejected_rows = []
        ingest_status_col = None
        
        # Find the ingestion status column
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == EXCEL_INGEST_COL:
                ingest_status_col = col
                break
        
        if ingest_status_col:
            for row in range(2, ws.max_row + 1):  # Skip header row
                try:
                    ingest_status = ws.cell(row=row, column=ingest_status_col).value
                    if ingest_status and str(ingest_status).lower() == "recipe_rejected":
                        recipe_rejected_rows.append(row)
                except:
                    continue
        
        # Step 2: Delete ALL recipe_rejected rows completely
        deleted_count = 0
        if recipe_rejected_rows:
            print(f"[proposals] Found {len(recipe_rejected_rows)} recipe_rejected rows to delete")
            # Delete from bottom to top to avoid index shifting issues
            for row in sorted(recipe_rejected_rows, reverse=True):
                ws.delete_rows(row)
                deleted_count += 1
        
        # Step 3: Check if recipes for this iteration already exist (after deletion)
        existing_iteration_rows = []
        
        for row in range(2, ws.max_row + 1):  # Skip header row
            try:
                pred_rate_col = None
                pred_range_col = None
                ingest_status_col = None
                
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == EXCEL_PRED_RATE_COL:
                        pred_rate_col = col
                    if ws.cell(row=1, column=col).value == EXCEL_PRED_RANGE_COL:
                        pred_range_col = col
                    if ws.cell(row=1, column=col).value == EXCEL_INGEST_COL:
                        ingest_status_col = col
                
                if pred_rate_col and pred_range_col and ingest_status_col:
                    pred_rate = ws.cell(row=row, column=pred_rate_col).value
                    pred_range = ws.cell(row=row, column=pred_range_col).value
                    ingest_status = ws.cell(row=row, column=ingest_status_col).value
                    
                    # Check if this row has the same predicted values as the new recipes
                    if (pred_rate == float(next_recipes[0]["rate"]) and 
                        pred_range == float(next_recipes[0]["range_nm"])):
                        
                        if ingest_status and str(ingest_status).lower() in ["approved", "waiting"]:
                            # This is a valid existing recipe, don't add duplicates
                            existing_iteration_rows.append(row)
            except:
                continue
        
        if existing_iteration_rows:
            print(f"[proposals] Recipes for iteration {iteration_num} already exist (approved or waiting), skipping duplicate addition")
            return True
        
        # Step 4: Add new recipes right after the last table row (Local Excel)
        next_row = _find_last_table_row(ws) + 1
        
        # Verify that next_row is actually empty (safety check)
        while next_row <= ws.max_row:
            row_empty = True
            for col in range(1, min(8, ws.max_column + 1)):
                cell_value = ws.cell(row=next_row, column=col).value
                if cell_value is not None and str(cell_value).strip():
                    row_empty = False
                    break
            if row_empty:
                break
            next_row += 1
        
        print(f"[proposals] Adding recipes starting at row: {next_row}")
        
        # Step 5: Expand Excel Table to include new rows
        _expand_excel_table(ws, len(next_recipes))
        
        # Add new recipes
        for i, recipe_data in enumerate(next_recipes):
            # Use the actual feature values from the selected recipe
            row_data = {
                # Feature columns (using constants as specified)
                "Chamber": 50,
                "Electrode": 15,
                "Etch_time": 5,
                
                # Recipe parameters (using actual values from the selected recipe)
                "O2_flow": float(recipe_data["O2_flow"]),
                "cf4_flow": float(recipe_data["cf4_flow"]),
                "Rf1_Pow": float(recipe_data["Rf1_Pow"]),
                "Rf2_Pow": float(recipe_data["Rf2_Pow"]),
                "Pressure": float(recipe_data["Pressure"]),
                
                # Predicted outcomes
                EXCEL_PRED_RATE_COL: float(recipe_data["rate"]),
                EXCEL_PRED_RANGE_COL: float(recipe_data["range_nm"]),
                
                # Status and metadata
                EXCEL_STATUS_COL: "pending",
                EXCEL_DATE_COL: "",  # Empty as specified
                EXCEL_LOT_COL: "",   # Empty as specified
                "IDRun": "",         # Empty as specified
                EXCEL_INGEST_COL: "waiting", # Start as waiting for manual approval
                "Comment": ""        # Empty as specified
            }
            
            # Add the row to the worksheet
            for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                # Find the column index for this column name
                col_found = False
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == col_name:
                        ws.cell(row=next_row, column=col, value=value)
                        col_found = True
                        break
                
                if not col_found:
                    print(f"[proposals] Warning: Column '{col_name}' not found in Excel")
            
            next_row += 1
        
        # Step 6: Clean up any empty rows to keep table clean
        _cleanup_empty_rows(ws)
        
        # Step 7: Additional cleanup - ensure no gaps between data rows
        _remove_gaps_between_data(ws)
        
        # Step 8: Final verification - ensure recipes were added consecutively
        final_last_row = _find_last_table_row(ws)
        expected_last_row = next_row - 1 + len(next_recipes)
        if final_last_row != expected_last_row:
            print(f"[proposals] Warning: Expected last row {expected_last_row}, but found {final_last_row}")
            print(f"[proposals] This might indicate gaps in the data")
        
        # Save workbook
        wb.save(LOCAL_RECIPES_XLSX)
        print(f"[proposals] Successfully added {len(next_recipes)} recipes to local Excel")
        return True
        
    except Exception as e:
        print(f"[proposals] Error updating local Excel: {e}")
        return False

# ==================== COMPREHENSIVE ITERATION PLOTTING SYSTEM ====================
def _create_comprehensive_iteration_plots(iteration_num, df_summary, current_front, 
                                        selected_points, recipes_df, iteration_status):
    """
    Create all 9 plots for a specific iteration:
    1. Pareto front with 3 proposed recipes (⓿₁ symbols with uncertainty bars)
    2. Etch rate parity with horizontal lines for proposed points
    3. Range parity with horizontal lines for proposed points
    4. Pareto front with both predicted (faded) and actual (full opacity) points
    5. Etch rate parity with actual points instead of horizontal lines
    6. Range parity with actual points instead of horizontal lines
    7. Metrics RMSE over time
    8. Metrics coverage over time
    9. DEBUG: Pareto front with hypothetical completion (proposed recipes as actual data)
    """
    print(f"[comprehensive_plots] Creating all 9 plots for iteration {iteration_num}")
    
    # Create iteration directory
    iter_dir = os.path.join(ITERATIONS_DIR, f"iteration_{iteration_num}")
    iter_plots_dir = os.path.join(iter_dir, "plots")
    os.makedirs(iter_dir, exist_ok=True)
    os.makedirs(iter_plots_dir, exist_ok=True)
    
    # Get proposed recipes for this iteration
    proposed_recipes = _get_proposed_recipes_for_iteration(recipes_df, iteration_num)
    
    # Get completed recipes from previous iterations for training progression
    completed_recipes = _get_completed_recipes_up_to_iteration(recipes_df, iteration_num - 1)
    
    # Get highlight lots for this iteration (completed recipes)
    highlight_lots = _get_highlight_lots_for_iteration(recipes_df, iteration_num)
    
    # Extract uncertainties for the proposed recipes
    selected_uncertainties = []
    if proposed_recipes:
        for i, r in enumerate(proposed_recipes):
            # Get uncertainties from the recipe data
            rate_uncertainty = r.get("rate_uncertainty", 0.0)
            range_uncertainty = r.get("range_uncertainty", 0.0)
            
            # Fallback to old column names if new ones don't exist
            if rate_uncertainty == 0.0:
                rate_uncertainty = r.get("Pred_etch_rate_uncertainty", 0.0)
            if range_uncertainty == 0.0:
                range_uncertainty = r.get("Pred_range_uncertainty", 0.0)
                
            selected_uncertainties.append((rate_uncertainty, range_uncertainty))
    
    # Use the selected_points passed from main function (these are the Excel recipes)
    # No need to convert again since they're already in the right format
    selected_points_for_plot = selected_points
    
    # Plot 1: Pareto front with 3 proposed recipes (⓿₁ symbols with uncertainty bars)
    print(f"[comprehensive_plots] Creating Plot 1: Pareto front with proposed recipes")
    front_plot_1 = _plot_front_with_proposed_recipes(df_summary, current_front, 
                                                    selected_points_for_plot, selected_uncertainties, 
                                                    iteration_num, iter_plots_dir)
    
    # Plot 2 & 3: Parity plots with horizontal lines for proposed points
    print(f"[comprehensive_plots] Creating Plots 2 & 3: Parity plots with horizontal lines")
    parity_plots_23 = _plot_parity_with_horizontal_lines(iteration_num, recipes_df, 
                                                        selected_points_for_plot, selected_uncertainties,
                                                        highlight_lots, iter_plots_dir)
    
    # Plot 4: Pareto front with both predicted (faded) and actual (full opacity) points
    print(f"[comprehensive_plots] Creating Plot 4: Pareto front with predicted and actual")
    front_plot_4 = _plot_front_with_predicted_and_actual(df_summary, current_front, 
                                                        selected_points_for_plot, selected_uncertainties,
                                                        highlight_lots, recipes_df, iteration_num, 
                                                        iter_plots_dir)
    
    # Plot 5 & 6: Parity plots with actual points instead of horizontal lines
    print(f"[comprehensive_plots] Creating Plots 5 & 6: Parity plots with actual points")
    parity_plots_56 = _plot_parity_with_actual_points(iteration_num, recipes_df, 
                                                      selected_points_for_plot, selected_uncertainties,
                                                      highlight_lots, iter_plots_dir)
    
    # Plot 7 & 8: Metrics plots
    print(f"[comprehensive_plots] Creating Plots 7 & 8: Metrics plots")
    metrics_plots = _plot_metrics_for_iteration(iteration_num, iter_plots_dir)
    
    # Plot 9: DEBUG - Pareto front with hypothetical completion
    print(f"[comprehensive_plots] Creating Plot 9: Debug Pareto front with hypothetical completion")
    debug_plot_9 = _plot_debug_pareto_with_hypothetical_completion(df_summary, current_front, 
                                                                  selected_points_for_plot, 
                                                                  iteration_num, iter_plots_dir)
    
    # Save iteration data
    _save_iteration_data_with_excel_info(iteration_num, df_summary, current_front, 
                                        selected_points_for_plot, iteration_status, recipes_df)
    
    print(f"[comprehensive_plots] Successfully created all 9 plots for iteration {iteration_num}")
    return front_plot_1

def _plot_front_with_proposed_recipes(df_summary, current_front, selected_points, 
                                    selected_uncertainties, iteration_num, plots_dir):
    """Plot 1: Pareto front with 3 proposed recipes using ⓿₁ symbols with uncertainty bars"""
    plt.figure(figsize=(12, 10))
    
    # Plot historical data (no label needed)
    plt.scatter(df_summary["AvgEtchRate"], df_summary["Range_nm"], s=80, edgecolor='k', 
               alpha=0.6)
    
    # Plot Pareto front
    plt.plot(current_front["AvgEtchRate"], current_front["Range_nm"], 'r--', lw=2.5, 
             label="Pareto front")
    plt.scatter(current_front["AvgEtchRate"], current_front["Range_nm"], s=100, 
               facecolors='none', edgecolors='r', linewidth=1.5)
    
    # Plot proposed recipes with ⓿₁ symbols
    if selected_points and len(selected_points) >= 3:
        # Use consistent colors from existing system
        highlight_colors = HIGHLIGHT_COLORS
        
        for i, (rate, range_nm) in enumerate(selected_points[:3]):
            # Check if coordinates are valid numbers (not nan or inf)
            if (isinstance(rate, (int, float)) and isinstance(range_nm, (int, float)) and
                not np.isnan(rate) and not np.isnan(range_nm) and
                not np.isinf(rate) and not np.isinf(range_nm)):
                
                color = highlight_colors[i % len(highlight_colors)]
                
                # Create the symbol: outer number = point number, inner number = iteration number
                point_num = i + 1
                symbol = _get_symbol_for_point(point_num, iteration_num)

                
                # Plot only the Unicode symbol (no circle needed)
                # plt.scatter removed - only symbol text remains
                
                # Add the symbol text directly on the data point for better visibility
                plt.text(rate, range_nm, symbol, color=color, fontsize=32, weight='bold', 
                        ha='center', va='center', zorder=11,
                        path_effects=[pe.withStroke(linewidth=3, foreground="white")])
                
                # Add uncertainty bars if provided
                if selected_uncertainties and i < len(selected_uncertainties):
                    rate_uncertainty, range_uncertainty = selected_uncertainties[i]
                    
                    # Check if uncertainties are valid numbers (not nan or inf)
                    if (isinstance(rate_uncertainty, (int, float)) and 
                        isinstance(range_uncertainty, (int, float)) and
                        not np.isnan(rate_uncertainty) and not np.isnan(range_uncertainty) and
                        not np.isinf(rate_uncertainty) and not np.isinf(range_uncertainty) and
                        rate_uncertainty > 0 and range_uncertainty > 0):
                        

                        
                        # Vertical uncertainty bar for thickness range (y-axis) - more translucent
                        plt.errorbar(rate, range_nm, yerr=range_uncertainty, fmt='none', 
                                   color=color, alpha=0.3, capsize=8, capthick=3, elinewidth=3)
                        
                        # Horizontal uncertainty bar for etch rate (x-axis) - more translucent
                        plt.errorbar(rate, range_nm, xerr=rate_uncertainty, fmt='none', 
                                   color=color, alpha=0.3, capsize=8, capthick=3, elinewidth=3)

    
    # Set labels and title
    plt.xlabel("Average Etch Rate (nm/min)", fontsize=22, weight='bold')
    plt.ylabel("Thickness Range (nm)", fontsize=22, weight='bold')
    plt.title(f"Pareto Front with Proposed Recipes - Iteration {iteration_num}", 
             fontsize=24, weight='bold', pad=20)
    
    # Add custom legend explaining the numbers
    legend_elements = [
        plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='gray', 
                   markersize=15, label='Historical Data'),
        plt.Line2D([0], [0], color='red', linestyle='--', linewidth=2.5, label='Pareto Front'),
        plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='lightblue', 
                   markersize=15, label='Outside number = Recipe #, Inside number = Iteration #')
    ]
    
    # Legend and grid
    plt.legend(handles=legend_elements, loc="upper left", fontsize=16)
    plt.grid(True, alpha=0.4, linewidth=1.5)
    plt.ylim(0, 25)  # Set y-axis limit to 25 for thickness range
    plt.xticks(fontsize=18)
    plt.yticks(fontsize=18)
    plt.tight_layout()
    
    # Save plot
    plot_path = os.path.join(plots_dir, "1_pareto_proposed.png")
    plt.savefig(plot_path, dpi=160)
    plt.close()
    
    return plot_path

def _plot_parity_with_horizontal_lines(iteration_num, recipes_df, selected_points, 
                                     selected_uncertainties, highlight_lots, plots_dir):
    """Plots 2 & 3: Parity plots with horizontal lines for proposed points"""
    if not selected_points or len(selected_points) < 3:
        print(f"[comprehensive_plots] No selected points for parity plots in iteration {iteration_num}")
        return None, None
    
    # Get LOOCV data for this iteration
    loocv_data = _get_loocv_data_for_iteration(iteration_num, recipes_df)
    if loocv_data is None:
        print(f"[comprehensive_plots] No LOOCV data available for iteration {iteration_num}")
        return None, None
    
    # Use consistent colors
    highlight_colors = HIGHLIGHT_COLORS
    
    # Plot 2: Etch Rate Parity with horizontal lines
    plt.figure(figsize=(10, 7))
    
    # Plot historical LOOCV data (no label needed)
    plt.errorbar(loocv_data["loo_true_rate"], loocv_data["loo_pred_rate"], 
                yerr=loocv_data["loo_std_rate"], fmt='o', alpha=0.6, ecolor='gray', 
                capsize=3, markersize=8, linewidth=1)
    
    # Add x=y reference line (no label needed)
    min_val = min(loocv_data["loo_true_rate"].min(), loocv_data["loo_pred_rate"].min())
    max_val = max(loocv_data["loo_true_rate"].max(), loocv_data["loo_pred_rate"].max())
    plt.plot([max(0, min_val), max_val], [max(0, min_val), max_val], 'k--', alpha=0.5, linewidth=1.5)
    
    # Set better axis limits for cleaner scaling (less empty space)
    margin = (max_val - min_val) * 0.05
    plt.xlim(max(0, min_val - margin), max_val + margin)  # Ensure x-axis starts at 0 or positive
    plt.ylim(max(0, min_val - margin), max_val + margin)  # Ensure y-axis starts at 0 or positive
    
    # Add horizontal lines for proposed points
    for i, (rate, range_nm) in enumerate(selected_points[:3]):
        color = highlight_colors[i % len(highlight_colors)]
        point_num = i + 1
        
        symbol = _get_symbol_for_point(point_num, iteration_num)
        
        # Get predicted rate from selected points
        pred_rate = rate
        
        # Add horizontal line extending full width
        plt.axhline(y=pred_rate, color=color, linestyle='--', alpha=0.7, linewidth=2)
        
        # Add very translucent uncertainty lines above and below
        if selected_uncertainties and i < len(selected_uncertainties):
            rate_uncertainty = selected_uncertainties[i][0]
            
            # Check if uncertainty is a valid number
            if (isinstance(rate_uncertainty, (int, float)) and 
                not np.isnan(rate_uncertainty) and not np.isinf(rate_uncertainty) and
                rate_uncertainty > 0):
                
                # Upper uncertainty line (slightly more visible)
                plt.axhline(y=pred_rate + rate_uncertainty, color=color, linestyle='-', alpha=0.25, linewidth=1)
                # Lower uncertainty line (slightly more visible)
                plt.axhline(y=pred_rate - rate_uncertainty, color=color, linestyle='-', alpha=0.25, linewidth=1)
        
            # Add symbol label in the right whitespace (increased spacing from right edge)
        plt.text(plt.xlim()[1] * 0.90, pred_rate, symbol, color=color, 
                fontsize=32, weight='bold', ha='right', va='center',
                path_effects=[pe.withStroke(linewidth=3, foreground="white")])
    
    # Calculate R² and RMSE scores using ONLY baseline historical data (not experimental points)
    # This ensures fair comparison across iterations
    baseline_mask = ~loocv_data["LOTNAME"].isin(highlight_lots) if highlight_lots else pd.Series([True] * len(loocv_data))
    baseline_data = loocv_data[baseline_mask]
    
    if len(baseline_data) > 0:
        r2 = r2_score(baseline_data["loo_true_rate"], baseline_data["loo_pred_rate"])
        rmse = np.sqrt(mean_squared_error(baseline_data["loo_true_rate"], baseline_data["loo_pred_rate"]))
    else:
        r2 = np.nan
        rmse = np.nan
    
    # Add clean legend with R² and RMSE inside the plot (like legend in Pareto front)
    # Use same positioning logic as Pareto front legend for neat top-left placement
    plt.text(0.02, 0.98, f"R² = {r2:.2f}\nRMSE = {rmse:.2f}", 
             fontsize=18, weight='bold', transform=plt.gca().transAxes,
             bbox=dict(boxstyle="round,pad=0.3", facecolor='white', alpha=0.9, edgecolor='black', linewidth=1),
             verticalalignment='top')
    
    # Set plot properties
    plt.xlabel("Actual AvgEtchRate (nm/min)", fontsize=22, weight='bold')
    plt.ylabel("Predicted AvgEtchRate (nm/min)", fontsize=22, weight='bold')
    plt.title(f"Etch Rate Parity with Proposed Points - Iteration {iteration_num}", 
             fontsize=20, weight='bold', pad=25)
    plt.grid(True, alpha=0.1, linewidth=0.5)
    plt.xticks(fontsize=18)
    plt.yticks(fontsize=18)
    plt.gca().set_aspect(1/1.7, adjustable='box')  # Make internal plot rectangular (1:1.7 ratio - height >> width)
    plt.tight_layout(pad=2.0)  # Increased padding for better margins
    
    # Save plot
    rate_plot_path = os.path.join(plots_dir, "2_parity_rate_horizontal.png")
    plt.savefig(rate_plot_path, dpi=160)
    plt.close()
    
    # Plot 3: Range Parity with horizontal lines
    plt.figure(figsize=(10, 7))
    
    # Plot historical LOOCV data (convert to thickness range, no label needed)
    plt.errorbar(loocv_data["loo_true_range"] * 5.0, loocv_data["loo_pred_range"] * 5.0, 
                yerr=loocv_data["loo_std_range"] * 5.0, fmt='o', alpha=0.6, ecolor='gray', 
                capsize=2, markersize=6, linewidth=0.5)
    
    # Add x=y reference line (no label needed)
    min_val = min((loocv_data["loo_true_range"] * 5.0).min(), (loocv_data["loo_pred_range"] * 5.0).min())
    max_val = max((loocv_data["loo_true_range"] * 5.0).max(), (loocv_data["loo_pred_range"] * 5.0).max())
    plt.plot([max(0, min_val), max_val], [max(0, min_val), max_val], 'k--', alpha=0.5, linewidth=1.5)
    
    # Add horizontal lines for proposed points
    for i, (rate, range_nm) in enumerate(selected_points[:3]):
        color = highlight_colors[i % len(highlight_colors)]
        point_num = i + 1
        
        symbol = _get_symbol_for_point(point_num, iteration_num)
        
        # Get predicted range from selected points (already in nm)
        pred_range = range_nm
        
        # Add horizontal line extending full width
        plt.axhline(y=pred_range, color=color, linestyle='--', alpha=0.7, linewidth=2)
        
        # Add very translucent uncertainty lines above and below
        if selected_uncertainties and i < len(selected_uncertainties):
            range_uncertainty = selected_uncertainties[i][1]
            
            # Check if uncertainty is a valid number
            if (isinstance(range_uncertainty, (int, float)) and 
                not np.isnan(range_uncertainty) and not np.isinf(range_uncertainty) and
                range_uncertainty > 0):
                
                # Upper uncertainty line (slightly more visible)
                plt.axhline(y=pred_range + range_uncertainty, color=color, linestyle='-', alpha=0.25, linewidth=1)
                # Lower uncertainty line (slightly more visible)
                plt.axhline(y=pred_range - range_uncertainty, color=color, linestyle='-', alpha=0.25, linewidth=1)
        
        # Add symbol label in the right whitespace (increased spacing from right edge)
        plt.text(plt.xlim()[1] * 0.90, pred_range, symbol, color=color, 
                fontsize=32, weight='bold', ha='right', va='center',
                path_effects=[pe.withStroke(linewidth=3, foreground="white")])
    
    # Calculate R² and RMSE scores using ONLY baseline historical data (not experimental points)
    # This ensures fair comparison across iterations
    baseline_mask = ~loocv_data["LOTNAME"].isin(highlight_lots) if highlight_lots else pd.Series([True] * len(loocv_data))
    baseline_data = loocv_data[baseline_mask]
    
    if len(baseline_data) > 0:
        r2 = r2_score(baseline_data["loo_true_range"] * 5.0, baseline_data["loo_pred_range"] * 5.0)
        rmse = np.sqrt(mean_squared_error(baseline_data["loo_true_range"] * 5.0, baseline_data["loo_pred_range"] * 5.0))
    else:
        r2 = np.nan
        rmse = np.nan
    
    # Add clean legend with R² and RMSE inside the plot (like legend in Pareto front)
    # Use same positioning logic as Pareto front legend for neat top-left placement
    plt.text(0.02, 0.98, f"R² = {r2:.2f}\nRMSE = {rmse:.2f}", 
             fontsize=18, weight='bold', transform=plt.gca().transAxes,
             bbox=dict(boxstyle="round,pad=0.3", facecolor='white', alpha=0.9, edgecolor='black', linewidth=1),
             verticalalignment='top')
    
    # Set plot properties
    plt.xlabel("Actual Thickness Range (nm)", fontsize=22, weight='bold')
    plt.ylabel("Predicted Thickness Range (nm)", fontsize=22, weight='bold')
    plt.title(f"Thickness Range Parity with Proposed Points - Iteration {iteration_num}", 
             fontsize=20, weight='bold', pad=25)
    plt.grid(True, alpha=0.1, linewidth=0.5)
    plt.xticks(fontsize=18)
    plt.gca().set_aspect(1/1.7, adjustable='box')  # Make internal plot rectangular (1:1.7 ratio - height >> width)
    plt.tight_layout(pad=2.0)  # Increased padding for better margins
    
    # Save plot
    range_plot_path = os.path.join(plots_dir, "3_parity_range_horizontal.png")
    plt.savefig(range_plot_path, dpi=160)
    plt.close()
    
    return rate_plot_path, range_plot_path

def _plot_front_with_predicted_and_actual(df_summary, current_front, selected_points, 
                                        selected_uncertainties, highlight_lots, recipes_df, 
                                        iteration_num, plots_dir):
    """Plot 4: Pareto front with both predicted (faded) and actual (full opacity) points"""
    plt.figure(figsize=(12, 10))
    
    # Plot ALL historical data (always show the full dataset)
    plt.scatter(df_summary["AvgEtchRate"], df_summary["Range_nm"], s=80, edgecolor='k', 
               alpha=0.6)
    
    # Plot Pareto front
    plt.plot(current_front["AvgEtchRate"], current_front["Range_nm"], 'r--', lw=2.5, 
             label="Pareto front")
    plt.scatter(current_front["AvgEtchRate"], current_front["Range_nm"], s=100, 
               facecolors='none', edgecolors='r', linewidth=1.5)
    
    # Plot predicted points (faded/translucent)
    if selected_points and len(selected_points) >= 3:
        highlight_colors = HIGHLIGHT_COLORS
        
        for i, (rate, range_nm) in enumerate(selected_points[:3]):
            # Check if coordinates are valid numbers (not nan, inf, or string)
            if (isinstance(rate, (int, float)) and isinstance(range_nm, (int, float)) and
                not np.isnan(rate) and not np.isnan(range_nm) and
                not np.isinf(rate) and not np.isinf(range_nm)):
                
                color = highlight_colors[i % len(highlight_colors)]
                point_num = i + 1
                
                symbol = _get_symbol_for_point(point_num, iteration_num)
                
                # Plot only the Unicode symbol for predicted point (faded)
                # plt.scatter removed - only symbol text remains
                
                # Add symbol text with path effects for better visibility (more faded)
                plt.text(rate, range_nm, symbol, color=color, fontsize=32, weight='bold', 
                        ha='center', va='center', zorder=11, alpha=0.4,
                        path_effects=[pe.withStroke(linewidth=3, foreground="white")])
                
                # Add uncertainty bars (faded) if valid
                if selected_uncertainties and i < len(selected_uncertainties):
                    rate_uncertainty, range_uncertainty = selected_uncertainties[i]
                    
                    # Check if uncertainties are valid numbers
                    if (isinstance(rate_uncertainty, (int, float)) and 
                        isinstance(range_uncertainty, (int, float)) and
                        not np.isnan(rate_uncertainty) and not np.isnan(range_uncertainty) and
                        not np.isinf(rate_uncertainty) and not np.isinf(range_uncertainty) and
                        rate_uncertainty > 0 and range_uncertainty > 0):
                        
                        plt.errorbar(rate, range_nm, yerr=range_uncertainty, fmt='none', 
                                   color=color, alpha=0.3, capsize=5, capthick=2, elinewidth=2)
                        plt.errorbar(rate, range_nm, xerr=rate_uncertainty, fmt='none', 
                                   color=color, alpha=0.3, capsize=5, capthick=2, elinewidth=2)

    
    # Plot actual points (full opacity) if available
    if highlight_lots and recipes_df is not None:
        # Separate current iteration points from previous iteration points
        current_iter_lots = []
        previous_iter_lots = []
        
        # Get current iteration lots
        current_iter_recipes = _get_proposed_recipes_for_iteration(recipes_df, iteration_num)
        if current_iter_recipes:
            for recipe in current_iter_recipes:
                status = recipe.get("status", "")
                if isinstance(status, str) and status.lower() == "completed":
                    lotname = recipe.get("lotname", "")
                    if lotname and isinstance(lotname, str) and lotname.strip():
                        current_iter_lots.append(lotname)
        
        # Get previous iteration lots
        if iteration_num > 1:
            previous_iter_recipes = _get_proposed_recipes_for_iteration(recipes_df, iteration_num - 1)
            if previous_iter_recipes:
                for recipe in previous_iter_recipes:
                    status = recipe.get("status", "")
                    if isinstance(status, str) and status.lower() == "completed":
                        lotname = recipe.get("lotname", "")
                        if lotname and isinstance(lotname, str) and lotname.strip():
                            previous_iter_lots.append(lotname)
        
        # Plot current iteration points (full opacity, full size, colored)
        for i, lot in enumerate(current_iter_lots):
            if lot in df_summary["LOTNAME"].values:
                row = df_summary[df_summary["LOTNAME"] == lot]
                actual_rate = row["AvgEtchRate"].iloc[0]
                actual_range = row["Range_nm"].iloc[0]
                
                # Get the correct point number for this lot
                point_num = i + 1
                symbol = _get_symbol_for_point(point_num, iteration_num)
                
                # Use highlight colors for current iteration
                color = highlight_colors[i % len(highlight_colors)]
                plt.text(actual_rate, actual_range, symbol, color=color, fontsize=32, weight='bold', 
                        ha='center', va='center', zorder=11,
                        path_effects=[pe.withStroke(linewidth=3, foreground="white")])
        
        # Plot previous iteration points (faded, smaller, black)
        for i, lot in enumerate(previous_iter_lots):
            if lot in df_summary["LOTNAME"].values:
                row = df_summary[df_summary["LOTNAME"] == lot]
                actual_rate = row["AvgEtchRate"].iloc[0]
                actual_range = row["Range_nm"].iloc[0]
                
                # Get the correct point number for this lot
                point_num = i + 1
                symbol = _get_symbol_for_point(point_num, iteration_num - 1)
                
                # Use black color, smaller size, and faded for previous iteration
                plt.text(actual_rate, actual_range, symbol, color='black', fontsize=24, weight='bold', 
                        ha='center', va='center', zorder=10, alpha=0.6,
                        path_effects=[pe.withStroke(linewidth=2, foreground="white")])
    
    # Set labels and title
    plt.xlabel("Average Etch Rate (nm/min)", fontsize=22, weight='bold')
    plt.ylabel("Thickness Range (nm)", fontsize=22, weight='bold')
    plt.title(f"Pareto Front with Predicted and Actual Points - Iteration {iteration_num}", 
             fontsize=24, weight='bold', pad=20)
    
    # Legend and grid
    plt.legend(loc="upper left", fontsize=20)
    plt.grid(True, alpha=0.4, linewidth=1.5)
    plt.ylim(0, 25)  # Set y-axis limit to 25 for thickness range
    plt.xticks(fontsize=18)
    plt.yticks(fontsize=18)
    plt.tight_layout()
    
    # Save plot
    plot_path = os.path.join(plots_dir, "4_pareto_predicted_actual.png")
    plt.savefig(plot_path, dpi=160)
    plt.close()
    
    return plot_path

def _plot_parity_with_actual_points(iteration_num, recipes_df, selected_points, 
                                   selected_uncertainties, highlight_lots, plots_dir):
    """Plots 5 & 6: Parity plots with actual points instead of horizontal lines"""
    if not selected_points or len(selected_points) < 3:
        print(f"[comprehensive_plots] No selected points for parity plots in iteration {iteration_num}")
        return None, None
    
    # Get LOOCV data for this iteration
    loocv_data = _get_loocv_data_for_iteration(iteration_num, recipes_df)
    if loocv_data is None:
        print(f"[comprehensive_plots] No LOOCV data available for iteration {iteration_num}")
        return None, None
    
    # Use consistent colors
    highlight_colors = HIGHLIGHT_COLORS
    
    # Plot 5: Etch Rate Parity with actual points
    plt.figure(figsize=(10, 7))
    
    # Plot historical LOOCV data (no label needed)
    plt.errorbar(loocv_data["loo_true_rate"], loocv_data["loo_pred_rate"], 
                yerr=loocv_data["loo_std_rate"], fmt='o', alpha=0.6, ecolor='gray', 
                capsize=3, markersize=8, linewidth=1)
    
    # Add x=y reference line (no label needed)
    min_val = min(loocv_data["loo_true_rate"].min(), loocv_data["loo_pred_rate"].min())
    max_val = max(loocv_data["loo_true_rate"].max(), loocv_data["loo_pred_rate"].max())
    plt.plot([min_val, max_val], [min_val, max_val], 'k--', alpha=0.5, linewidth=1.5)
    
    # Add actual points for highlighted lots
    if highlight_lots and recipes_df is not None:
        # Separate current iteration points from previous iteration points
        current_iter_lots = []
        previous_iter_lots = []
        
        # Get current iteration lots
        current_iter_recipes = _get_proposed_recipes_for_iteration(recipes_df, iteration_num)
        if current_iter_recipes:
            for recipe in current_iter_recipes:
                status = recipe.get("status", "")
                if isinstance(status, str) and status.lower() == "completed":
                    lotname = recipe.get("lotname", "")
                    if lotname and isinstance(lotname, str) and lotname.strip():
                        current_iter_lots.append(lotname)
        
        # Get previous iteration lots
        if iteration_num > 1:
            previous_iter_recipes = _get_proposed_recipes_for_iteration(recipes_df, iteration_num - 1)
            if previous_iter_recipes:
                for recipe in previous_iter_recipes:
                    status = recipe.get("status", "")
                    if isinstance(status, str) and status.lower() == "completed":
                        lotname = recipe.get("lotname", "")
                        if lotname and isinstance(lotname, str) and lotname.strip():
                            previous_iter_lots.append(lotname)
        
        # Plot current iteration points (full opacity, full size, colored)
        for i, lot in enumerate(current_iter_lots):
            if lot in loocv_data["LOTNAME"].values:
                lot_mask = loocv_data["LOTNAME"] == lot
                actual_rate = loocv_data.loc[lot_mask, "loo_true_rate"].iloc[0]
                pred_rate = loocv_data.loc[lot_mask, "loo_pred_rate"].iloc[0]
                
                # Get the correct point number for this lot
                point_num = i + 1
                symbol = _get_symbol_for_point(point_num, iteration_num)
                
                # Use highlight colors for current iteration
                color = highlight_colors[i % len(highlight_colors)]
                plt.text(actual_rate, pred_rate, symbol, color=color, fontsize=32, 
                        weight='bold', ha='center', va='center', zorder=6,
                        path_effects=[pe.withStroke(linewidth=3, foreground="white")])
        
        # Plot previous iteration points (faded, smaller, black)
        for i, lot in enumerate(previous_iter_lots):
            if lot in loocv_data["LOTNAME"].values:
                lot_mask = loocv_data["LOTNAME"] == lot
                actual_rate = loocv_data.loc[lot_mask, "loo_true_rate"].iloc[0]
                pred_rate = loocv_data.loc[lot_mask, "loo_pred_rate"].iloc[0]
                
                # Get the correct point number for this lot
                point_num = i + 1
                symbol = _get_symbol_for_point(point_num, iteration_num - 1)
                
                # Use black color, smaller size, and faded for previous iteration
                plt.text(actual_rate, pred_rate, symbol, color='black', fontsize=24, weight='bold', 
                        ha='center', va='center', zorder=5, alpha=0.6,
                        path_effects=[pe.withStroke(linewidth=2, foreground="white")])
    
    # Calculate R² and RMSE scores using ONLY baseline historical data (not experimental points)
    # This ensures fair comparison across iterations
    baseline_mask = ~loocv_data["LOTNAME"].isin(highlight_lots) if highlight_lots else pd.Series([True] * len(loocv_data))
    baseline_data = loocv_data[baseline_mask]
    
    if len(baseline_data) > 0:
        r2 = r2_score(baseline_data["loo_true_rate"], baseline_data["loo_pred_rate"])
        rmse = np.sqrt(mean_squared_error(baseline_data["loo_true_rate"], baseline_data["loo_pred_rate"]))
    else:
        r2 = np.nan
        rmse = np.nan
    
    # Add clean legend with R² and RMSE inside the plot (like legend in Pareto front)
    # Use same positioning logic as Pareto front legend for neat top-left placement
    plt.text(0.02, 0.98, f"R² = {r2:.2f}\nRMSE = {rmse:.2f}", 
             fontsize=18, weight='bold', transform=plt.gca().transAxes,
             bbox=dict(boxstyle="round,pad=0.3", facecolor='white', alpha=0.9, edgecolor='black', linewidth=1),
             verticalalignment='top')
    
    # Set plot properties
    plt.xlabel("Actual AvgEtchRate (nm/min)", fontsize=22, weight='bold')
    plt.ylabel("Predicted AvgEtchRate (nm/min)", fontsize=22, weight='bold')
    plt.title(f"Etch Rate Parity with Actual Points - Iteration {iteration_num}", 
             fontsize=20, weight='bold', pad=25)
    plt.grid(True, alpha=0.1, linewidth=0.5)
    plt.xticks(fontsize=18)
    plt.yticks(fontsize=18)
    plt.gca().set_aspect(1/1.7, adjustable='box')  # Make internal plot rectangular (1:1.7 ratio - height >> width)
    plt.tight_layout(pad=2.0)  # Increased padding for better margins
    
    # Save plot
    rate_plot_path = os.path.join(plots_dir, "5_parity_rate_actual.png")
    plt.savefig(rate_plot_path, dpi=160)
    plt.close()
    
    # Plot 6: Range Parity with actual points
    plt.figure(figsize=(10, 7))
    
    # Plot historical LOOCV data (convert to thickness range, no label needed)
    plt.errorbar(loocv_data["loo_true_range"] * 5.0, loocv_data["loo_pred_range"] * 5.0, 
                yerr=loocv_data["loo_std_range"] * 5.0, fmt='o', alpha=0.6, ecolor='gray', 
                capsize=2, markersize=6, linewidth=0.5)
    
    # Add x=y reference line (no label needed)
    min_val = min((loocv_data["loo_true_range"] * 5.0).min(), (loocv_data["loo_pred_range"] * 5.0).min())
    max_val = max((loocv_data["loo_true_range"] * 5.0).max(), (loocv_data["loo_pred_range"] * 5.0).max())
    plt.plot([max(0, min_val), max_val], [max(0, min_val), max_val], 'k--', alpha=0.5, linewidth=1.5)
    
    # Add actual points for highlighted lots
    if highlight_lots and recipes_df is not None:
        # Separate current iteration points from previous iteration points
        current_iter_lots = []
        previous_iter_lots = []
        
        # Get current iteration lots
        current_iter_recipes = _get_proposed_recipes_for_iteration(recipes_df, iteration_num)
        if current_iter_recipes:
            for recipe in current_iter_recipes:
                status = recipe.get("status", "")
                if isinstance(status, str) and status.lower() == "completed":
                    lotname = recipe.get("lotname", "")
                    if lotname and isinstance(lotname, str) and lotname.strip():
                        current_iter_lots.append(lotname)
        
        # Get previous iteration lots
        if iteration_num > 1:
            previous_iter_recipes = _get_proposed_recipes_for_iteration(recipes_df, iteration_num - 1)
            if previous_iter_recipes:
                for recipe in previous_iter_recipes:
                    status = recipe.get("status", "")
                    if isinstance(status, str) and status.lower() == "completed":
                        lotname = recipe.get("lotname", "")
                        if lotname and isinstance(lotname, str) and lotname.strip():
                            previous_iter_lots.append(lotname)
        
        # Plot current iteration points (full opacity, full size, colored)
        for i, lot in enumerate(current_iter_lots):
            if lot in loocv_data["LOTNAME"].values:
                lot_mask = loocv_data["LOTNAME"] == lot
                actual_range = loocv_data.loc[lot_mask, "loo_true_range"].iloc[0] * 5.0
                pred_range = loocv_data.loc[lot_mask, "loo_pred_range"].iloc[0] * 5.0
                
                # Get the correct point number for this lot
                point_num = i + 1
                symbol = _get_symbol_for_point(point_num, iteration_num)
                
                # Use highlight colors for current iteration
                color = highlight_colors[i % len(highlight_colors)]
                plt.text(actual_range, pred_range, symbol, color=color, fontsize=32, 
                        weight='bold', ha='center', va='center', zorder=6,
                        path_effects=[pe.withStroke(linewidth=3, foreground="white")])
        
        # Plot previous iteration points (faded, smaller, black)
        for i, lot in enumerate(previous_iter_lots):
            if lot in loocv_data["LOTNAME"].values:
                lot_mask = loocv_data["LOTNAME"] == lot
                actual_range = loocv_data.loc[lot_mask, "loo_true_range"].iloc[0] * 5.0
                pred_range = loocv_data.loc[lot_mask, "loo_pred_range"].iloc[0] * 5.0
                
                # Get the correct point number for this lot
                point_num = i + 1
                symbol = _get_symbol_for_point(point_num, iteration_num - 1)
                
                # Use black color, smaller size, and faded for previous iteration
                plt.text(actual_range, pred_range, symbol, color='black', fontsize=24, weight='bold', 
                        ha='center', va='center', zorder=5, alpha=0.6,
                        path_effects=[pe.withStroke(linewidth=2, foreground="white")])
    
    # Calculate R² and RMSE scores using ONLY baseline historical data (not experimental points)
    # This ensures fair comparison across iterations
    baseline_mask = ~loocv_data["LOTNAME"].isin(highlight_lots) if highlight_lots else pd.Series([True] * len(loocv_data))
    baseline_data = loocv_data[baseline_mask]
    
    if len(baseline_data) > 0:
        r2 = r2_score(baseline_data["loo_true_range"] * 5.0, baseline_data["loo_pred_range"] * 5.0)
        rmse = np.sqrt(mean_squared_error(baseline_data["loo_true_range"] * 5.0, baseline_data["loo_pred_range"] * 5.0))
    else:
        r2 = np.nan
        rmse = np.nan
    
    # Add clean legend with R² and RMSE inside the plot (like legend in Pareto front)
    # Use same positioning logic as Pareto front legend for neat top-left placement
    plt.text(0.02, 0.98, f"R² = {r2:.2f}\nRMSE = {rmse:.2f}", 
             fontsize=18, weight='bold', transform=plt.gca().transAxes,
             bbox=dict(boxstyle="round,pad=0.3", facecolor='white', alpha=0.9, edgecolor='black', linewidth=1),
             verticalalignment='top')
    
    # Set plot properties
    plt.xlabel("Actual Thickness Range (nm)", fontsize=22, weight='bold')
    plt.ylabel("Predicted Thickness Range (nm)", fontsize=22, weight='bold')
    plt.title(f"Thickness Range Parity with Actual Points - Iteration {iteration_num}", 
             fontsize=20, weight='bold', pad=25)
    plt.grid(True, alpha=0.1, linewidth=0.5)
    plt.xticks(fontsize=18)
    plt.yticks(fontsize=18)
    plt.gca().set_aspect(1/1.7, adjustable='box')  # Make internal plot rectangular (1:1.7 ratio - height >> width)
    plt.tight_layout(pad=2.0)  # Increased padding for better margins
    
    # Save plot
    range_plot_path = os.path.join(plots_dir, "6_parity_range_actual.png")
    plt.savefig(range_plot_path, dpi=160)
    plt.close()
    
    return rate_plot_path, range_plot_path

def _plot_metrics_for_iteration(iteration_num, plots_dir):
    """Plots 7 & 8: Metrics plots for the iteration"""
    # Check if metrics data exists
    metrics_path = os.path.join(ROLLING_DIR, "metrics_over_time.csv")
    if not os.path.exists(metrics_path):
        print(f"[comprehensive_plots] No metrics data available for iteration {iteration_num}")
        return None, None
    
    # Create metrics plots using existing function
    try:
        metrics_plots = _plot_metrics_over_time(metrics_path, iteration_num, plots_dir)
        return metrics_plots
    except Exception as e:
        print(f"[comprehensive_plots] Error creating metrics plots: {e}")
        return None, None

def _plot_debug_pareto_with_hypothetical_completion(df_summary, current_front, selected_points, 
                                                   iteration_num, plots_dir):
    """
    Plot 9: DEBUG - Pareto front with hypothetical completion
    
    This plot shows what the Pareto front would look like if the proposed recipes
    were perfect predictions and became part of the actual dataset. This helps
    validate if the Pareto logic is working correctly.
    
    Args:
        df_summary: Historical data summary
        current_front: Current Pareto front
        selected_points: Proposed recipes for this iteration
        iteration_num: Current iteration number
        plots_dir: Directory to save the plot
    """
    plt.figure(figsize=(12, 10))
    
    # Plot historical data (light blue, no label)
    plt.scatter(df_summary["AvgEtchRate"], df_summary["Range_nm"], s=80, edgecolor='k', 
               alpha=0.6, color='lightblue', label="Historical data")
    
    # Plot current Pareto front (red dashed line)
    plt.plot(current_front["AvgEtchRate"], current_front["Range_nm"], 'r--', lw=2.5, 
             label="Current Pareto front")
    plt.scatter(current_front["AvgEtchRate"], current_front["Range_nm"], s=100, 
               color='red', edgecolor='k', alpha=0.8, zorder=3)
    
    # Create hypothetical dataset by adding proposed recipes to historical data
    if selected_points and len(selected_points) > 0:
        # Extract proposed recipe data
        proposed_rates = []
        proposed_ranges = []
        
        for i, recipe in enumerate(selected_points):
            # Handle both tuple format (rate, range) and dict format
            if isinstance(recipe, tuple):
                rate, range_nm = recipe
            else:
                rate = recipe.get("Pred_avg_etch_rate", recipe.get("rate", 0))
                range_nm = recipe.get("Range_nm", recipe.get("range_nm", 0))
            
            if rate > 0 and range_nm > 0:
                proposed_rates.append(rate)
                proposed_ranges.append(range_nm)
        
        if proposed_rates:
            # Plot proposed recipes as if they were actual data (large, colored circles)
            colors = ['purple', 'orange', 'green']
            for i, (rate, range_nm) in enumerate(zip(proposed_rates, proposed_ranges)):
                color = colors[i % len(colors)]
                symbol = _get_symbol_for_point(i + 1, iteration_num)
                
                plt.scatter(rate, range_nm, s=200, color=color, edgecolor='k', 
                           alpha=0.9, zorder=5, label=f"Proposed recipe {i+1}")
                
                # Add symbol annotation
                plt.text(rate, range_nm, symbol, color='white', fontsize=20, 
                        weight='bold', ha='center', va='center', zorder=6,
                        path_effects=[pe.withStroke(linewidth=2, foreground="black")])
            
            print(f"[debug_plot] Processing {len(proposed_rates)} proposed recipes for hypothetical Pareto front")
            
            # Calculate NEW Pareto front including proposed recipes
            # Create combined dataset
            combined_data = df_summary[["AvgEtchRate", "Range_nm"]].copy()
            proposed_data = pd.DataFrame({
                "AvgEtchRate": proposed_rates,
                "Range_nm": proposed_ranges
            })
            combined_data = pd.concat([combined_data, proposed_data], ignore_index=True)
            
            print(f"[debug_plot] Combined dataset: {len(combined_data)} total points (historical + proposed)")
            
            # Calculate new Pareto front
            new_front = _pareto_front(combined_data)
            
            print(f"[debug_plot] New Pareto front has {len(new_front)} points")
            
            # Plot NEW Pareto front (blue dashed line, thicker)
            plt.plot(new_front["AvgEtchRate"], new_front["Range_nm"], 'b--', lw=3.5, 
                     label="Hypothetical Pareto front (with proposed recipes)")
            plt.scatter(new_front["AvgEtchRate"], new_front["Range_nm"], s=120, 
                       color='blue', edgecolor='k', alpha=0.9, zorder=4)
            
            # Add improvement analysis
            improvement_text = []
            for i, (rate, range_nm) in enumerate(zip(proposed_rates, proposed_ranges)):
                # Check if this point dominates any existing Pareto front points
                dominates = _dominates_existing((rate, range_nm), 
                                             current_front[["AvgEtchRate", "Range_nm"]].values)
                if dominates:
                    improvement_text.append(f"Recipe {i+1}: PUSHES front forward ✅")
                else:
                    improvement_text.append(f"Recipe {i+1}: Fills gap only ⚠️")
            
            # Add improvement summary to plot
            if improvement_text:
                improvement_str = "\n".join(improvement_text)
                plt.text(0.02, 0.02, f"Improvement Analysis:\n{improvement_str}", 
                         fontsize=14, weight='bold', transform=plt.gca().transAxes,
                         bbox=dict(boxstyle="round,pad=0.3", facecolor='lightyellow', 
                                  alpha=0.9, edgecolor='black', linewidth=1),
                         verticalalignment='bottom')
    
    # Set plot properties
    plt.xlabel("Average Etch Rate (nm/min)", fontsize=22, weight='bold')
    plt.ylabel("Thickness Range (nm)", fontsize=22, weight='bold')
    plt.title(f"DEBUG: Pareto Front with Hypothetical Completion - Iteration {iteration_num}", 
             fontsize=20, weight='bold', pad=25)
    plt.grid(True, alpha=0.3, linewidth=0.5)
    plt.xticks(fontsize=18)
    plt.yticks(fontsize=18)
    plt.legend(fontsize=16, loc='upper left', bbox_to_anchor=(0.02, 0.98))
    plt.tight_layout(pad=2.0)
    
    # Save plot
    debug_plot_path = os.path.join(plots_dir, "9_debug_pareto_hypothetical.png")
    plt.savefig(debug_plot_path, dpi=160)
    plt.close()
    
    print(f"[comprehensive_plots] Created debug plot: {debug_plot_path}")
    return debug_plot_path

def _get_symbol_for_point(point_num, iteration_num):
    """Generate proper Unicode symbol for a point in an iteration
    
    Uses the correct logic: circled[iteration-1] + subs[point-1]
    - point_num: 1-based point number (1, 2, 3)
    - iteration_num: 1-based iteration number (1, 2, 3, ...)
    """
    # Convert to 0-based indices for array access
    iter_idx = iteration_num - 1
    point_idx = point_num - 1
    
    # Safety checks
    if iter_idx < 0 or iter_idx >= len(CIRCLED_NUMBERS):
        return f"?{iteration_num}"
    if point_idx < 0 or point_idx >= len(SUBSCRIPT_NUMBERS):
        return f"{CIRCLED_NUMBERS[iter_idx]}?"
    
    return CIRCLED_NUMBERS[iter_idx] + SUBSCRIPT_NUMBERS[point_idx]

def _get_loocv_data_for_iteration(iteration_num, recipes_df):
    """Get LOOCV data for a specific iteration with progressive training"""
    try:
        # For iteration 1, use iteration-specific LOOCV if available, otherwise global
        iter_dir = os.path.join(ITERATIONS_DIR, f"iteration_{iteration_num}")
        loocv_path = os.path.join(iter_dir, "loocv_predictions.csv")
        
        if os.path.exists(loocv_path):
            print(f"[comprehensive_plots] Using iteration-specific LOOCV for iteration {iteration_num}")
            return pd.read_csv(loocv_path)
        else:
            # Fallback to global LOOCV
            global_loocv_path = os.path.join(ROLLING_DIR, "loocv_predictions.csv")
            if os.path.exists(global_loocv_path):
                print(f"[comprehensive_plots] Using global LOOCV for iteration {iteration_num}")
                return pd.read_csv(global_loocv_path)
        
        return None
    except Exception as e:
        print(f"[comprehensive_plots] Error loading LOOCV data: {e}")
        return None

def _get_completed_recipes_up_to_iteration(recipes_df, max_iteration):
    """Get all completed recipes up to a specific iteration for training progression"""
    if recipes_df is None or max_iteration < 1:
        return []
    
    completed_recipes = []
    for iter_num in range(1, max_iteration + 1):
        iter_recipes = _get_proposed_recipes_for_iteration(recipes_df, iter_num)
        if iter_recipes:
            for recipe in iter_recipes:
                status = recipe.get("status", "")
                # Handle both string and numeric status values
                if isinstance(status, str) and status.lower() == "completed":
                    completed_recipes.append(recipe)
    
    return completed_recipes

def _get_highlight_lots_for_iteration(recipes_df, iteration_num):
    """
    Get lotnames for highlighting in plots for a specific iteration.
    
    For iterations 2+, we show:
    - Current iteration: Full opacity, full size
    - Previous iteration only: Faded, smaller size
    - Remove all older iterations: Don't show iteration 1 in iteration 3 plots
    """
    if iteration_num == 1:
        # For iteration 1, only show current iteration points
        highlight_lots = []
        iter_recipes = _get_proposed_recipes_for_iteration(recipes_df, iteration_num)
        if iter_recipes:
            for recipe in iter_recipes:
                status = recipe.get("status", "")
                # Handle both string and numeric status values
                if isinstance(status, str) and status.lower() == "completed":
                    lotname = recipe.get("lotname", "")
                    if lotname and isinstance(lotname, str) and lotname.strip():
                        highlight_lots.append(lotname)
        
        print(f"[highlight] Iteration {iteration_num}: Found {len(highlight_lots)} completed lots for highlighting")
        return highlight_lots
    
    # For iterations 2+, show ONLY current iteration + previous iteration
    highlight_lots = []
    
    # Get current iteration points
    current_iter_recipes = _get_proposed_recipes_for_iteration(recipes_df, iteration_num)
    if current_iter_recipes:
        for recipe in current_iter_recipes:
            status = recipe.get("status", "")
            if isinstance(status, str) and status.lower() == "completed":
                lotname = recipe.get("lotname", "")
                if lotname and isinstance(lotname, str) and lotname.strip():
                    highlight_lots.append(lotname)
    
    # Get previous iteration points only (not all iterations)
    previous_iter_recipes = _get_proposed_recipes_for_iteration(recipes_df, iteration_num - 1)
    if previous_iter_recipes:
        for recipe in previous_iter_recipes:
            status = recipe.get("status", "")
            if isinstance(status, str) and status.lower() == "completed":
                lotname = recipe.get("lotname", "")
                if lotname and isinstance(lotname, str) and lotname.strip():
                    highlight_lots.append(lotname)
    
            print(f"[highlight] Iteration {iteration_num}: Found {len(highlight_lots)} completed lots for highlighting "
              f"({len([r for r in current_iter_recipes if str(r.get('status', '')).lower() == 'completed'])} current + "
              f"{len([r for r in previous_iter_recipes if str(r.get('status', '')).lower() == 'completed'])} previous)")
    return highlight_lots

# REMOVED: This function was part of the broken secondary plotting system
# The main comprehensive plotting system handles previous iteration points correctly

# REMOVED: These functions were overwriting the main plots with broken versions
# The main comprehensive plotting system already handles previous iteration points correctly
# and preserves all the formatting, uncertainty bars, and historical data

def regenerate_all_iterations():
    """Regenerate all iterations with the new comprehensive 8-plot system"""
    print("[regenerate] Starting regeneration of all iterations with comprehensive plotting system...")
    
    # Load required data
    try:
        df = _load_dataset()
        df_summary = df[["LOTNAME","FIMAP_FILE","AvgEtchRate","RangeEtchRate"]].copy()
        df_summary = df_summary[df_summary["AvgEtchRate"] <= 250].copy()
        df_summary["Range_nm"] = df_summary["RangeEtchRate"] * 5.0
        
        # Calculate Pareto front
        df_complete = df[["LOTNAME","FIMAP_FILE","AvgEtchRate","RangeEtchRate"]].copy()
        df_complete["Range_nm"] = df_complete["RangeEtchRate"] * 5.0
        current_front = _pareto_front(df_complete)
        
        # Read recipes Excel
        recipes = _read_recipes_excel()
        if recipes is None:
            print("[regenerate] No recipes Excel available, cannot regenerate iterations")
            return
        
        # Get iteration status
        iteration_status = _get_excel_iteration_status(recipes)
        if not iteration_status:
            print("[regenerate] No iterations found in Excel, cannot regenerate")
            return
        
        print(f"[regenerate] Found {len(iteration_status)} iterations to regenerate")
        
        # Generate fresh proposals for each iteration
        for iteration_num in range(1, max(iteration_status.keys()) + 2):
            print(f"[regenerate] Regenerating iteration {iteration_num}...")
            
            # Generate fresh proposals for this iteration using iterative training
            selected_points = _generate_fresh_proposals_for_iteration(iteration_num, df, current_front, recipes)
            
            if selected_points is None:
                print(f"[regenerate] Skipping iteration {iteration_num} - no proposals generated")
                continue
            
            # Generate iteration-specific LOOCV data first
            print(f"[regenerate] Generating LOOCV data for iteration {iteration_num}...")
            loocv_results = _run_loocv_iteration_specific(df, iteration_num, recipes)
            if loocv_results is not None:
                # Save LOOCV data to iteration directory
                iter_dir = os.path.join(ITERATIONS_DIR, f"iteration_{iteration_num}")
                os.makedirs(iter_dir, exist_ok=True)
                loocv_path = os.path.join(iter_dir, "loocv_predictions.csv")
                loocv_results.to_csv(loocv_path, index=False)
                print(f"[regenerate] Saved LOOCV data for iteration {iteration_num}")
            
            # Create comprehensive plots for this iteration
            front_plot = _create_comprehensive_iteration_plots(iteration_num, df_summary, current_front, 
                                                             selected_points, recipes, iteration_status)
            
            if front_plot is None:
                print(f"[regenerate] Failed to create plots for iteration {iteration_num}")
                continue
            
            # Note: Previous iteration points are now handled directly in the comprehensive plotting system
            # No need for separate updates that were overwriting the main plots
            
            print(f"[regenerate] Successfully regenerated iteration {iteration_num}")
        
        print("[regenerate] All iterations regenerated successfully!")
        
    except Exception as e:
        print(f"[regenerate] Error during regeneration: {e}")
        import traceback
        traceback.print_exc()

def _generate_fresh_proposals_for_iteration(iteration_num, df, current_front, recipes_df=None):
    """Generate fresh proposals for a specific iteration using iterative training"""
    try:
        # Get the proper training data for this iteration
        if recipes_df is not None:
            training_df = _get_training_data_for_iteration(df, recipes_df, iteration_num)
            print(f"[regenerate] Iteration {iteration_num}: Using {len(training_df)} training points "
                  f"(run_date < cutoff date for iteration {iteration_num})")
        else:
            training_df = df.copy()
            print(f"[regenerate] Iteration {iteration_num}: No recipes available, using full dataset")
        
        # Prepare features and targets from training data
        X = training_df[FEATURES].astype(float).values
        y_rate = training_df["AvgEtchRate"].values
        y_range = training_df["RangeEtchRate"].values
        
        # Train models on iterative training data
        model_rate = _make_extratrees().fit(X, y_rate)
        model_range = _make_rf().fit(X, y_range)
        
        # Sample new candidates
        lower = np.array([FEATURE_RANGES[c][0] for c in FEATURES], float)
        upper = np.array([FEATURE_RANGES[c][1] for c in FEATURES], float)
        Xcand = _sample_candidates(SAMPLING_METHOD, N_SAMPLES, lower, upper, SAMPLING_SEED)
        Xcand = _quantize(Xcand, FEATURES)
        
        # Predict outcomes
        mu_r, sd_r = _pred_stats(model_rate, Xcand)
        mu_g, sd_g = _pred_stats(model_range, Xcand)
        mu_g_nm = mu_g * 5.0
        
        # Apply target rate filter
        mask = (mu_r >= TARGET_RATE_MIN) & (mu_r <= TARGET_RATE_MAX)
        if not np.any(mask):
            mask = np.ones(len(mu_r), dtype=bool)
        
        # Apply rf2 > rf1 filter
        rf_mask = np.ones(len(mu_r), dtype=bool)
        for i in range(len(mu_r)):
            if mask[i]:  # Only check if already passed target rate filter
                rf1_power = Xcand[i, FEATURES.index("Etch_Avg_Rf1_Pow")]
                rf2_power = Xcand[i, FEATURES.index("Etch_Avg_Rf2_Pow")]
                if rf2_power <= rf1_power:
                    rf_mask[i] = False
        
        # Combine both filters
        combined_mask = mask & rf_mask
        
        if not np.any(combined_mask):
            print(f"[regenerate] Warning: No candidates meet both target rate and rf2 > rf1 requirements")
            print(f"[regenerate] Target rate filter: {np.sum(mask)} candidates")
            print(f"[regenerate] Rf2 > rf1 filter: {np.sum(rf_mask)} candidates")
            print(f"[regenerate] Combined filter: {np.sum(combined_mask)} candidates")
            # Fallback to just target rate filter if no candidates meet both
            combined_mask = mask
        
        Xb = Xcand[combined_mask]
        mur = mu_r[combined_mask]
        mug = mu_g_nm[combined_mask]
        sdr = sd_r[combined_mask]
        sdg = sd_g[combined_mask] * 5.0
        
        print(f"[regenerate] Filtered candidates: {np.sum(mask)} (target rate) → {np.sum(combined_mask)} (with rf2 > rf1)")
        
        # Get Pareto front points for comparison from training data
        if recipes_df is not None:
            # Calculate Pareto front from the training data used for this iteration
            df_complete = training_df[["LOTNAME","FIMAP_FILE","AvgEtchRate","RangeEtchRate"]].copy()
            df_complete["Range_nm"] = df_complete["RangeEtchRate"] * 5.0
            
            # Calculate Pareto front from training data
            pts = df_complete[["AvgEtchRate","Range_nm"]].values
            flags = _is_pareto(pts)
            pareto_front = df_complete[flags].sort_values("AvgEtchRate")
            pareto_pts = pareto_front[["AvgEtchRate","Range_nm"]].values
            
            print(f"[regenerate] Iteration {iteration_num}: Using training data Pareto front with {len(pareto_pts)} points")
        else:
            # Fallback to provided current_front
            pareto_pts = current_front[["AvgEtchRate","Range_nm"]].values if len(current_front) > 0 else np.empty((0,2))
        
        # Select top K proposals
        selected = []
        selected_knobs = []
        aug_front = pareto_pts.copy()
        
        for k in range(K):
            # Filter to Pareto-improving points
            valid_candidates = []
            for i in range(len(mur)):
                if i not in selected:
                    if len(aug_front) == 0 or _improves((mur[i], mug[i]), aug_front):
                        valid_candidates.append(i)
            
            # NEW: Apply rf2 > rf1 filter to valid candidates
            rf_filtered_candidates = []
            for i in valid_candidates:
                rf1_power = Xb[i, FEATURES.index("Etch_Avg_Rf1_Pow")]
                rf2_power = Xb[i, FEATURES.index("Etch_Avg_Rf2_Pow")]
                if rf2_power > rf1_power:
                    rf_filtered_candidates.append(i)
                else:
                    print(f"[regenerate] Filtered out candidate {i}: Rf1={rf1_power:.1f}, Rf2={rf2_power:.1f} (rf2 must be > rf1)")
            
            if len(rf_filtered_candidates) == 0:
                print(f"⚠️ No candidates with rf2 > rf1 available after selecting {k} recipes!")
                break
            
            valid_candidates = rf_filtered_candidates
            
            if len(valid_candidates) == 0:
                break
            
            valid_candidates = np.array(valid_candidates)
            
            # Score candidates
            r_scale, rn_scale = _objective_scales(mur, mug)
            expl_raw = np.array([_exploit_distance_norm((mur[i], mug[i]), aug_front, r_scale, rn_scale) if len(aug_front) else 0 for i in valid_candidates])
            expl_norm = _norm01(expl_raw)
            
            explore_sigma_rn = sdr[valid_candidates] / r_scale
            explore_sigma_gn = sdg[valid_candidates] / rn_scale
            explr_raw = np.hypot(explore_sigma_rn, explore_sigma_gn)
            explr_norm = _norm01(explr_raw)
            
            if not selected_knobs:
                center = np.array([0.5]*len(FEATURES))
                div_raw = []
                for i in valid_candidates:
                    normalized_vec = []
                    for n in FEATURES:
                        val = (Xb[i][FEATURES.index(n)] - FEATURE_RANGES[n][0]) / (FEATURE_RANGES[n][1] - FEATURE_RANGES[n][0])
                        normalized_vec.append(val)
                    div_raw.append(np.linalg.norm(np.array(normalized_vec) - center))
                div_raw = np.array(div_raw)
            else:
                div_raw = np.array([_diversity_score(Xb[i], selected_knobs, FEATURES) for i in valid_candidates])
            div_norm = _norm01(div_raw)
            
            combined = ALPHA*expl_norm + BETA*explr_norm + GAMMA*div_norm
            
            # Select best candidate
            best_valid_idx = int(np.argmax(combined))
            best = valid_candidates[best_valid_idx]
            
            # Debug logging for selected recipe
            rf1_power = Xb[best, FEATURES.index("Etch_Avg_Rf1_Pow")]
            rf2_power = Xb[best, FEATURES.index("Etch_Avg_Rf2_Pow")]
            print(f"[regenerate] Selected recipe {k+1}: Rate={mur[best]:.1f}, Range={mug[best]:.1f}, Rf1={rf1_power:.1f}, Rf2={rf2_power:.1f}")
            
            selected.append(best)
            selected_knobs.append(Xb[best])
            aug_front = np.vstack([aug_front, [mur[best], mug[best]]]) if len(aug_front) else np.array([[mur[best], mug[best]]])
        
        # Return selected points
        sel_idx = [r for r in selected]
        if len(sel_idx) > 0:
            rates = mur[sel_idx]
            ranges_nm = mug[sel_idx]
            return list(zip(rates, ranges_nm))
        
        return None
        
    except Exception as e:
        print(f"[regenerate] Error generating proposals for iteration {iteration_num}: {e}")
        return None

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Pareto optimization script with enhanced plotting and iteration system")
    parser.add_argument("--date", type=str, help="View plots from specific date (YYYY-MM-DD)")
    parser.add_argument("--last", action="store_true", help="View latest available plots")
    parser.add_argument("--interactive", action="store_true", help="Interactive mode to select date for viewing plots")
    parser.add_argument("--clear-cache", action="store_true", help="Clear all cached data")
    parser.add_argument("--clear-backtesting", action="store_true", help="Clear only backtesting cache to force rebuild")
    parser.add_argument("--iteration", type=int, help="View plots for specific iteration number")
    parser.add_argument("--list-iterations", action="store_true", help="List all available iterations")
    parser.add_argument("--clear-iterations", action="store_true", help="Clear all iteration cache data")
    parser.add_argument("--regenerate-iterations", action="store_true", help="Regenerate all iterations with comprehensive 8-plot system")
    
    args = parser.parse_args()
    
    if args.clear_cache:
        _clear_cache()
    elif args.clear_backtesting:
        _clear_backtesting_cache()
    elif args.date:
        view_plots(args.date)
    elif args.last:
        view_latest()
    elif args.interactive:
        view_plots()  # Will prompt interactively
    elif args.iteration is not None:
        view_iteration(args.iteration)
    elif args.list_iterations:
        list_iterations()
    elif args.clear_iterations:
        clear_iteration_cache()
    elif args.regenerate_iterations:
        regenerate_all_iterations()
    else:
        main()
