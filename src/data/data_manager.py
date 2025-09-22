"""
Data management module for Pareto optimization system.
Handles data loading, processing, validation, and Excel integration.
"""

import os
import io
import json
import hashlib
import numpy as np
import pandas as pd
import openpyxl
import datetime as dt
from datetime import datetime, timezone
from typing import Optional, Dict, Any, List, Tuple

from ..core.config import (
    DATASET_CSV, ROOT_DIR, ROLLING_DIR, SNAPSHOTS_DIR, MANIFESTS_DIR, 
    ITERATIONS_DIR, SNAP_DIR, PLOTS_DIR, FEATURES, FEATURE_RANGES,
    EXCEL_STATUS_COL, EXCEL_DATE_COL, EXCEL_LOT_COL, EXCEL_INGEST_COL,
    EXCEL_PRED_RATE_COL, EXCEL_PRED_RANGE_COL, EXCEL_RATE_UNCERTAINTY_COL,
    EXCEL_RANGE_UNCERTAINTY_COL, GRAPH_CLIENT_ID, GRAPH_CLIENT_SECRET,
    GRAPH_TENANT_ID, GRAPH_TENANT_NAME, GRAPH_SITE_NAME, RECIPES_FILE_PATH,
    LOCAL_RECIPES_XLSX, _HAS_MSAL, CODE_VERSION, POINTS_PER_ITERATION,
    TRAINING_CUTOFF_METHOD
)
from .database_manager import DatabaseManager


class DataManager:
    """Manages all data operations for the Pareto optimization system."""
    
    def __init__(self):
        self._ensure_dirs()
        self.database = DatabaseManager()  # Initialize SQLite database manager
        from .excel_manager import ExcelManager
        self.excel_manager = ExcelManager()  # Initialize Excel manager
    
    def _ensure_dirs(self):
        """Create necessary directories"""
        for d in [ROOT_DIR, ROLLING_DIR, SNAPSHOTS_DIR, MANIFESTS_DIR, 
                  ITERATIONS_DIR, SNAP_DIR, PLOTS_DIR]:
            os.makedirs(d, exist_ok=True)
    
    def _atomic_write(self, path: str, data_bytes: bytes):
        """Atomically write data to file"""
        tmp = path + ".__tmp__"
        with open(tmp, "wb") as f:
            f.write(data_bytes)
        os.replace(tmp, path)
    
    def _read_csv(self, path: str) -> Optional[pd.DataFrame]:
        """Read CSV file safely"""
        if not os.path.exists(path):
            return None
        return pd.read_csv(path)
    
    def _append_csv_atomic(self, path: str, df: pd.DataFrame):
        """Atomically append DataFrame to CSV"""
        header = not os.path.exists(path)
        tmp = path + ".__tmp__"
        mode = "a" if os.path.exists(path) else "w"
        df.to_csv(tmp, index=False, header=header, mode=mode)
        if os.path.exists(path):
            os.remove(path)
        os.replace(tmp, path)
    
    def _hash_bytes(self, b: bytes) -> str:
        """Hash bytes using SHA256"""
        return hashlib.sha256(b).hexdigest()
    
    def _hash_df(self, df: pd.DataFrame, cols: Optional[List[str]] = None) -> str:
        """Hash DataFrame"""
        if cols is not None:
            df = df[cols].copy()
        bio = io.BytesIO()
        df.to_csv(bio, index=False)
        return self._hash_bytes(bio.getvalue())
    
    def _manifest_write(self, obj: Dict[str, Any]):
        """Write manifest to file"""
        path = os.path.join(MANIFESTS_DIR, "latest.json")
        self._atomic_write(path, json.dumps(obj, indent=2).encode("utf-8"))
    
    def _manifest_read(self) -> Dict[str, Any]:
        """Read manifest from file"""
        path = os.path.join(MANIFESTS_DIR, "latest.json")
        if not os.path.exists(path):
            return {}
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    
    def _parse_date_from_lotname(self, x: str) -> pd.Timestamp:
        """Parse date from lotname string"""
        if isinstance(x, str) and len(x) >= 10 and x[2] == "_" and x[5] == "_":
            m = x[:2]
            d = x[3:5]
            y = x[6:10]
            try:
                return pd.to_datetime(f"{y}-{m}-{d}")
            except Exception:
                return pd.NaT
        return pd.NaT
    
    def load_dataset(self) -> pd.DataFrame:
        """Load and validate the main dataset"""
        df = pd.read_csv(DATASET_CSV)
        
        # Handle date column
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"])
        else:
            if "LOTNAME" in df.columns:
                df["Date"] = df["LOTNAME"].apply(self._parse_date_from_lotname)
            else:
                raise ValueError("Dataset needs a Date column or a LOTNAME with MM_DD_YYYY prefix")
        
        # Validate required columns
        for c in ["AvgEtchRate", "RangeEtchRate"]:
            if c not in df.columns:
                raise ValueError(f"Missing column {c}")
        
        for f in FEATURES:
            if f not in df.columns:
                raise ValueError(f"Missing feature {f}")
        
        return df
    
    def read_recipes_excel(self) -> Optional[pd.DataFrame]:
        """Read recipes Excel file from SharePoint or local file"""
        # Prefer Graph if creds provided
        if _HAS_MSAL and GRAPH_CLIENT_ID and GRAPH_CLIENT_SECRET and GRAPH_TENANT_ID:
            return self._read_recipes_from_sharepoint()
        else:
            print("[recipes] ERROR: Graph credentials not configured. SharePoint Excel file is required.")
            print("[recipes] Please configure GRAPH_CLIENT_ID, GRAPH_CLIENT_SECRET, and GRAPH_TENANT_ID in your environment.")
            print("[recipes] The system will not fall back to local files as SharePoint is the only source.")
            return None
    
    def _read_recipes_from_sharepoint(self) -> Optional[pd.DataFrame]:
        """Read recipes Excel from SharePoint"""
        try:
            print("[recipes] Authenticating with Microsoft Graph…")
            import requests
            import msal
            
            authority = f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}"
            scopes = ["https://graph.microsoft.com/.default"]
            app = msal.ConfidentialClientApplication(
                GRAPH_CLIENT_ID, authority=authority, client_credential=GRAPH_CLIENT_SECRET
            )
            token_resp = app.acquire_token_for_client(scopes=scopes)
            if "access_token" not in token_resp:
                raise RuntimeError(f"MSAL token error: {token_resp}")
            token = token_resp["access_token"]
            headers = {"Authorization": f"Bearer {token}"}

            print("[recipes] Downloading Recipes Excel from SharePoint…")
            site_url = f"https://graph.microsoft.com/v1.0/sites/{GRAPH_TENANT_NAME}.sharepoint.com:/sites/{GRAPH_SITE_NAME}"
            site_id = requests.get(site_url, headers=headers).json()["id"]
            drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
            drives = requests.get(drives_url, headers=headers).json()["value"]
            drive_id = next(d["id"] for d in drives if d["name"] == "Documents")
            download_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{RECIPES_FILE_PATH}:/content"
            resp = requests.get(download_url, headers=headers)
            resp.raise_for_status()
            df = pd.read_excel(io.BytesIO(resp.content))
            
            return self._normalize_excel_data(df)
            
        except Exception as e:
            print(f"[recipes] Error reading from SharePoint: {e}")
            return None
    
    def write_recipes_to_sharepoint(self, recipes_df: pd.DataFrame) -> bool:
        """Write recipes DataFrame to SharePoint Excel file"""
        try:
            print(f"[sharepoint] Writing {len(recipes_df)} recipes to SharePoint Excel...")
            
            # Remove internal columns that shouldn't be in SharePoint
            clean_df = recipes_df.copy()
            internal_columns = ['Status_norm', 'Ingest_norm', 'Iteration']
            for col in internal_columns:
                if col in clean_df.columns:
                    clean_df = clean_df.drop(columns=[col])
                    print(f"[sharepoint] Removed internal column: {col}")
            
            # Authenticate with Microsoft Graph
            import requests
            import msal
            import openpyxl
            import io
            
            authority = f"https://login.microsoftonline.com/{GRAPH_TENANT_ID}"
            scopes = ["https://graph.microsoft.com/.default"]
            app = msal.ConfidentialClientApplication(
                GRAPH_CLIENT_ID, authority=authority, client_credential=GRAPH_CLIENT_SECRET
            )
            token_resp = app.acquire_token_for_client(scopes=scopes)
            if "access_token" not in token_resp:
                raise RuntimeError(f"MSAL token error: {token_resp}")
            token = token_resp["access_token"]
            headers = {"Authorization": f"Bearer {token}"}
            
            # Get site and drive information
            site_url = f"https://graph.microsoft.com/v1.0/sites/{GRAPH_TENANT_NAME}.sharepoint.com:/sites/{GRAPH_SITE_NAME}"
            site_id = requests.get(site_url, headers=headers).json()["id"]
            drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
            drives = requests.get(drives_url, headers=headers).json()["value"]
            drive_id = next(d["id"] for d in drives if d["name"] == "Documents")
            
            # Download current file
            download_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{RECIPES_FILE_PATH}:/content"
            resp = requests.get(download_url, headers=headers)
            resp.raise_for_status()
            
            # Load workbook and update with new data
            wb = openpyxl.load_workbook(io.BytesIO(resp.content))
            ws = wb.active
            
            # Clear existing data (except header)
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None
            
            # Write new data
            for row_idx, (_, recipe) in enumerate(clean_df.iterrows(), start=2):
                for col_idx, (col_name, value) in enumerate(recipe.items(), start=1):
                    # Convert timezone-aware datetime to timezone-naive for Excel compatibility
                    if pd.isna(value):
                        ws.cell(row=row_idx, column=col_idx).value = None
                    elif isinstance(value, pd.Timestamp) and value.tz is not None:
                        ws.cell(row=row_idx, column=col_idx).value = value.replace(tzinfo=None)
                    else:
                        ws.cell(row=row_idx, column=col_idx).value = value
            
            # Save back to SharePoint
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            upload_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/root:/{RECIPES_FILE_PATH}:/content"
            upload_headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
            
            upload_resp = requests.put(upload_url, headers=upload_headers, data=buffer.getvalue())
            upload_resp.raise_for_status()
            
            print(f"[sharepoint] Successfully wrote {len(clean_df)} recipes to SharePoint Excel")
            return True
            
        except Exception as e:
            print(f"[sharepoint] Error writing to SharePoint Excel: {e}")
            return False
    
    def _read_recipes_from_local(self) -> Optional[pd.DataFrame]:
        """Read recipes Excel from local file"""
        if not LOCAL_RECIPES_XLSX:
            print("[recipes] No Graph creds and no DT_RECIPES_XLSX set; skipping Excel features.")
            return None
        
        try:
            print(f"[recipes] Reading local Excel: {LOCAL_RECIPES_XLSX}")
            df = pd.read_excel(LOCAL_RECIPES_XLSX)
            return self._normalize_excel_data(df)
        except Exception as e:
            print(f"[recipes] Error reading local Excel: {e}")
            return None
    
    def _normalize_excel_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Normalize Excel data columns and clean up data"""
        print(f"[normalize] Starting with {len(df)} rows")
        
        # CRITICAL: Delete recipe_rejected rows FIRST
        if EXCEL_INGEST_COL in df.columns:
            rejected_mask = df[EXCEL_INGEST_COL].astype(str).str.strip().str.lower() == 'recipe_rejected'
            rejected_count = rejected_mask.sum()
            if rejected_count > 0:
                print(f"[normalize] Deleting {rejected_count} recipe_rejected rows")
                df = df[~rejected_mask].copy()
                print(f"[normalize] After deletion: {len(df)} rows remaining")
            else:
                print("[normalize] No recipe_rejected rows found")
        
        # Normalize column names
        df.columns = df.columns.str.strip()
        
        # Normalize status column
        if EXCEL_STATUS_COL in df.columns:
            df["Status_norm"] = df[EXCEL_STATUS_COL].astype(str).str.strip().str.lower()
        
        # Normalize ingestion column
        if EXCEL_INGEST_COL in df.columns:
            df["Ingest_norm"] = df[EXCEL_INGEST_COL].astype(str).str.strip().str.lower()
        
        # Normalize date column
        if EXCEL_DATE_COL in df.columns:
            df[EXCEL_DATE_COL] = pd.to_datetime(df[EXCEL_DATE_COL], errors="coerce", utc=True)
        
        # CRITICAL: Fix iteration numbering issues
        df = self._fix_iteration_numbering(df)
        
        return df
    
    def _fix_iteration_numbering(self, df: pd.DataFrame) -> pd.DataFrame:
        """Fix iteration numbering issues with double verification and renumbering"""
        if 'Iteration_num' not in df.columns:
            print("[fix_iterations] No Iteration_num column found, skipping iteration fix")
            return df
        
        print("[fix_iterations] Analyzing iteration numbering...")
        
        # Get unique iteration numbers and their counts
        iteration_counts = df['Iteration_num'].value_counts().sort_index()
        print(f"[fix_iterations] Current iteration distribution: {dict(iteration_counts)}")
        
        # Check for gaps or invalid sequences
        unique_iterations = sorted(df['Iteration_num'].unique())
        expected_iterations = list(range(1, max(unique_iterations) + 1))
        missing_iterations = [i for i in expected_iterations if i not in unique_iterations]
        
        # Check for non-sequential jumps (like 6 → 19)
        gaps = []
        for i in range(len(unique_iterations) - 1):
            current = unique_iterations[i]
            next_iter = unique_iterations[i + 1]
            if next_iter - current > 1:
                gaps.append((current, next_iter))
        
        if missing_iterations or gaps:
            print(f"[fix_iterations] ISSUES DETECTED:")
            if missing_iterations:
                print(f"[fix_iterations] Missing iterations: {missing_iterations}")
            if gaps:
                print(f"[fix_iterations] Iteration gaps: {gaps}")
            
            # Renumber iterations to be sequential
            print("[fix_iterations] Renumbering iterations to be sequential...")
            df_copy = df.copy()
            
            # Create mapping from old to new iteration numbers
            old_to_new = {}
            new_iteration = 1
            
            for old_iter in unique_iterations:
                old_to_new[old_iter] = new_iteration
                new_iteration += 1
            
            # Apply renumbering
            df_copy['Iteration_num'] = df_copy['Iteration_num'].map(old_to_new)
            
            # Verify the fix
            new_iteration_counts = df_copy['Iteration_num'].value_counts().sort_index()
            print(f"[fix_iterations] After renumbering: {dict(new_iteration_counts)}")
            
            # Double verification: ensure each iteration has exactly POINTS_PER_ITERATION recipes
            verification_passed = True
            for iter_num in df_copy['Iteration_num'].unique():
                count = (df_copy['Iteration_num'] == iter_num).sum()
                if count != POINTS_PER_ITERATION:
                    print(f"[fix_iterations] WARNING: Iteration {iter_num} has {count} recipes, expected {POINTS_PER_ITERATION}")
                    verification_passed = False
            
            if verification_passed:
                print("[fix_iterations] ✅ Double verification passed - all iterations have correct recipe counts")
            else:
                print("[fix_iterations] ⚠️ Double verification failed - some iterations have incorrect recipe counts")
            
            return df_copy
        else:
            print("[fix_iterations] ✅ Iteration numbering is already correct")
            return df
    
    def get_iteration_training_cutoff_date(self, recipes_df: pd.DataFrame, iteration_num: int) -> Optional[datetime]:
        """Get training data cutoff date for a specific iteration"""
        if recipes_df is None or EXCEL_DATE_COL not in recipes_df.columns:
            return None
        
        # Get the cutoff method from configuration
        cutoff_method = TRAINING_CUTOFF_METHOD
        
        if cutoff_method == 'last_recipe_previous':
            return self._get_cutoff_from_previous_iteration(recipes_df, iteration_num)
        else:
            return self._get_cutoff_from_current_iteration(recipes_df, iteration_num)
    
    def _get_cutoff_from_current_iteration(self, recipes_df: pd.DataFrame, iteration_num: int) -> Optional[datetime]:
        """Original approach: Use completion date of FIRST recipe in current iteration"""
        if 'Iteration_num' in recipes_df.columns:
            # Get recipes for this specific iteration
            iter_recipes = recipes_df[recipes_df['Iteration_num'] == iteration_num]
            if len(iter_recipes) == 0:
                print(f"[training] Iteration {iteration_num}: No recipes found for this iteration")
                return None
            
            # Get the completion date of the FIRST row of this iteration
            first_row = iter_recipes.iloc[0]
            completion_date = first_row.get(EXCEL_DATE_COL)
            print(f"[training] Iteration {iteration_num}: Using cutoff from first recipe of iteration {iteration_num}")
        else:
            # Fallback to row-based indexing for backward compatibility
            first_row_idx = (iteration_num - 1) * POINTS_PER_ITERATION
            
            if first_row_idx >= len(recipes_df):
                print(f"[training] Iteration {iteration_num}: No data available for this iteration")
                return None
            
            # Get the completion date of the first row of this iteration
            first_row = recipes_df.iloc[first_row_idx]
            completion_date = first_row.get(EXCEL_DATE_COL)
        
        # Convert to datetime and subtract one day to get the day before
        processed_date = self._process_completion_date(completion_date, iteration_num, "first recipe of current iteration")
        if processed_date:
            from datetime import timedelta
            cutoff_date = processed_date - timedelta(days=1)
            print(f"[training] Iteration {iteration_num}: Adjusted cutoff to day before: {cutoff_date.strftime('%Y-%m-%d')}")
            return cutoff_date
        return processed_date
    
    def _get_cutoff_from_previous_iteration(self, recipes_df: pd.DataFrame, iteration_num: int) -> Optional[datetime]:
        """Alternative approach: Use completion date of LAST recipe in previous iteration"""
        if iteration_num == 1:
            print(f"[training] Iteration {iteration_num}: No previous iteration available, using all data")
            return None
        
        prev_iteration_num = iteration_num - 1
        
        if 'Iteration_num' in recipes_df.columns:
            # Get recipes for the previous iteration
            prev_iter_recipes = recipes_df[recipes_df['Iteration_num'] == prev_iteration_num]
            if len(prev_iter_recipes) == 0:
                print(f"[training] Iteration {iteration_num}: No recipes found for previous iteration {prev_iteration_num}")
                return None
            
            # Get the completion date of the LAST row of the previous iteration
            last_row = prev_iter_recipes.iloc[-1]
            completion_date = last_row.get(EXCEL_DATE_COL)
            print(f"[training] Iteration {iteration_num}: Using cutoff from last recipe of iteration {prev_iteration_num}")
        else:
            # Fallback to row-based indexing for backward compatibility
            prev_iter_end_idx = iteration_num * POINTS_PER_ITERATION - 1
            
            if prev_iter_end_idx >= len(recipes_df):
                print(f"[training] Iteration {iteration_num}: No data available for previous iteration")
                return None
            
            # Get the completion date of the last row of the previous iteration
            last_row = recipes_df.iloc[prev_iter_end_idx]
            completion_date = last_row.get(EXCEL_DATE_COL)
        
        return self._process_completion_date(completion_date, iteration_num, f"last recipe of iteration {prev_iteration_num}")
    
    def _process_completion_date(self, completion_date, iteration_num: int, source: str) -> Optional[datetime]:
        """Process and validate completion date"""
        if pd.isna(completion_date) or completion_date is None:
            print(f"[training] Iteration {iteration_num}: No completion date for {source}")
            return None
        
        # Convert Excel date format if needed
        if isinstance(completion_date, str):
            try:
                completion_date = pd.to_datetime(completion_date, format='%m/%d/%Y')
            except:
                completion_date = pd.to_datetime(completion_date, errors='coerce')
        
        if pd.isna(completion_date):
            print(f"[training] Iteration {iteration_num}: Could not parse completion date from {source}")
            return None
        
        # Ensure timezone-naive datetime
        if completion_date.tz is not None:
            completion_date = completion_date.tz_localize(None)
        
        print(f"[training] Iteration {iteration_num}: Training cutoff date = {completion_date.strftime('%Y-%m-%d')}")
        return completion_date
    
    def get_training_data_for_iteration(self, df: pd.DataFrame, recipes_df: pd.DataFrame, 
                                       iteration_num: int) -> pd.DataFrame:
        """Get training data for a specific iteration"""
        print(f"[training] Getting training data for iteration {iteration_num}")
        
        # Get the cutoff date for this iteration
        cutoff_date = self.get_iteration_training_cutoff_date(recipes_df, iteration_num)
        
        if cutoff_date is None:
            # If no cutoff date available, this means we're proposing recipes for a future iteration
            # In this case, use the entire dataset for training
            print(f"[training] Iteration {iteration_num}: No cutoff date available - using entire dataset for proposal generation")
            print(f"[training] Iteration {iteration_num}: Training on {len(df)} points (entire dataset)")
            return df.copy()
        
        # Ensure run_date column is properly converted to datetime
        df_copy = self._ensure_run_date_datetime(df)
        
        # Apply cutoff-based filtering: keep only data strictly before (or equal to day-before) the cutoff
        # Note: get_iteration_training_cutoff_date already subtracts one day from the first recipe date
        if "run_date" in df_copy.columns and cutoff_date is not None:
            try:
                # Use <= because cutoff_date was already adjusted to the day before
                df_copy = df_copy[df_copy["run_date"] <= pd.to_datetime(cutoff_date)]
            except Exception:
                # If any parsing issue occurs, leave df_copy as-is (fail-safe)
                pass
        
        # Get lot names for this iteration to exclude from training
        iteration_lotnames = self._get_iteration_lotnames(recipes_df, iteration_num)
        
        # Filter data to exclude recipe points from this iteration
        training_mask = ~df_copy["LOTNAME"].isin(iteration_lotnames)
        training_data = df_copy[training_mask].copy()
        
        print(f"[training] Iteration {iteration_num}: Training on {len(training_data)} points "
              f"(excluded {len(iteration_lotnames)} points: {iteration_lotnames})")
        
        return training_data
    
    def _get_iteration_lotnames(self, recipes_df: pd.DataFrame, iteration_num: int) -> List[str]:
        """Get lot names for a specific iteration"""
        iteration_recipes = recipes_df[recipes_df['Iteration_num'] == iteration_num]
        return iteration_recipes['Lotname'].tolist()
    
    def get_training_data_debug_info(self, original_df: pd.DataFrame, recipes_df: pd.DataFrame, 
                                   iteration_num: int, rate_model_type: str, range_model_type: str,
                                   rate_params: dict, range_params: dict) -> dict:
        """Get debug information for training data"""
        debug_info = {}
        
        try:
            # Get training data for this iteration
            training_data = self.get_training_data_for_iteration(original_df, recipes_df, iteration_num)
            
            # Get cutoff date info
            cutoff_date = self.get_iteration_training_cutoff_date(recipes_df, iteration_num)
            
            if cutoff_date is not None:
                debug_info['training_cutoff_date'] = cutoff_date.strftime('%Y-%m-%d')
                debug_info['total_training_points'] = len(training_data)
                
                # Get first and last lots
                if len(training_data) > 0:
                    # Check for both 'Lotname' and 'LOTNAME' columns (case sensitivity issue)
                    lotname_col = None
                    if 'Lotname' in training_data.columns:
                        lotname_col = 'Lotname'
                    elif 'LOTNAME' in training_data.columns:
                        lotname_col = 'LOTNAME'
                    
                    if lotname_col is not None:
                        debug_info['first_lot'] = training_data[lotname_col].iloc[0]
                        debug_info['last_lot'] = training_data[lotname_col].iloc[-1]
                    else:
                        debug_info['first_lot'] = 'No lot data available'
                        debug_info['last_lot'] = 'No lot data available'
                else:
                    debug_info['first_lot'] = 'No lot data available'
                    debug_info['last_lot'] = 'No lot data available'
            else:
                # No cutoff date means we're using the entire dataset for proposal generation
                debug_info['training_cutoff_date'] = 'N/A (using entire dataset for proposal generation)'
                debug_info['total_training_points'] = len(training_data)
                
                # Get first and last lots from the entire dataset
                if len(training_data) > 0:
                    # Check for both 'Lotname' and 'LOTNAME' columns (case sensitivity issue)
                    lotname_col = None
                    if 'Lotname' in training_data.columns:
                        lotname_col = 'Lotname'
                    elif 'LOTNAME' in training_data.columns:
                        lotname_col = 'LOTNAME'
                    
                    if lotname_col is not None:
                        debug_info['first_lot'] = training_data[lotname_col].iloc[0]
                        debug_info['last_lot'] = training_data[lotname_col].iloc[-1]
                    else:
                        debug_info['first_lot'] = 'No lot data available'
                        debug_info['last_lot'] = 'No lot data available'
                else:
                    debug_info['first_lot'] = 'No lot data available'
                    debug_info['last_lot'] = 'No lot data available'
            
            # Add model info
            debug_info['rate_model'] = rate_model_type
            debug_info['range_model'] = range_model_type
            debug_info['rate_params'] = str(rate_params) if rate_params else 'N/A'
            debug_info['range_params'] = str(range_params) if range_params else 'N/A'
            
        except Exception as e:
            print(f"[debug] Error in get_training_data_debug_info: {e}")
            import traceback
            traceback.print_exc()
            # Return minimal debug info
            debug_info = {
                'training_cutoff_date': 'ERROR',
                'total_training_points': 'ERROR',
                'first_lot': 'ERROR',
                'last_lot': 'ERROR',
                'rate_model': rate_model_type,
                'range_model': range_model_type,
                'rate_params': str(rate_params) if rate_params else 'N/A',
                'range_params': str(range_params) if range_params else 'N/A'
            }
        
        return debug_info
    
    def _ensure_run_date_datetime(self, df: pd.DataFrame) -> pd.DataFrame:
        """Ensure run_date column is datetime"""
        if "run_date" not in df.columns:
            print("[training] Warning: No run_date column found in dataset")
            return df
        
        if df["run_date"].dtype == 'object':
            try:
                df_copy = df.copy()
                df_copy["run_date"] = pd.to_datetime(df_copy["run_date"])
                print("[training] Successfully converted run_date column to datetime")
            except Exception as e:
                print(f"[training] Warning: Could not convert run_date to datetime: {e}")
                return df
        else:
            df_copy = df.copy()
        
        # Ensure timezone-naive datetime
        if df_copy["run_date"].dt.tz is not None:
            df_copy["run_date"] = df_copy["run_date"].dt.tz_localize(None)
            print("[training] Converted timezone-aware dates to timezone-naive")
        
        return df_copy
    
    def get_completed_experimental_data_for_iteration(self, df: pd.DataFrame, 
                                                     recipes_df: pd.DataFrame, 
                                                     iteration_num: int) -> pd.DataFrame:
        """Get completed experimental data for a specific iteration"""
        if recipes_df is None or EXCEL_LOT_COL not in recipes_df.columns:
            return pd.DataFrame()
        
        # Get the lotnames for this iteration
        start_idx = (iteration_num - 1) * POINTS_PER_ITERATION
        end_idx = start_idx + POINTS_PER_ITERATION
        
        if start_idx >= len(recipes_df):
            return pd.DataFrame()
        
        iter_recipes = recipes_df.iloc[start_idx:end_idx]
        
        # Get lotnames that have completion dates
        completed_lots = []
        for _, recipe in iter_recipes.iterrows():
            lotname = recipe.get(EXCEL_LOT_COL)
            completion_date = recipe.get(EXCEL_DATE_COL)
            
            if (isinstance(lotname, str) and lotname.strip() and 
                not pd.isna(completion_date) and completion_date is not None):
                completed_lots.append(lotname)
        
        if not completed_lots:
            print(f"[training] Iteration {iteration_num}: No completed lots found")
            return pd.DataFrame()
        
        # Find these lots in the full dataset
        experimental_data = df[df["LOTNAME"].isin(completed_lots)].copy()
        
        print(f"[training] Iteration {iteration_num}: Found {len(experimental_data)} experimental points "
              f"for lots: {', '.join(completed_lots)}")
        
        return experimental_data
    
    def get_excel_iteration_status(self, recipes_df: pd.DataFrame) -> Dict[int, Dict[str, Any]]:
        """Analyze Excel file to determine iteration status"""
        if recipes_df is None or EXCEL_STATUS_COL not in recipes_df.columns:
            return {}
        
        iteration_status = {}
        
        # Use Iteration_num column if available, otherwise fall back to row-based indexing
        if 'Iteration_num' in recipes_df.columns:
            unique_iterations = sorted(recipes_df['Iteration_num'].unique())
            
            # Process all iterations from 1 to the maximum iteration number
            max_iteration = max(unique_iterations) if unique_iterations else 0
            for iter_num in range(1, max_iteration + 1):
                # Get recipes for this iteration using Iteration_num column
                iter_recipes = recipes_df[recipes_df['Iteration_num'] == iter_num]
                
                if len(iter_recipes) == 0:
                    # This iteration doesn't exist yet
                    iteration_status[iter_num] = {
                        "completed_count": 0,
                        "pending_count": 0,
                        "is_completed": False,
                        "status": "not_started",
                        "recipes": pd.DataFrame()
                    }
                else:
                    # Count completed vs pending
                    completed_count = len(iter_recipes[iter_recipes["Status_norm"] == "completed"])
                    pending_count = len(iter_recipes[iter_recipes["Status_norm"] == "pending"])
                    
                    # Determine status
                    if completed_count == POINTS_PER_ITERATION:
                        status = "completed"
                    elif pending_count > 0:
                        status = "pending"
                    else:
                        status = "incomplete"
                    
                    iteration_status[iter_num] = {
                        "completed_count": completed_count,
                        "pending_count": pending_count,
                        "is_completed": completed_count == POINTS_PER_ITERATION,
                        "status": status,
                        "recipes": iter_recipes
                    }
        else:
            # Fallback to row-based indexing for backward compatibility
            total_recipes = len(recipes_df)
            total_iterations = total_recipes // POINTS_PER_ITERATION
            
            for iter_num in range(total_iterations):
                start_idx = iter_num * POINTS_PER_ITERATION
                end_idx = start_idx + POINTS_PER_ITERATION
                
                # Get recipes for this iteration
                iter_recipes = recipes_df.iloc[start_idx:end_idx]
                
                # Count completed vs pending
                completed_count = len(iter_recipes[iter_recipes["Status_norm"] == "completed"])
                pending_count = len(iter_recipes[iter_recipes["Status_norm"] == "pending"])
                
                iteration_status[iter_num] = {
                    "start_idx": start_idx,
                    "end_idx": end_idx,
                    "completed_count": completed_count,
                    "pending_count": pending_count,
                    "is_completed": completed_count == POINTS_PER_ITERATION,
                    "recipes": iter_recipes
                }
        
        return iteration_status
    
    def get_proposed_recipes_for_iteration(self, recipes_df: pd.DataFrame, 
                                          iteration_num: int) -> Optional[List[Dict[str, Any]]]:
        """Get proposed recipes for a specific iteration"""
        if recipes_df is None or EXCEL_STATUS_COL not in recipes_df.columns:
            return None
        
        # Use Iteration_num column to filter recipes for the specific iteration
        if 'Iteration_num' in recipes_df.columns:
            iter_recipes = recipes_df[recipes_df['Iteration_num'] == iteration_num]
        else:
            # Fallback to row-based indexing for backward compatibility
            start_idx = (iteration_num - 1) * POINTS_PER_ITERATION
            end_idx = start_idx + POINTS_PER_ITERATION
            
            if start_idx >= len(recipes_df):
                return None
            
            iter_recipes = recipes_df.iloc[start_idx:end_idx]
        
        if len(iter_recipes) == 0:
            return None
        
        # Extract the recipe data
        proposed_recipes = []
        for i, (_, recipe) in enumerate(iter_recipes.iterrows()):
            recipe_data = {
                "O2_flow": recipe.get("O2_flow", 0.0),
                "cf4_flow": recipe.get("cf4_flow", 0.0),
                "Rf1_Pow": recipe.get("Rf1_Pow", 0.0),
                "Rf2_Pow": recipe.get("Rf2_Pow", 0.0),
                "Pressure": recipe.get("Pressure", 0.0),
                "predicted_rate": recipe.get(EXCEL_PRED_RATE_COL, 0.0),
                "predicted_range": recipe.get(EXCEL_PRED_RANGE_COL, 0.0),
                "rate_uncertainty": recipe.get(EXCEL_RATE_UNCERTAINTY_COL, 0.0),
                "range_uncertainty": recipe.get(EXCEL_RANGE_UNCERTAINTY_COL, 0.0),
                "status": recipe.get(EXCEL_STATUS_COL, ""),
                "lotname": recipe.get(EXCEL_LOT_COL, ""),
                "date_completed": recipe.get(EXCEL_DATE_COL, "")
            }
            proposed_recipes.append(recipe_data)
        
        return proposed_recipes
    
    def should_propose_new_recipes(self, recipes_df: pd.DataFrame, current_iteration: int) -> bool:
        """Determine if we should propose new recipes for the next iteration"""
        if recipes_df is None:
            return False
        
        iteration_status = self.get_excel_iteration_status(recipes_df)
        
        # If current iteration doesn't exist in Excel, we definitely need to propose recipes for it
        if current_iteration not in iteration_status:
            return True
        
        # If current iteration exists but isn't completed, don't propose new recipes yet
        if not iteration_status[current_iteration]["is_completed"]:
            return False
        
        # Check if next iteration already has proposed recipes
        next_iteration = current_iteration + 1
        if next_iteration in iteration_status:
            return False
        
        return True
    
    def get_completed_lots_by_iteration(self, recipes_df: pd.DataFrame) -> Dict[int, List[str]]:
        """Group completed lots by iteration based on completion order"""
        if recipes_df is None or EXCEL_STATUS_COL not in recipes_df.columns:
            return {}
        
        # Get completed lots sorted by completion date
        completed_df = recipes_df[recipes_df["Status_norm"] == "completed"].copy()
        if completed_df.empty:
            return {}
        
        # Sort by completion date
        if EXCEL_DATE_COL in completed_df.columns:
            completed_df = completed_df.sort_values(EXCEL_DATE_COL)
        
        # Group into iterations
        iterations = {}
        current_iteration = 0
        current_points = []
        
        for _, row in completed_df.iterrows():
            lot_name = row[EXCEL_LOT_COL]
            current_points.append(lot_name)
            
            if len(current_points) == POINTS_PER_ITERATION:
                iterations[current_iteration] = current_points.copy()
                current_iteration += 1
                current_points = []
        
        # Handle partial iteration
        if current_points:
            iterations[current_iteration] = current_points
        
        return iterations
    
    def get_cumulative_highlight_lots(self, iterations: Dict[int, List[str]], 
                                     target_iteration: int) -> List[str]:
        """Get cumulative highlight lots up to a specific iteration"""
        if target_iteration < 0:
            return []
        
        highlight_lots = []
        for i in range(target_iteration + 1):
            if i in iterations:
                highlight_lots.extend(iterations[i])
        
        return highlight_lots
    
    def get_completed_recipes_up_to_iteration(self, recipes_df: pd.DataFrame, iteration_num: int) -> List[Dict[str, Any]]:
        """Get completed recipes from previous iterations for training progression"""
        completed_recipes = []
        
        if recipes_df is None or recipes_df.empty:
            return completed_recipes
        
        # Filter for completed recipes up to the specified iteration
        if "Status_norm" in recipes_df.columns:
            # Calculate iteration based on row position (every 3 rows = 1 iteration)
            # Add iteration column based on row position
            recipes_df_copy = recipes_df.copy()
            recipes_df_copy["Iteration"] = (recipes_df_copy.index // POINTS_PER_ITERATION) + 1
            
            completed_mask = (recipes_df_copy["Status_norm"] == "completed") & (recipes_df_copy["Iteration"] <= iteration_num)
            completed_df = recipes_df_copy[completed_mask]
            
            for _, recipe in completed_df.iterrows():
                completed_recipes.append({
                    "lotname": recipe.get(EXCEL_LOT_COL, ""),
                    "predicted_rate": recipe.get(EXCEL_PRED_RATE_COL, 0.0),
                    "predicted_range": recipe.get(EXCEL_PRED_RANGE_COL, 0.0),
                    "iteration": recipe.get("Iteration", 0),
                    "status": recipe.get("Status_norm", "")
                })
        
        return completed_recipes
    
    def get_highlight_lots_for_iteration(self, recipes_df: pd.DataFrame, iteration_num: int) -> List[str]:
        """Get highlight lots for this iteration (completed recipes)"""
        highlight_lots = []
        
        if recipes_df is None or recipes_df.empty:
            return highlight_lots
        
        # Get completed lots from this iteration and previous iterations
        if "Status_norm" in recipes_df.columns:
            # Calculate iteration based on row position (every 3 rows = 1 iteration)
            # Add iteration column based on row position
            recipes_df_copy = recipes_df.copy()
            recipes_df_copy["Iteration"] = (recipes_df_copy.index // POINTS_PER_ITERATION) + 1
            
            # Get completed lots from current iteration
            current_completed_mask = (recipes_df_copy["Status_norm"] == "completed") & (recipes_df_copy["Iteration"] == iteration_num)
            current_completed_lots = recipes_df_copy.loc[current_completed_mask, EXCEL_LOT_COL].tolist()
            
            # Get completed lots from previous iterations
            previous_completed_mask = (recipes_df_copy["Status_norm"] == "completed") & (recipes_df_copy["Iteration"] < iteration_num)
            previous_completed_lots = recipes_df_copy.loc[previous_completed_mask, EXCEL_LOT_COL].tolist()
            
            # Combine and count
            all_completed_lots = current_completed_lots + previous_completed_lots
            print(f"[highlight] Iteration {iteration_num}: Found {len(current_completed_lots)} completed lots for highlighting (3 current + {len(previous_completed_lots)} previous)")
            
            # Take the most recent completed lots (up to 6 total)
            highlight_lots = all_completed_lots[-6:] if len(all_completed_lots) > 6 else all_completed_lots
            
            print(f"[highlight] Iteration {iteration_num}: Found {len(highlight_lots)} completed lots for highlighting")
        
        return highlight_lots
    
    # Hash functions for caching
    def dataset_hash(self, df: pd.DataFrame) -> str:
        """Calculate hash of dataset"""
        cols = FEATURES + ["AvgEtchRate", "RangeEtchRate", "run_date"]
        return self._hash_df(df[cols])
    
    def features_hash(self) -> str:
        """Calculate hash of features"""
        payload = json.dumps({"features": FEATURES, "ranges": FEATURE_RANGES}, 
                           sort_keys=True).encode("utf-8")
        return self._hash_bytes(payload)
    
    def model_config_hash(self) -> str:
        """Calculate hash of model configuration"""
        payload = json.dumps({
            "K": 3,  # Hardcoded for now
            "SAMPLING": "sobol",  # Hardcoded for now
            "SEED": 42,  # Hardcoded for now
            "NSAMPLES": 200000,  # Hardcoded for now
            "ALPHA": 1.0,  # Hardcoded for now
            "BETA": 0.25,  # Hardcoded for now
            "GAMMA": 0.4,  # Hardcoded for now
            "models": {"rate": "ExtraTrees+Scaler/200", "range": "RF+Scaler/200"}
        }, sort_keys=True).encode("utf-8")
        return self._hash_bytes(payload)
    
    def code_hash(self) -> str:
        """Calculate hash of code version"""
        return self._hash_bytes(CODE_VERSION.encode("utf-8"))
    
    def cache_key(self, dataset_h: str, features_h: str, model_h: str, code_h: str) -> str:
        """Generate cache key from component hashes"""
        return self._hash_bytes("|".join([dataset_h, features_h, model_h, code_h]).encode("utf-8"))
    
    # File path getters
    def loocv_path(self) -> str:
        """Get LOOCV predictions file path"""
        return os.path.join(ROLLING_DIR, "loocv_predictions.csv")
    
    def pareto_history_path(self) -> str:
        """Get Pareto front history file path"""
        return os.path.join(ROLLING_DIR, "pareto_front_history.csv")
    
    def manifest_path(self) -> str:
        """Get manifest file path"""
        return os.path.join(MANIFESTS_DIR, "latest.json")
    
    def save_iteration_data_to_database(self, iteration_num: int, debug_info: Dict[str, str], 
                                       highlight_lots: List[str], pareto_front_data: pd.DataFrame,
                                       selected_points: List[Tuple[float, float]], 
                                       full_dataset: pd.DataFrame = None,
                                       loocv_data: pd.DataFrame = None,
                                       uncertainties: Dict[str, List[float]] = None,
                                       metrics: Dict[str, float] = None) -> bool:
        """Save iteration data to both file cache and SQLite database"""
        try:
            # Save to database
            success = self.database.save_iteration_data(
                iteration_num, debug_info, highlight_lots, pareto_front_data, selected_points, 
                full_dataset, loocv_data, uncertainties, metrics
            )
            
            if success:
                print(f"[database] Successfully saved iteration {iteration_num} to SQLite database")
            else:
                print(f"[database] Failed to save iteration {iteration_num} to SQLite database")
            
            return success
            
        except Exception as e:
            print(f"[database] Error saving iteration {iteration_num} to database: {e}")
            return False
    
    def get_database_info(self) -> Dict[str, Any]:
        """Get database statistics and information"""
        return self.database.get_database_info()
    
    def export_database_to_csv(self) -> bool:
        """Export database tables to CSV files"""
        return self.database.export_to_csv()
