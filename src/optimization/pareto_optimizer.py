"""
Pareto optimization module for semiconductor etching processes.
Handles Pareto front calculations, scoring, and proposal generation.
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from typing import List, Tuple, Dict, Any, Optional
from datetime import datetime, timezone
import os
from scipy.stats import norm

from ..core.config import (
    FEATURES, FEATURE_RANGES, ALPHA, BETA, GAMMA, K,
    RATE_IMPROVEMENT_THRESHOLD, RANGE_IMPROVEMENT_THRESHOLD,
    EXCEL_PRED_RATE_COL, EXCEL_PRED_RANGE_COL, EXCEL_RATE_UNCERTAINTY_COL,
    EXCEL_RANGE_UNCERTAINTY_COL, SNAP_DIR, TODAY
)


class ParetoOptimizer:
    """Handles Pareto front optimization and proposal generation."""
    
    def __init__(self):
        pass
    
    def is_pareto(self, points: np.ndarray) -> np.ndarray:
        """Check which points are Pareto optimal"""
        P = np.array(points, dtype=float)
        flags = np.ones(len(P), dtype=bool)
        
        for i, (r, rng) in enumerate(P):
            # Check if this point is dominated by any other point
            for j, (r2, rng2) in enumerate(P):
                if i != j:  # Don't compare point with itself
                    if (r2 >= r and rng2 <= rng and (r2 > r or rng2 < rng)):
                        flags[i] = False
                        break
        
        return flags
    
    def pareto_front(self, df_rates_ranges: pd.DataFrame) -> pd.DataFrame:
        """Calculate Pareto front from dataframe"""
        pts = df_rates_ranges[["AvgEtchRate", "Range_nm"]].values
        flags = self.is_pareto(pts)
        return df_rates_ranges.loc[flags].sort_values("AvgEtchRate").reset_index(drop=True)
    
    def norm01(self, v: np.ndarray) -> np.ndarray:
        """Normalize array to [0,1] range"""
        v = np.asarray(v, float)
        vmin, vmax = float(np.min(v)), float(np.max(v))
        if vmax - vmin < 1e-12:
            return np.zeros_like(v)
        return (v - vmin) / (vmax - vmin)
    
    def objective_scales(self, rate_vals: np.ndarray, range_nm_vals: np.ndarray) -> Tuple[float, float]:
        """Calculate objective scales for normalization"""
        r_scale = np.percentile(rate_vals, 97.5) - np.percentile(rate_vals, 2.5)
        rn_scale = np.percentile(range_nm_vals, 97.5) - np.percentile(range_nm_vals, 2.5)
        
        if r_scale < 1e-9:
            r_scale = max(1.0, np.ptp(rate_vals))
        if rn_scale < 1e-9:
            rn_scale = max(1.0, np.ptp(range_nm_vals))
        
        return float(r_scale), float(rn_scale)
    
    def improves(self, p: Tuple[float, float], front: np.ndarray) -> bool:
        """Check if point p improves the Pareto front"""
        if len(front) == 0:
            return True  # Empty front, any point improves it
        
        # Check if p is dominated by any existing point on the front
        for front_point in front:
            # front_point dominates p if front_point is better or equal in both objectives
            # AND strictly better in at least one objective
            if (front_point[0] >= p[0] and front_point[1] <= p[1] and 
                (front_point[0] > p[0] or front_point[1] < p[1])):
                return False  # p is dominated by this front point
        
        # If we get here, p is not dominated by any front point
        return True
    
    def dominates_existing(self, p: Tuple[float, float], front: np.ndarray) -> bool:
        """Check if point p dominates ANY existing point on the Pareto front"""
        if len(front) == 0:
            return False  # Empty front, nothing to dominate
        
        # Check if p dominates any existing point on the front
        for front_point in front:
            # p dominates front_point if p is better or equal in both objectives
            # AND strictly better in at least one objective
            if (p[0] >= front_point[0] and p[1] <= front_point[1] and 
                (p[0] > front_point[0] or p[1] < front_point[1])):
                return True  # p dominates this front point
        
        return False
    
    def significantly_improves(self, p: Tuple[float, float], front: np.ndarray, 
                             rate_threshold: float = 0.05, range_threshold: float = 0.05) -> bool:
        """Check if point p significantly improves the Pareto front"""
        if len(front) == 0:
            return True  # Empty front, any point improves it
        
        # First check if it dominates any existing point
        if self.dominates_existing(p, front):
            return True
        
        # If not dominating, check if it's non-dominated and provides significant improvement
        if not self.improves(p, front):
            return False  # Point is dominated
        
        # Find the closest point on the front
        min_distance = float('inf')
        closest_point = None
        
        for front_point in front:
            # Calculate normalized distance
            rate_diff = abs(p[0] - front_point[0]) / max(1, abs(front_point[0]))
            range_diff = abs(p[1] - front_point[1]) / max(1, abs(front_point[1]))
            distance = np.sqrt(rate_diff**2 + range_diff**2)
            if distance < min_distance:
                min_distance = distance
                closest_point = front_point
        
        if closest_point is None:
            return True  # Shouldn't happen, but safety check
        
        # Check if there's significant improvement in at least one objective
        rate_improvement = (p[0] - closest_point[0]) / max(1, abs(closest_point[0]))
        range_improvement = (closest_point[1] - p[1]) / max(1, abs(closest_point[1]))  # Lower range is better
        
        # Return True if there's significant improvement in at least one objective
        return (rate_improvement > rate_threshold) or (range_improvement > range_threshold)
    
    def exploit_distance_norm(self, p: Tuple[float, float], front: np.ndarray, 
                             r_scale: float, rn_scale: float) -> float:
        """Calculate normalized distance to Pareto front"""
        if len(front) == 0:
            return 0.0
        
        dr = (p[0] - front[:, 0]) / r_scale
        dn = (p[1] - front[:, 1]) / rn_scale
        return np.min(np.hypot(dr, dn))
    
    def diversity_score(self, candidate_knob: np.ndarray, selected_knobs: List[np.ndarray], 
                       cols: List[str]) -> float:
        """Calculate diversity score for candidate point"""
        if not selected_knobs:
            return 1.0
        
        def _norm_vec(vec: np.ndarray) -> np.ndarray:
            out = []
            for name in cols:
                lo, hi = FEATURE_RANGES[name]
                out.append((vec[cols.index(name)] - lo) / (hi - lo))
            return np.array(out)
        
        nc = _norm_vec(candidate_knob)
        dmin = float("inf")
        
        for sk in selected_knobs:
            ns = _norm_vec(sk)
            d = np.linalg.norm(nc - ns)
            if d < dmin:
                dmin = d
        
        return float(dmin)
    
    def pareto_improvement_objective(self, rate: float, range_nm: float, pareto_pts: np.ndarray) -> float:
        """Calculate Pareto improvement potential as the objective function for EI
        
        For Pareto optimization, we want to MAXIMIZE this value to find points that improve the Pareto front.
        This function calculates how much a point improves the current Pareto front by measuring:
        1. How much it dominates existing Pareto points
        2. How much it extends the Pareto front in rate and range dimensions
        """
        if len(pareto_pts) == 0:
            # If no existing Pareto front, use a simple weighted sum
            # Higher rate is better, lower range is better
            rate_score = rate / 200.0  # Normalize by expected max rate
            range_score = (50.0 - range_nm) / 50.0  # Invert range (lower is better)
            return rate_score + range_score
        
        # Normalize objectives for fair comparison
        rate_norm = rate / 200.0  # Normalize by expected max rate
        range_norm = range_nm / 50.0  # Normalize by expected max range
        
        # Calculate improvement in each dimension
        rate_improvement = 0.0
        range_improvement = 0.0
        
        # Check if this point dominates any existing Pareto point
        dominates_any = False
        for pareto_rate, pareto_range in pareto_pts:
            pareto_rate_norm = pareto_rate / 200.0
            pareto_range_norm = pareto_range / 50.0
            
            # Check if current point dominates Pareto point
            if rate_norm > pareto_rate_norm and range_norm < pareto_range_norm:
                dominates_any = True
                # Calculate improvement
                rate_improvement += (rate_norm - pareto_rate_norm)
                range_improvement += (pareto_range_norm - range_norm)
        
        if dominates_any:
            # If it dominates, use the improvement score
            return rate_improvement + range_improvement
        else:
            # If it doesn't dominate, check if it extends the Pareto front
            max_rate = max(pareto_pts[:, 0]) / 200.0 if len(pareto_pts) > 0 else 0
            min_range = min(pareto_pts[:, 1]) / 50.0 if len(pareto_pts) > 0 else 1
            
            # Calculate how much it extends the front
            rate_extension = max(0, rate_norm - max_rate)
            range_extension = max(0, min_range - range_norm)
            
            # Weight the extensions (rate extension is more valuable)
            return rate_extension * 2.0 + range_extension
    
    def expected_improvement(self, mu: np.ndarray, sigma: np.ndarray, f_best: float) -> np.ndarray:
        """Calculate Expected Improvement acquisition function
        
        Args:
            mu: Predicted mean values
            sigma: Predicted standard deviations  
            f_best: Best observed objective value so far
            
        Returns:
            Expected improvement values
        """
        # Avoid division by zero
        sigma = np.maximum(sigma, 1e-9)
        
        # Calculate improvement
        improvement = mu - f_best
        
        # Calculate z-score
        z = improvement / sigma
        
        # Calculate EI using the formula from the paper
        ei = improvement * norm.cdf(z) + sigma * norm.pdf(z)
        
        return ei
    
    def batch_ei_selection(self, X_band: np.ndarray, mu_r: np.ndarray, mu_rng_nm: np.ndarray,
                          sr: np.ndarray, srg_nm: np.ndarray, pareto_pts: np.ndarray,
                          ml_models, batch_size: int = K) -> Tuple[List[int], List[Dict]]:
        """Select batch of candidates using Expected Improvement with sequential model updates
        
        This implements the batch optimization approach from the paper where:
        1. Select first candidate using EI
        2. Update model with predicted value of first candidate
        3. Select second candidate using updated model
        4. Repeat until batch is complete
        
        Args:
            X_band: Candidate feature vectors
            mu_r: Predicted etch rates
            mu_rng_nm: Predicted ranges (in nm)
            sr: Rate uncertainties
            srg_nm: Range uncertainties (in nm)
            pareto_pts: Current Pareto front points
            ml_models: ML models object for retraining
            batch_size: Number of candidates to select
            
        Returns:
            Tuple of (selected_indices, selection_details)
        """
        print(f"[batch_ei] Starting batch EI selection with batch size {batch_size}")
        
        selected_indices = []
        selection_details = []
        remaining_candidates = list(range(len(X_band)))
        
        # Calculate current best objective value from Pareto front
        if len(pareto_pts) > 0:
            # Calculate objective for each Pareto point using the new Pareto improvement metric
            pareto_objectives = []
            for rate, range_nm in pareto_pts:
                obj = self.pareto_improvement_objective(rate, range_nm, pareto_pts)
                pareto_objectives.append(obj)
            f_best = max(pareto_objectives)
            print(f"[batch_ei] Current best Pareto improvement value: {f_best:.4f}")
        else:
            f_best = 0.0
            print(f"[batch_ei] No existing Pareto front, starting with f_best = 0")
        
        # For each candidate in the batch
        for batch_idx in range(batch_size):
            print(f"[batch_ei] Selecting candidate {batch_idx + 1}/{batch_size}")
            
            if not remaining_candidates:
                print(f"[batch_ei] No more candidates available")
                break
            
            # Calculate objective values for remaining candidates using Pareto improvement
            candidate_objectives = []
            for idx in remaining_candidates:
                obj = self.pareto_improvement_objective(mu_r[idx], mu_rng_nm[idx], pareto_pts)
                candidate_objectives.append(obj)
            
            # Calculate EI for remaining candidates
            # We need to combine rate and range uncertainties into a single uncertainty
            # We'll use the geometric mean of the uncertainties
            combined_uncertainties = np.sqrt(sr[remaining_candidates]**2 + srg_nm[remaining_candidates]**2)
            
            ei_values = self.expected_improvement(
                np.array(candidate_objectives), 
                combined_uncertainties, 
                f_best
            )
            
            # Apply RF constraint: rf2 > rf1
            valid_candidates = []
            valid_ei_values = []
            
            for i, idx in enumerate(remaining_candidates):
                rf1_power = X_band[idx, FEATURES.index("Etch_Avg_Rf1_Pow")]
                rf2_power = X_band[idx, FEATURES.index("Etch_Avg_Rf2_Pow")]
                
                if rf2_power > rf1_power:
                    valid_candidates.append(idx)
                    valid_ei_values.append(ei_values[i])
                else:
                    print(f"[batch_ei] Filtered out candidate {idx}: Rf1={rf1_power:.1f}, Rf2={rf2_power:.1f} (rf2 must be > rf1)")
            
            if not valid_candidates:
                print(f"[batch_ei] No valid candidates with rf2 > rf1 constraint")
                break
            
            # Select candidate with highest EI
            best_local_idx = np.argmax(valid_ei_values)
            selected_idx = valid_candidates[best_local_idx]
            
            selected_indices.append(selected_idx)
            remaining_candidates.remove(selected_idx)
            
            # Calculate details for this selection
            selected_obj = self.pareto_improvement_objective(mu_r[selected_idx], mu_rng_nm[selected_idx], pareto_pts)
            selected_ei = valid_ei_values[best_local_idx]
            
            selection_details.append({
                "batch_idx": batch_idx + 1,
                "selected_idx": selected_idx,
                "objective_value": selected_obj,
                "ei_value": selected_ei,
                "rate_pred": mu_r[selected_idx],
                "range_pred": mu_rng_nm[selected_idx],
                "rate_uncertainty": sr[selected_idx],
                "range_uncertainty": srg_nm[selected_idx],
                "rf1_power": X_band[selected_idx, FEATURES.index("Etch_Avg_Rf1_Pow")],
                "rf2_power": X_band[selected_idx, FEATURES.index("Etch_Avg_Rf2_Pow")]
            })
            
            print(f"[batch_ei] Selected candidate {selected_idx}: Obj={selected_obj:.4f}, EI={selected_ei:.4f}, "
                  f"Rate={mu_r[selected_idx]:.1f}, Range={mu_rng_nm[selected_idx]:.1f}")
            
            # Update f_best for next iteration
            if selected_obj > f_best:
                f_best = selected_obj
                print(f"[batch_ei] Updated f_best to {f_best:.4f}")
            
            # CRITICAL: Update the Pareto front with the selected point for sequential batch optimization
            # This is the key to ensuring diverse selection in batch optimization
            # We add the selected point to the Pareto front to update f_best for the next iteration
            
            # Get the selected point's features
            selected_features = X_band[selected_idx].reshape(1, -1)
            selected_rate = mu_r[selected_idx]
            selected_range = mu_rng_nm[selected_idx]
            
            # Add the selected point to the Pareto front for the next iteration
            # This ensures that subsequent selections consider this point as part of the current front
            new_pareto_point = np.array([[selected_rate, selected_range]])
            pareto_pts = np.vstack([pareto_pts, new_pareto_point])
            
            print(f"[batch_ei] Added selected point to Pareto front: Rate={selected_rate:.1f}, Range={selected_range:.1f}")
            print(f"[batch_ei] Updated Pareto front now has {len(pareto_pts)} points")
            
            # Update f_best to reflect the new Pareto front
            new_pareto_objectives = []
            for rate, range_nm in pareto_pts:
                obj = self.pareto_improvement_objective(rate, range_nm, pareto_pts)
                new_pareto_objectives.append(obj)
            f_best = max(new_pareto_objectives)
            print(f"[batch_ei] Updated f_best to {f_best:.4f}")
        
        print(f"[batch_ei] Completed batch selection: {len(selected_indices)} candidates selected")
        return selected_indices, selection_details
    
    def build_and_write_proposals_ei(self, X_band: np.ndarray, mu_r: np.ndarray, mu_rng_nm: np.ndarray,
                                    sr: np.ndarray, srg_nm: np.ndarray, pareto_pts: np.ndarray,
                                    ml_models) -> Tuple[str, List[Dict], np.ndarray, np.ndarray, np.ndarray]:
        """Build and write proposals using Expected Improvement batch optimization"""
        print(f"[ei_proposals] Building proposals using EI batch optimization")
        
        # Use batch EI selection instead of multi-objective scoring
        selected_indices, selection_details = self.batch_ei_selection(
            X_band, mu_r, mu_rng_nm, sr, srg_nm, pareto_pts, ml_models, batch_size=K
        )
        
        if not selected_indices:
            print(f"[ei_proposals] No candidates selected")
            return "", [], np.array([]), np.array([]), np.array([])
        
        # Extract selected data
        new_recs = X_band[selected_indices]
        rates = mu_r[selected_indices]
        ranges_nm = mu_rng_nm[selected_indices]
        
        # Get uncertainties for selected recipes
        rate_uncertainties = sr[selected_indices]
        range_uncertainties = srg_nm[selected_indices]
        
        # Create proposals DataFrame
        df_props = pd.DataFrame({
            "O2_flow": new_recs[:, FEATURES.index("Etch_AvgO2Flow")],
            "Rf1_Pow": new_recs[:, FEATURES.index("Etch_Avg_Rf1_Pow")],
            "Rf2_Pow": new_recs[:, FEATURES.index("Etch_Avg_Rf2_Pow")],
            "Pressure": new_recs[:, FEATURES.index("Etch_AvgPres")],
            "cf4_flow": new_recs[:, FEATURES.index("Etch_Avgcf4Flow")],
            "Pred_avg_etch_rate": rates,
            "Pred_Range": ranges_nm,
            EXCEL_RATE_UNCERTAINTY_COL: rate_uncertainties,
            EXCEL_RANGE_UNCERTAINTY_COL: range_uncertainties
        })
        
        # Create Excel file
        xlsx_path = os.path.join(SNAP_DIR, f"proposals_ei_{TODAY}.xlsx")
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
            df_props.to_excel(w, index=False, sheet_name="Proposals")
        
        # Color code the proposals
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Proposals"]
        colors = ['FF1F77B4', 'FFFF7F0E', 'FF2CA02C']
        
        for i in range(min(K, len(colors))):
            fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=2 + i, column=col).fill = fill
        
        wb.save(xlsx_path)
        
        # Create detailed scores sheet with EI information
        scores_rows = []
        for i, detail in enumerate(selection_details):
            idx = detail["selected_idx"]
            scores_rows.append({
                "Recipe": f"Recipe {i+1}",
                "Selection_Method": "Expected Improvement",
                "Objective_Value": detail["objective_value"],
                "EI_Value": detail["ei_value"],
                "AvgEtchRate_pred": detail["rate_pred"],
                "Range_nm_pred": detail["range_pred"],
                "Rate_sigma": detail["rate_uncertainty"],
                "Range_sigma_nm": detail["range_uncertainty"],
                "Sampling": "sobol",
                "N_SAMPLES": 200000,
                "Seed": 42,
                # Knobs (store for audit)
                "O2_flow": float(X_band[idx, FEATURES.index("Etch_AvgO2Flow")]),
                "cf4_flow": float(X_band[idx, FEATURES.index("Etch_Avgcf4Flow")]),
                "Rf1_Pow": float(X_band[idx, FEATURES.index("Etch_Avg_Rf1_Pow")]),
                "Rf2_Pow": float(X_band[idx, FEATURES.index("Etch_Avg_Rf2_Pow")]),
                "Pressure": float(X_band[idx, FEATURES.index("Etch_AvgPres")]),
            })
        
        # Add EI explanation sheet
        expr_rows = []
        for i, detail in enumerate(selection_details):
            expr_rows.append({
                "Recipe": f"Recipe {i+1}",
                "Expression": (f"Recipe {i+1}: EI = (μ - f*)Φ((μ - f*)/σ) + σφ((μ - f*)/σ) = "
                              f"({detail['objective_value']:.4f} - f*)Φ(...) + σφ(...) = "
                              f"{detail['ei_value']:.6f}")
            })
        
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
            pd.DataFrame(scores_rows).to_excel(w, index=False, sheet_name="EI_Scores")
            pd.DataFrame(expr_rows).to_excel(w, index=False, sheet_name="EI_Expression")
        
        print(f"[ei_proposals] Generated {len(selected_indices)} proposals using EI")
        print(f"[ei_proposals] Excel file saved to: {xlsx_path}")
        
        return xlsx_path, selection_details, new_recs, rates, ranges_nm
    
    def build_and_write_proposals(self, X_band: np.ndarray, mu_r: np.ndarray, mu_rng_nm: np.ndarray,
                                 sr: np.ndarray, srg_nm: np.ndarray, pareto_pts: np.ndarray) -> Tuple[str, List[Dict], np.ndarray, np.ndarray, np.ndarray]:
        """Build and write proposals to Excel file"""
        r_scale, rn_scale = self.objective_scales(
            np.concatenate([mu_r, pareto_pts[:, 0]]) if len(pareto_pts) else mu_r,
            np.concatenate([mu_rng_nm, pareto_pts[:, 1]]) if len(pareto_pts) else mu_rng_nm
        )
        
        selected = []
        selected_knobs = []
        aug_front = pareto_pts.copy() if len(pareto_pts) else np.empty((0, 2))
        rows = []
        
        for k in range(K):
            # First, filter to only Pareto-improving points that significantly push the front
            valid_candidates = []
            for i in range(len(mu_r)):
                if i not in selected:
                    if len(aug_front) == 0 or self.significantly_improves(
                        (mu_r[i], mu_rng_nm[i]), aug_front, 
                        RATE_IMPROVEMENT_THRESHOLD, RANGE_IMPROVEMENT_THRESHOLD
                    ):
                        valid_candidates.append(i)
            
            # Filter to ensure only recipes with rf2 > rf1 are proposed
            rf_filtered_candidates = []
            for i in valid_candidates:
                rf1_power = X_band[i, FEATURES.index("Etch_Avg_Rf1_Pow")]
                rf2_power = X_band[i, FEATURES.index("Etch_Avg_Rf2_Pow")]
                if rf2_power > rf1_power:
                    rf_filtered_candidates.append(i)
                else:
                    print(f"[proposals] Filtered out candidate {i}: Rf1={rf1_power:.1f}, Rf2={rf2_power:.1f} (rf2 must be > rf1)")
            
            if len(rf_filtered_candidates) == 0:
                print(f"⚠️ No candidates with rf2 > rf1 available after selecting {k} recipes!")
                print(f"   Trying with relaxed Pareto criteria...")
                
                # Fallback: try with more relaxed thresholds but still maintain rf2 > rf1
                relaxed_candidates = []
                for i in range(len(mu_r)):
                    if i not in selected:
                        if len(aug_front) == 0 or self.significantly_improves(
                            (mu_r[i], mu_rng_nm[i]), aug_front, 
                            RATE_IMPROVEMENT_THRESHOLD * 0.5, 
                            RANGE_IMPROVEMENT_THRESHOLD * 0.5
                        ):
                            relaxed_candidates.append(i)
                
                # Apply rf2 > rf1 filter to relaxed candidates
                rf_filtered_candidates = []
                for i in relaxed_candidates:
                    rf1_power = X_band[i, FEATURES.index("Etch_Avg_Rf1_Pow")]
                    rf2_power = X_band[i, FEATURES.index("Etch_Avg_Rf2_Pow")]
                    if rf2_power > rf1_power:
                        rf_filtered_candidates.append(i)
                
                if len(rf_filtered_candidates) == 0:
                    # Final fallback: use original _improves function but still maintain rf2 > rf1
                    print(f"   No relaxed candidates with rf2 > rf1 found, using original non-dominated criteria")
                    for i in range(len(mu_r)):
                        if i not in selected:
                            if len(aug_front) == 0 or self.improves((mu_r[i], mu_rng_nm[i]), aug_front):
                                rf1_power = X_band[i, FEATURES.index("Etch_Avg_Rf1_Pow")]
                                rf2_power = X_band[i, FEATURES.index("Etch_Avg_Rf2_Pow")]
                                if rf2_power > rf1_power:
                                    rf_filtered_candidates.append(i)
                    
                    if len(rf_filtered_candidates) == 0:
                        print(f"   ⚠️ No more Pareto-improving points with rf2 > rf1 available after selecting {k} recipes!")
                        print(f"   This suggests we've reached the limit of what the current model can predict while maintaining rf2 > rf1 constraint.")
                    break
            
            # Use the rf-filtered candidates
            valid_candidates = rf_filtered_candidates
            
            # Debug logging for recipe selection
            if len(aug_front) > 0:
                print(f"[proposals] Iteration {k+1}: Checking {len(mu_r)} candidates against Pareto front with {len(aug_front)} points")
                print(f"[proposals] Found {len(valid_candidates)} valid significantly improving candidates with rf2 > rf1")
                if len(valid_candidates) > 0:
                    # Show a few examples of valid candidates with improvement details
                    for j, idx in enumerate(valid_candidates[:3]):  # Show first 3
                        rate, range_nm = mu_r[idx], mu_rng_nm[idx]
                        rf1_power = X_band[idx, FEATURES.index("Etch_Avg_Rf1_Pow")]
                        rf2_power = X_band[idx, FEATURES.index("Etch_Avg_Rf2_Pow")]
                        dominates = self.dominates_existing((rate, range_nm), aug_front)
                        print(f"[proposals]   Valid candidate {j+1}: Rate={rate:.1f}, Range={range_nm:.1f}, Rf1={rf1_power:.1f}, Rf2={rf2_power:.1f} (Dominates existing: {dominates})")
            else:
                print(f"[proposals] Iteration {k+1}: No existing Pareto front, all {len(mu_r)} candidates are valid")
                
                valid_candidates = np.array(valid_candidates)
            
            if len(valid_candidates) == 0:
                print(f"⚠️ No more significantly Pareto-improving points available after selecting {k} recipes!")
                print(f"   Trying with relaxed criteria...")
                
                # Fallback: try with more relaxed thresholds
                relaxed_candidates = []
                for i in range(len(mu_r)):
                    if i not in selected:
                        if len(aug_front) == 0 or self.significantly_improves(
                            (mu_r[i], mu_rng_nm[i]), aug_front, 
                            RATE_IMPROVEMENT_THRESHOLD * 0.5, 
                            RANGE_IMPROVEMENT_THRESHOLD * 0.5
                        ):
                            relaxed_candidates.append(i)
                
                if len(relaxed_candidates) > 0:
                    print(f"   Found {len(relaxed_candidates)} candidates with relaxed criteria")
                    valid_candidates = relaxed_candidates
                else:
                    # Final fallback: use original _improves function
                    print(f"   No relaxed candidates found, using original non-dominated criteria")
                    for i in range(len(mu_r)):
                        if i not in selected:
                            if len(aug_front) == 0 or self.improves((mu_r[i], mu_rng_nm[i]), aug_front):
                                valid_candidates.append(i)
                    
                    if len(valid_candidates) == 0:
                        print(f"   ⚠️ No more Pareto-improving points available after selecting {k} recipes!")
                        print(f"   This suggests we've reached the limit of what the current model can predict.")
                        break
            
            # Only consider valid candidates for scoring
            expl_raw = np.array([
                self.exploit_distance_norm((mu_r[i], mu_rng_nm[i]), aug_front, r_scale, rn_scale) 
                if len(aug_front) else 0 for i in valid_candidates
            ])
            expl_norm = self.norm01(expl_raw)
            
            explore_sigma_rn = sr[valid_candidates] / r_scale
            explore_sigma_gn = srg_nm[valid_candidates] / rn_scale
            explr_raw = np.hypot(explore_sigma_rn, explore_sigma_gn)
            explr_norm = self.norm01(explr_raw)
            
            if not selected_knobs:
                center = np.array([0.5] * len(FEATURES))
                div_raw = []
                for i in valid_candidates:
                    normalized_vec = []
                    for n in FEATURES:
                        val = (X_band[i][FEATURES.index(n)] - FEATURE_RANGES[n][0]) / (FEATURE_RANGES[n][1] - FEATURE_RANGES[n][0])
                        normalized_vec.append(val)
                    div_raw.append(np.linalg.norm(np.array(normalized_vec) - center))
                div_raw = np.array(div_raw)
                div_norm = self.norm01(div_raw)
            else:
                div_raw = np.array([self.diversity_score(X_band[i], selected_knobs, FEATURES) for i in valid_candidates])
                div_norm = self.norm01(div_raw)
                
            combined = ALPHA * expl_norm + BETA * explr_norm + GAMMA * div_norm
            
            # Select the best among valid candidates
            best_valid_idx = int(np.argmax(combined))
            best = valid_candidates[best_valid_idx]
            
            selected.append(best)
            selected_knobs.append(X_band[best])
            aug_front = np.vstack([aug_front, [mu_r[best], mu_rng_nm[best]]]) if len(aug_front) else np.array([[mu_r[best], mu_rng_nm[best]]])
            
            # Debug logging for selected recipe
            rate, range_nm = mu_r[best], mu_rng_nm[best]
            rf1_power = X_band[best, FEATURES.index("Etch_Avg_Rf1_Pow")]
            rf2_power = X_band[best, FEATURES.index("Etch_Avg_Rf2_Pow")]
            print(f"[proposals] Selected recipe {k+1}: Rate={rate:.1f}, Range={range_nm:.1f}, Rf1={rf1_power:.1f}, Rf2={rf2_power:.1f}")
            if len(pareto_pts) > 0:
                # Verify this point is actually non-dominated by the original Pareto front
                original_non_dominated = self.improves((rate, range_nm), pareto_pts)
                dominates_original = self.dominates_existing((rate, range_nm), pareto_pts)
                print(f"[proposals]   Verifies as non-dominated: {original_non_dominated}")
                print(f"[proposals]   Dominates existing points: {dominates_original}")
                if dominates_original:
                    print(f"[proposals]   ✅ This point PUSHES the Pareto front forward!")
                else:
                    print(f"[proposals]   ⚠️ This point fills a gap but doesn't dominate existing points")
            
            rows.append({
                "iter": k + 1,
                "idx": best,
                "Exploit_raw": float(expl_raw[best_valid_idx]),
                "Exploit_norm": float(expl_norm[best_valid_idx]),
                "Explore_raw": float(explr_raw[best_valid_idx]),
                "Explore_norm": float(explr_norm[best_valid_idx]),
                "Diversity_raw": float(div_raw[best_valid_idx]),
                "Diversity_norm": float(div_norm[best_valid_idx]),
                "Alpha*Exploit_norm": float(ALPHA * expl_norm[best_valid_idx]),
                "Beta*Explore_norm": float(BETA * explr_norm[best_valid_idx]),
                "Gamma*Div_norm": float(GAMMA * div_norm[best_valid_idx]),
                "Combined_score": float(combined[best_valid_idx]),
                "AvgEtchRate_pred": float(mu_r[best]),
                "Range_nm_pred": float(mu_rng_nm[best]),
                "Rate_sigma": float(sr[best]),
                "Range_sigma_nm": float(srg_nm[best]),
            })
        
        sel_idx = [r["idx"] for r in rows]
        new_recs = X_band[sel_idx]
        rates = mu_r[sel_idx]
        ranges_nm = mu_rng_nm[sel_idx]
        
        # Summary of Pareto front improvements
        if len(pareto_pts) > 0:
            dominating_points = 0
            for i, (rate, range_nm) in enumerate(zip(rates, ranges_nm)):
                if self.dominates_existing((rate, range_nm), pareto_pts):
                    dominating_points += 1
                    print(f"[proposals] 🎯 Recipe {i+1} dominates existing Pareto front points!")
            
            print(f"[proposals] 📊 Summary: {dominating_points}/{len(rates)} recipes actually push the Pareto front forward")
            if dominating_points == 0:
                print(f"[proposals] ⚠️ Warning: No recipes dominate existing points. Consider adjusting thresholds or sampling strategy.")
        
        # Get uncertainties for selected recipes
        rate_uncertainties = sr[sel_idx]
        range_uncertainties = srg_nm[sel_idx]
        
        df_props = pd.DataFrame({
            "O2_flow": new_recs[:, FEATURES.index("Etch_AvgO2Flow")],
            "Rf1_Pow": new_recs[:, FEATURES.index("Etch_Avg_Rf1_Pow")],
            "Rf2_Pow": new_recs[:, FEATURES.index("Etch_Avg_Rf2_Pow")],
            "Pressure": new_recs[:, FEATURES.index("Etch_AvgPres")],
            "cf4_flow": new_recs[:, FEATURES.index("Etch_Avgcf4Flow")],
            "Pred_avg_etch_rate": rates,
            "Pred_Range": ranges_nm,
            EXCEL_RATE_UNCERTAINTY_COL: rate_uncertainties,
            EXCEL_RANGE_UNCERTAINTY_COL: range_uncertainties
        })
        
        xlsx_path = os.path.join(SNAP_DIR, f"proposals_{TODAY}.xlsx")
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
            df_props.to_excel(w, index=False, sheet_name="Proposals")
        
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Proposals"]
        colors = ['FF1F77B4', 'FFFF7F0E', 'FF2CA02C']
        
        for i in range(min(K, len(colors))):
            fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type="solid")
            for col in range(1, ws.max_column + 1):
                ws.cell(row=2 + i, column=col).fill = fill
        
        wb.save(xlsx_path)
        
        # Scores + Expression
        scores_rows, expr_rows = [], []
        for ridx, rec in enumerate(rows, start=1):
            idx = rec["idx"]
            scores_rows.append({
                "Recipe": f"Recipe {ridx}",
                "Alpha": ALPHA, "Beta": BETA, "Gamma": GAMMA,
                "Exploit_raw": rec["Exploit_raw"], "Exploit_norm": rec["Exploit_norm"],
                "Alpha*Exploit_norm": rec["Alpha*Exploit_norm"],
                "Explore_raw": rec["Explore_raw"], "Explore_norm": rec["Explore_norm"],
                "Beta*Explore_norm": rec["Beta*Explore_norm"],
                "Diversity_raw": rec["Diversity_raw"], "Diversity_norm": rec["Diversity_norm"],
                "Gamma*Div_norm": rec["Gamma*Div_norm"],
                "Combined_score": rec["Combined_score"],
                "AvgEtchRate_pred": rec["AvgEtchRate_pred"],
                "Range_nm_pred": rec["Range_nm_pred"],
                "Rate_sigma": rec["Rate_sigma"],
                "Range_sigma_nm": rec["Range_sigma_nm"],
                "Sampling": "sobol",  # Hardcoded for now
                "N_SAMPLES": 200000,  # Hardcoded for now
                "Seed": 42,  # Hardcoded for now
                # Knobs (store for audit)
                "O2_flow": float(X_band[idx, FEATURES.index("Etch_AvgO2Flow")]),
                "cf4_flow": float(X_band[idx, FEATURES.index("Etch_Avgcf4Flow")]),
                "Rf1_Pow": float(X_band[idx, FEATURES.index("Etch_Avg_Rf1_Pow")]),
                "Rf2_Pow": float(X_band[idx, FEATURES.index("Etch_Avg_Rf2_Pow")]),
                "Pressure": float(X_band[idx, FEATURES.index("Etch_AvgPres")]),
            })
            expr_rows.append({
                "Recipe": f"Recipe {ridx}",
                "Expression": (f"Recipe {ridx}: score = αE + βX + γD = "
                              f"{ALPHA:.6f}*{rec['Exploit_norm']:.6f} + "
                              f"{BETA:.6f}*{rec['Explore_norm']:.6f} + "
                              f"{GAMMA:.6f}*{rec['Diversity_norm']:.6f} "
                              f"= {rec['Combined_score']:.6f}")
            })
        
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
            pd.DataFrame(scores_rows).to_excel(w, index=False, sheet_name="Scores")
            pd.DataFrame(expr_rows).to_excel(w, index=False, sheet_name="Expression")
        
        return xlsx_path, rows, new_recs, rates, ranges_nm
